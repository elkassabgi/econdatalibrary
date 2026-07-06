"""S1/S5 fetcher — IEP Ecological Threat Report (ETR).

CC BY-NC-SA 4.0 (Institute for Economics & Peace; granted to the Elkassabgi Data
Library for non-commercial re-hosting 2026-07-06). Single grouped parquet
clean_full/etr/etr.parquet, schema (series_key, obs_date, value),
series_key = "ETR:<indicator>:<country_name>" (annual, Dec-31).

Workbook layout (verified against ETR-2024-Public-Release-Data.xlsx): an
'Overview' sheet (licence text only) plus a 'Data' sheet holding a single
latest-year cross-section. The real header is a few rows down; columns are
Country | Overall Score | Food Insecurity | Impact of Natural Events |
Demographic Pressure | Water Risk. There is NO ISO3 column and NO year column in
the sheet, so the id is the country name and the reporting YEAR is taken from the
release URL (e.g. ETR-2024 -> 2024). Higher score = greater ecological threat.
"""
from __future__ import annotations
import io
import re
from datetime import date

from ._iep import probe_vintage, run_update

SOURCE = "etr"
URLS = [
    "https://www.economicsandpeace.org/wp-content/uploads/2025/08/ETR-2024-Public-Release-Data.xlsx",
]


def _report_year() -> int:
    """Reporting year of the current ETR edition, read from the release URL
    (the ETR workbook itself carries no year column). Falls back to 2024."""
    m = re.search(r"ETR[-_](20\d{2})", URLS[0])
    return int(m.group(1)) if m else 2024


def _slug(header):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(header).strip()).strip("_")


def parse(xlsx_bytes):
    """IEP ETR 'Data' sheet (single cross-section) -> (keys, dates, vals)."""
    import openpyxl
    obs = date(_report_year(), 12, 31)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    header_idx = country_col = None
    for i, row in enumerate(grid):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() == "country":
                header_idx, country_col = i, j
                break
        if header_idx is not None:
            break
    if header_idx is None:
        raise ValueError("ETR: could not find header row containing a 'Country' column")

    header = grid[header_idx]
    indicator_cols = [
        (j, str(header[j]).strip())
        for j in range(country_col + 1, len(header))
        if header[j] is not None and str(header[j]).strip() != ""
    ]

    keys, dates, vals = [], [], []
    for row in grid[header_idx + 1:]:
        if country_col >= len(row):
            continue
        country = row[country_col]
        if country is None or str(country).strip() == "":
            continue
        cid = str(country).strip()
        for j, hdr in indicator_cols:
            if j >= len(row):
                continue
            v = row[j]
            if v is None or isinstance(v, bool) or not isinstance(v, (int, float)):
                continue
            keys.append(f"ETR:{_slug(hdr)}:{cid}")
            dates.append(obs)
            vals.append(float(v))

    wb.close()
    return keys, dates, vals


def current_vintage(unit):
    return probe_vintage(URLS)


def update(unit, since):
    return run_update(SOURCE, URLS, parse)
