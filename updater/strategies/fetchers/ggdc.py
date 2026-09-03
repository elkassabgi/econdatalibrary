"""S1 fetcher — GGDC (Groningen Growth and Development Centre): PWT + Maddison.

Four complete static country-year panels, each re-estimated as a whole-panel release
between versions (no date delta):
  PWT 10.0          rug.nl/ggdc/docs/pwt100.xlsx                       -> pwt10.parquet
  PWT 11.0          dataverse.nl datafile 554105                      -> pwt11.parquet
  Maddison 2020     rug.nl .../maddison/data/mpd2020.xlsx             -> maddison2020.parquet
  Maddison 2023     dataverse.nl datafile 421302                      -> maddison2023.parquet

Each parquet is schema (series_key, obs_date, value); series_key is version-prefixed
(PWT10:/PWT11:/MADDISON:/MADDISON23:) so the four panels are disjoint by design. obs_date
is Dec-31 of the country-year. We MERGE per file (dedup series_key+obs_date, new wins on
revision, never-shrink) — an unchanged release merges to 0 new rows and reads no_change.

S1 (overwrite_if_changed): the WHOLE workbook is re-fetched + parsed by REUSING
jobs/ingest_ggdc.py's URLs + parse shape; we publish ONLY via merge.merge_and_write.
Cheap vintage: rug.nl HEAD Last-Modified/ETag for PWT10/Maddison2020; the Dataverse
/api/files/{id} SHA-1 checksum+filesize for the two pinned datafiles (Dataverse HEAD
403s, so we use the API). Composite token over all four; changes iff any component moved.

Per-component honesty: a 200 with a real (>10KB) body parsing 0 rows is structural; a
timeout/5xx/429/network failure is transient; a 403/404/tiny body is a missing component
(empty). If EVERY component fails to yield data, finalize's all-empty-window guard
escalates to a structural break.
"""
from __future__ import annotations
import datetime as dt
import importlib.util
import io
import os

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import http_vintage, UA

SOURCE = "ggdc"
DEDUP = ("series_key", "obs_date")
_TRANSIENT_STATUS = (429, 500, 502, 503, 504)

# Dataverse file-info API (HEAD 403s; this GET is cheap and exposes a SHA-1 checksum).
_DV_FILE_API = "https://dataverse.nl/api/files/{fid}"


def _ingest_mod():
    """Load jobs/ingest_ggdc.py by path to reuse its URLs + id-column set (the script is
    a standalone job, not an importable package). We reuse its constants and replicate the
    per-sheet parse shape here — we must NOT call its ingest_*()/save(), which write parquet
    directly (bypassing merge's never-shrink guard) and short-circuit when a file exists."""
    path = os.path.join(config.JOBS_DIR, "ingest_ggdc.py")
    spec = importlib.util.spec_from_file_location("_ingest_ggdc", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _dv_vintage(fid: str) -> str | None:
    """Cheap Dataverse vintage: SHA-1 checksum + filesize of the pinned datafile.
    Returns None on transient error (caller then fetches anyway — safe under merge)."""
    try:
        r = requests.get(_DV_FILE_API.format(fid=fid), headers=UA, timeout=60)
    except (requests.Timeout, requests.ConnectionError):
        return None
    if r.status_code != 200:
        return None
    try:
        df = (r.json().get("data") or {}).get("dataFile") or {}
    except ValueError:
        return None
    chk = df.get("checksum") or {}
    val = chk.get("value")
    size = df.get("filesize")
    if val:
        return f"{fid}:{val}:{size}"
    return None


def current_vintage(unit):
    """Composite cheap probe over all four components; changes iff any upstream moved.
    None only if NOTHING was determinable (strategy then fetches anyway, which is safe)."""
    ing = _ingest_mod()
    parts = []
    # rug.nl: HEAD Last-Modified/ETag works for these two.
    parts.append("pwt10=" + (http_vintage(ing.PWT_URL) or "?"))
    parts.append("mad20=" + (http_vintage(ing.MADDISON_URL) or "?"))
    # dataverse.nl: HEAD 403s, so use the file-info API SHA-1 checksum.
    parts.append("pwt11=" + (_dv_vintage("554105") or "?"))
    parts.append("mad23=" + (_dv_vintage("421302") or "?"))
    token = "|".join(parts)
    return None if token.count("?") == 4 else token


def _fetch(url):
    """Re-fetch one workbook. Returns (bytes|None, outcome) in {'ok','transient','notfound'}.
    Mirrors the ingester's accept rule (200 and >10KB)."""
    try:
        r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
    except (requests.Timeout, requests.ConnectionError):
        return None, "transient"
    if r.status_code in _TRANSIENT_STATUS:
        return None, "transient"
    if r.status_code == 200 and len(r.content) > 10_000:
        return r.content, "ok"
    return None, "notfound"


def _parse_panel(data: bytes, sheet_names, prefix: str, id_cols, min_year: int):
    """Parse a PWT/Maddison workbook into (keys, dates, vals) — the shared shape from
    jobs/ingest_ggdc.py: find the first matching sheet, treat row 0 as the header, take
    countrycode+year as ids and every other non-id column as a numeric variable, key as
    '<prefix>:<var>:<ccode>', obs_date = Dec-31 of the year (>= min_year)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], []
    header = rows[0]

    col_map = {}
    ctry_idx = year_idx = None
    for ci, col in enumerate(header):
        if col is None:
            continue
        col = str(col).strip()
        if col == "countrycode":
            ctry_idx = ci
        elif col == "year":
            year_idx = ci
        elif col not in id_cols:
            col_map[ci] = col
    if ctry_idx is None or year_idx is None or not col_map:
        return [], [], []

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None or row[year_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip()
        try:
            yr = int(row[year_idx])
            if yr < min_year:
                continue
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for ci, varname in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            try:
                v = float(cell)
            except (TypeError, ValueError):
                continue
            if v != v or v in (float("inf"), float("-inf")):
                continue
            keys.append(f"{prefix}:{varname}:{ctry}")
            dates.append(obs_d)
            vals.append(v)
    return keys, dates, vals


def _series_maxes(keys, dates, out):
    for k, d in zip(keys, dates):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return out


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    ing = _ingest_mod()
    tally = Tally()

    # (filename, url, sheet-name candidates, series prefix, id-col set, min-year)
    components = [
        ("pwt10.parquet",        ing.PWT_URL,      ["Data"],                          "PWT10",      ing.PWT_ID_COLS, 1),
        ("pwt11.parquet",        ing.PWT11_URL,    ["Data"],                          "PWT11",      ing.PWT_ID_COLS, 1),
        ("maddison2020.parquet", ing.MADDISON_URL, ["Full data"],                     "MADDISON",   {"countrycode", "country", "year"}, 1),
        ("maddison2023.parquet", ing.MAD23_URL,    ["Full data", "Data", "data", "full_data"], "MADDISON23", {"countrycode", "country", "year", "iso3", "code", "region"}, 1),
    ]

    total_rows = 0
    last_obs = None
    cursors: dict = {}

    for fname, url, sheets, prefix, id_cols, min_year in components:
        path = os.path.join(out_dir, fname)
        before = blob.row_count(path)

        data, outcome = _fetch(url)
        if outcome == "transient":
            tally.transient_unit(f"{fname}: fetch failed (transient)")
            total_rows += before
            continue
        if outcome == "notfound" or not data:
            # A pinned URL/datafile-id that 403/404s or returns a tiny body: this component
            # is missing/rotated. Count empty; if EVERY component is empty, finalize's
            # all-empty-window guard escalates to a structural break (don't fake success).
            tally.empty_unit(f"{fname}: pinned URL 403/404 or a tiny body — component rotated")
            total_rows += before
            continue

        try:
            keys, dates, vals = _parse_panel(data, sheets, prefix, id_cols, min_year)
        except Exception as e:  # noqa: BLE001
            # 200 with a real body that failed to parse -> schema break
            tally.structural_unit(f"{fname}: body will not parse — {type(e).__name__}")
            total_rows += before
            continue

        if not vals:
            # 200, >10KB body, but parsed 0 rows -> structural break
            tally.structural_unit(f"{fname}: real body parsed 0 rows over {before:,} stored")
            total_rows += before
            continue

        tbl = pa.table({
            "series_key": pa.array(keys, pa.string()),
            "obs_date":   pa.array(dates, pa.date32()),
            "value":      pa.array(vals, pa.float64()),
        })
        # Publish ONLY via merge (atomic, dedup, never-shrink). Honest new-row count is the
        # merge delta (n - before), not parsed rows — an unchanged release merges to 0 new.
        n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
        tally.added_unit(max(0, n - before))
        total_rows += n
        if md is not None and (last_obs is None or md > last_obs):
            last_obs = md
        _series_maxes(keys, dates, cursors)

    cursors_iso = {k: v.isoformat() for k, v in cursors.items()}
    return finalize(tally, total_rows, last_obs, source=SOURCE, series_cursors=cursors_iso)
