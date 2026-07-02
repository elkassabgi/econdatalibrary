"""S1 fetcher — SIPRI Military Expenditure + Polity5 + MEPV (annual bulk XLSX/XLS).

Three tiny annual bulk files, no delta API, re-published/re-estimated each release:
  SIPRI milex : www.sipri.org versioned dated workbook -> clean_full/sipri/milex.parquet
  Polity5     : www.systemicpeace.org p5v2018.xls      -> clean_full/polity/polity5.parquet
  MEPV        : www.systemicpeace.org MEPVv2018.xls     -> clean_full/polity/mepv.parquet

All three share schema (series_key, obs_date, value), dedup (series_key, obs_date).
We re-fetch each whole workbook and MERGE (new wins on revision, never-shrink), so a
revised vintage overwrites in place without dropping history. The three output dirs are
HARDCODED (sipri/ and polity/) and are shared with other sources — we overwrite per-file,
never clear a dir.

Vintage signal (registry adapter.vintage_signal): SIPRI's versioned dated filename
SIPRI-Milex-data-1949-<year>_v<X.Y>.xlsx (auto-discover the live year ahead of the
hardcoded list) + its Last-Modified; Polity/MEPV gate on Last-Modified (frozen 2018
vintages — poll for a newer file). Composite token over all three so any one moving
triggers a re-fetch.

Each workbook is one sub-unit: a 200 that parses 0 rows from a real body is structural;
a timeout/5xx/429/network drop is transient (existing data kept, retry next run).
"""
from __future__ import annotations
import datetime as dt
import io
import os

import pyarrow as pa
import requests

from ... import config, blob, merge
from ._common import Tally, finalize
from ._vintage import http_vintage, UA

SOURCE = "sipri_polity"
DEDUP = ("series_key", "obs_date")

# Output paths are HARDCODED to match the existing ingester (jobs/ingest_sipri_polity.py).
# These are NOT config.source_dir(SOURCE) — milex shares sipri/ with the separate `sipri`
# source, and polity files live under polity/.
SIPRI_OUT = os.path.join(config.DATA_ROOT, "sipri", "milex.parquet")
POLITY_OUT = os.path.join(config.DATA_ROOT, "polity", "polity5.parquet")
MEPV_OUT = os.path.join(config.DATA_ROOT, "polity", "mepv.parquet")

# SIPRI versioned dated workbook. Hardcoded current vintage; current_vintage() also
# auto-probes newer years so we notice (and the strategy re-fetches) on a new release.
SIPRI_URL = "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2025_v1.2.xlsx"
POLITY_URL = "https://www.systemicpeace.org/inscr/p5v2018.xls"
MEPV_URL = "https://www.systemicpeace.org/inscr/MEPVv2018.xls"

# SIPRI xlsx sheet name -> short label for series key (matches the ingester).
SIPRI_SHEETS = {
    "Current US$": "usd_curr",
    "Constant (2024) US$": "usd_const",
    "Share of GDP": "gdp_share",
    "Share of Govt. spending": "govt_share",
    "Per capita": "per_capita",
}

POLITY_SKIP = {
    "cyear", "ccode", "scode", "country", "year", "byear", "bmonth", "bday",
    "eyear", "emonth", "eday", "flag", "fragment", "prior", "emonth2", "eday2",
    "eyear2", "eseq", "post", "lead", "change", "d5", "sf", "regtrans", "p5",
}
MEPV_SKIP = {"cyear", "ccode", "scode", "country", "year", "version"}

_TRANSIENT = (429, 500, 502, 503, 504)


# ---------------------------------------------------------------------------
# vintage
# ---------------------------------------------------------------------------

def _discover_sipri_url():
    """Return (url, vintage_token) for the live SIPRI workbook. Auto-probes the
    current/next year ahead of the hardcoded 2025 file so we pick up a new annual
    release; falls back to the hardcoded URL. token = filename + Last-Modified/len."""
    candidates = [SIPRI_URL]
    yr_now = dt.date.today().year
    base = "https://www.sipri.org/sites/default/files/"
    # probe a couple years ahead, common version suffixes, plain + versioned
    for y in range(yr_now + 1, 2024, -1):
        for suf in ("", "_v1.0", "_v1.1", "_v1.2", "_v1.3", "_v2.0"):
            candidates.append(f"{base}SIPRI-Milex-data-1949-{y}{suf}.xlsx")
    seen = set()
    for url in candidates:
        if url in seen:
            continue
        seen.add(url)
        tok = http_vintage(url)  # HEAD; None on 404/transient/no-header
        if tok is not None:
            fname = url.rsplit("/", 1)[-1]
            return url, f"{fname}|{tok}"
    return SIPRI_URL, None


def current_vintage(unit):
    """Composite cheap probe: live SIPRI filename+Last-Modified plus Polity/MEPV
    Last-Modified. Changes iff any of the three upstream files moves. None only if
    nothing can be determined (strategy then fetches anyway — safe via never-shrink)."""
    parts = []
    _, sipri_tok = _discover_sipri_url()
    if sipri_tok:
        parts.append("sipri=" + sipri_tok)
    p_tok = http_vintage(POLITY_URL)
    if p_tok:
        parts.append("polity=" + p_tok)
    m_tok = http_vintage(MEPV_URL)
    if m_tok:
        parts.append("mepv=" + m_tok)
    return "; ".join(parts) if parts else None


# ---------------------------------------------------------------------------
# fetch + parse (reused from the ingester)
# ---------------------------------------------------------------------------

def _get(url, tally):
    """GET with the failure contract. Returns bytes on 200, or None after recording
    the right Tally outcome (transient for 5xx/429/network, structural for other 4xx)."""
    try:
        r = requests.get(url, headers=UA, timeout=120)
    except (requests.Timeout, requests.ConnectionError):
        tally.transient_unit()
        return None
    if r.status_code in _TRANSIENT:
        tally.transient_unit()
        return None
    if r.status_code != 200:
        tally.structural_unit()  # 404/410/403 on a previously-live bulk file = structural break
        return None
    return r.content


def _parse_sipri_sheet(ws, sheet_label):
    """Parse one SIPRI sheet (header row = years; col 0 = country)."""
    keys, dates, vals = [], [], []
    header_row = None
    year_cols = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        year_count = sum(1 for c in row if isinstance(c, (int, float)) and 1900 < c < 2100)
        if year_count > 10:
            header_row = r_idx
            for c_idx, cell in enumerate(row):
                if isinstance(cell, (int, float)) and 1900 < cell < 2100:
                    year_cols[c_idx] = int(cell)
            break
    if header_row is None:
        return [], [], []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row or not row or row[0] is None:
            continue
        country = str(row[0]).strip()
        if not country or country.lower() in ("country", "notes", "note", "source", "sources"):
            continue
        for c_idx, yr in year_cols.items():
            if c_idx >= len(row):
                continue
            cv = row[c_idx]
            if cv is None or cv == "" or str(cv).strip() in ("", "...", "xxx", "NA", "N/A"):
                continue
            try:
                v = float(cv)
                if v != v or v in (float("inf"), float("-inf")):
                    continue
                keys.append(f"SIPRI:milex:{sheet_label}:{country}")
                dates.append(dt.date(yr, 12, 31))
                vals.append(v)
            except (TypeError, ValueError):
                pass
    return keys, dates, vals


def _parse_sipri(data):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    keys, dates, vals = [], [], []
    for sheet_name, sheet_label in SIPRI_SHEETS.items():
        ws = None
        for s in wb.sheetnames:
            if sheet_name.lower() in s.lower() or s.lower() in sheet_name.lower():
                ws = wb[s]
                break
        if ws is None:
            continue
        k, d, v = _parse_sipri_sheet(ws, sheet_label)
        keys.extend(k); dates.extend(d); vals.extend(v)
    return keys, dates, vals


def _parse_csp(data, skip, prefix):
    """Parse a Center-for-Systemic-Peace XLS (Polity5 / MEPV) into long form."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheets()[0]
    headers = [str(ws.cell(0, c).value).strip().lower() for c in range(ws.ncols)]
    if "year" not in headers:
        return [], [], []
    year_i = headers.index("year")
    ctry_i = (headers.index("scode") if "scode" in headers else
              headers.index("ccode") if "ccode" in headers else
              headers.index("country") if "country" in headers else None)
    num_idx = [(h, i) for i, h in enumerate(headers) if h not in skip and i not in (year_i, ctry_i)]
    # Polity5 sentinels are special-cased; MEPV has none.
    sentinels = (-99.0, -88.0, -77.0, -66.0) if prefix == "POLITY" else ()
    keys, dates, vals = [], [], []
    for r in range(1, ws.nrows):
        try:
            yr = int(ws.cell(r, year_i).value)
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue
        ctry = str(ws.cell(r, ctry_i).value).strip() if ctry_i is not None else ""
        for col, ci in num_idx:
            cell = ws.cell(r, ci)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            if prefix == "POLITY" and str(cell.value).strip() in ("-99", "-88", "-77", "-66"):
                continue
            try:
                v = float(cell.value)
            except (TypeError, ValueError):
                continue
            if v in sentinels:
                continue
            key = f"{prefix}:{col}:{ctry}" if ctry else f"{prefix}:{col}"
            keys.append(key); dates.append(obs_d); vals.append(v)
    return keys, dates, vals


# ---------------------------------------------------------------------------
# update
# ---------------------------------------------------------------------------

def _to_table(keys, dates, vals):
    return pa.table({
        "series_key": pa.array(keys, pa.string()),
        "obs_date": pa.array(dates, pa.date32()),
        "value": pa.array(vals, pa.float64()),
    })


def _publish(path, keys, dates, vals, tally, before, cursors):
    """Merge one workbook's rows into its parquet, recording an honest Tally outcome.
    Adds {series_key: last_obs} into `cursors`. Returns (rows_now, last_obs)."""
    tbl = _to_table(keys, dates, vals)
    if tbl.num_rows == 0:
        tally.structural_unit()  # 200 parsed nothing from a real body -> schema break
        return before, None
    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))
    for k, d in zip(keys, dates):
        if k not in cursors or d.isoformat() > cursors[k]:
            cursors[k] = d.isoformat()
    return n, md


def update(unit, since):
    os.makedirs(os.path.dirname(SIPRI_OUT), exist_ok=True)
    os.makedirs(os.path.dirname(POLITY_OUT), exist_ok=True)
    tally = Tally()
    cursors = {}
    total_rows = 0
    last_obs = None

    # --- SIPRI milex (auto-discovered URL) ---
    sipri_url, _ = _discover_sipri_url()
    before = blob.row_count(SIPRI_OUT)
    data = _get(sipri_url, tally)
    if data is not None:
        try:
            k, d, v = _parse_sipri(data)
        except Exception:
            k, d, v = [], [], []
        n, md = _publish(SIPRI_OUT, k, d, v, tally, before, cursors)
        total_rows += n
        if md and (last_obs is None or md > last_obs):
            last_obs = md
    else:
        total_rows += before

    # --- Polity5 ---
    before = blob.row_count(POLITY_OUT)
    data = _get(POLITY_URL, tally)
    if data is not None:
        try:
            k, d, v = _parse_csp(data, POLITY_SKIP, "POLITY")
        except Exception:
            k, d, v = [], [], []
        n, md = _publish(POLITY_OUT, k, d, v, tally, before, cursors)
        total_rows += n
        if md and (last_obs is None or md > last_obs):
            last_obs = md
    else:
        total_rows += before

    # --- MEPV ---
    before = blob.row_count(MEPV_OUT)
    data = _get(MEPV_URL, tally)
    if data is not None:
        try:
            k, d, v = _parse_csp(data, MEPV_SKIP, "MEPV")
        except Exception:
            k, d, v = [], [], []
        n, md = _publish(MEPV_OUT, k, d, v, tally, before, cursors)
        total_rows += n
        if md and (last_obs is None or md > last_obs):
            last_obs = md
    else:
        total_rows += before

    return finalize(tally, total_rows, last_obs, source=SOURCE, series_cursors=cursors)
