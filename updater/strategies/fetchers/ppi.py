"""S1/S5 fetcher — IEP Positive Peace Index (PPI).

CC BY-NC-SA 4.0 (Institute for Economics & Peace; granted to the Elkassabgi Data
Library for non-commercial re-hosting 2026-07-06). Single grouped parquet
clean_full/ppi/ppi.parquet, schema (series_key, obs_date, value),
series_key = "PPI:<indicator>:<country_name>" (annual, Dec-31).

Workbook layout (verified against PPI-Public-Release-Data-2023.xlsx): one sheet
PER YEAR '2009'..'2022' plus a non-data 'Overview'. On each year sheet the real
header is the first row whose first non-empty cell is 'Country' (branding + a
Pillars/Indicators grouping row sit above). Columns are matched BY HEADER NAME,
not position (the 2018 sheet has a duplicate 'Country' column that shifts
everything right). 'Country' and 'Region' are metadata; every other header is a
numeric indicator (Overall Score, the 8 pillars, 24 indicators). There is NO ISO3
column in the file, so the series id is the country name. Lower score = stronger
positive peace.
"""
from __future__ import annotations
import io
import re
from datetime import date

from ._iep import probe_vintage, run_update

SOURCE = "ppi"
URLS = [
    "https://www.economicsandpeace.org/wp-content/uploads/2025/08/PPI-Public-Release-Data-2023.xlsx",
]


def _slug(text):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(text).strip()).strip("_")


def parse(xlsx_bytes):
    """IEP PPI per-year sheets -> (keys, dates, vals). id = country name."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    keys, dates, vals = [], [], []
    for sheet_name in wb.sheetnames:
        if not re.fullmatch(r"\d{4}", sheet_name):
            continue  # skip 'Overview' / any non-year sheet
        obs = date(int(sheet_name), 12, 31)
        rows = list(wb[sheet_name].iter_rows(values_only=True))

        header_idx = None
        for i, row in enumerate(rows):
            first = next((c for c in row if c is not None and str(c).strip() != ""), None)
            if first == "Country":
                header_idx = i
                break
        if header_idx is None:
            continue

        header = list(rows[header_idx])
        metadata = {"Country", "Region"}
        indicator_cols = [
            (i, h) for i, h in enumerate(header)
            if h is not None and str(h).strip() != "" and str(h).strip() not in metadata
        ]
        country_col = header.index("Country")

        for row in rows[header_idx + 1:]:
            country = row[country_col] if country_col < len(row) else None
            if country is None or str(country).strip() == "":
                continue
            country = str(country).strip()
            for col_idx, col_header in indicator_cols:
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None or isinstance(v, bool) or not isinstance(v, (int, float)):
                    continue
                keys.append(f"PPI:{_slug(col_header)}:{country}")
                dates.append(obs)
                vals.append(float(v))

    wb.close()
    return keys, dates, vals


def current_vintage(unit):
    return probe_vintage(URLS)


def update(unit, since):
    return run_update(SOURCE, URLS, parse)
