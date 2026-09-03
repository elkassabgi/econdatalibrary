"""S1 fetcher — Global Peace Index (GPI), annual, 163 countries, 23 indicators.

Institute for Economics and Peace (IEP). Single grouped parquet
clean_full/gpi/gpi.parquet, schema (series_key, obs_date, value),
series_key = "GPI:{indicator}:{iso3-or-country}". IEP re-publishes the whole
annual table (and revises prior years) each June, so this is a whole-table
overwrite_if_changed source: re-fetch the workbook/CSV and MERGE (dedup
series_key+obs_date, new wins on revision, never-shrink). Reuses the URL list +
parse logic from jobs/ingest_gpi.py.

BLOCKER (verified live 2026-06): every hardcoded GPI URL 404s. IEP moved its
structured data behind a licensing wall (economicsandpeace.org/consulting/
data-licensing/); the wp-content path now serves only PDFs. OWID no longer hosts
a global-peace-index grapher slug; the GitHub `datasets/global-peace-index`
mirror is gone; the Mendeley CC-BY mirror (DOI 10.17632/yjxnfkcv4h, 2008-2023)
is reachable in a browser but its public-API/file endpoints return 403 to server
clients (Cloudflare WAF). With no free programmatic full-table source, this
fetcher does NOT fake success: current_vintage returns None (no usable signal)
and update surfaces the wholesale-404 honestly via the Tally/finalize contract
(every URL 404 -> structural sub-units -> DefinitiveError; timeouts/5xx -> the
unit goes 'partial' and retries). If/when a working URL is restored, add it to
the front of GPI_URLS and the fetcher resumes with no other change.
"""
from __future__ import annotations
import datetime as dt
import io
import os
import re

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import UA

SOURCE = "gpi"
DEDUP = ("series_key", "obs_date")  # matches jobs/ingest_gpi.py output schema

# URL list (IEP report files change name each year, so several candidates are tried
# in order; first one that parses >0 rows wins).
#
# PRIMARY (2026-07-06): IEP granted non-commercial re-hosting under CC BY-NC-SA 4.0
# (Prof. Elkassabgi's data-licensing request; info@economicsandpeace.org). This is the
# current WORKING source — the old visionofhumanity.org/GitHub/OWID URLs below all 404
# now (IEP moved GPI onto economicsandpeace.org behind the licensing confirmation).
# When IEP publishes the next edition, prepend its ..._YYYY.xlsx URL here.
GPI_URLS = [
    # 2026 edition — IEP public-release (granted, CC BY-NC-SA 4.0)
    "https://www.economicsandpeace.org/wp-content/uploads/2026/06/GPI_Public_Release_2026.xlsx",
    # --- legacy candidates (currently 404; kept as ordered fallbacks) ---
    # 2024 edition
    "https://www.visionofhumanity.org/wp-content/uploads/2024/06/GPI-2024-web.xlsx",
    "https://www.visionofhumanity.org/wp-content/uploads/2024/07/GPI-2024-full-report-data.xlsx",
    "https://www.visionofhumanity.org/wp-content/uploads/2024/06/GPI-2024-download.xlsx",
    # 2023 edition
    "https://www.visionofhumanity.org/wp-content/uploads/2023/06/GPI-2023-web.xlsx",
    "https://www.visionofhumanity.org/wp-content/uploads/2023/06/GPI-2023-Results-Overall-Scores-and-Domains.xlsx",
    # 2022
    "https://www.visionofhumanity.org/wp-content/uploads/2022/06/GPI-2022-web.xlsx",
    # GitHub / OWID mirrors (CSV)
    "https://raw.githubusercontent.com/datasets/global-peace-index/master/data/global-peace-index.csv",
    "https://ourworldindata.org/grapher/global-peace-index.csv",
]

_TRANSIENT_HTTP = (429, 500, 502, 503, 504)


def current_vintage(unit):
    """Cheap probe: ETag/Last-Modified of the first URL that returns 200.

    A bare http_vintage() also returns a 404 page's Content-Length, which would be
    a junk "vintage" that never moves; so we require a 200 first (a real file) and
    only then read its validator headers. Every GPI_URL currently 404s, so this
    returns None — correct: there is genuinely no cheap signal. The strategy then
    fetches anyway (cadence-gated), and update() surfaces the dead source honestly.
    """
    for url in GPI_URLS:
        try:
            r = requests.head(url, headers=UA, timeout=60, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError):
            continue
        if r.status_code != 200:
            continue
        h = r.headers
        v = h.get("ETag") or h.get("Last-Modified") or h.get("Content-Length")
        if v:
            return v
    return None


def _parse_owid_csv(data: bytes):
    """OWID/GitHub CSV: columns Entity, Code, Year, <value cols...>."""
    import csv
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() for h in (reader.fieldnames or [])]

    entity_col = next((h for h in headers if h.lower() in ("entity", "country", "name")), None)
    code_col = next((h for h in headers if h.lower() in ("code", "iso3", "iso")), None)
    year_col = next((h for h in headers if h.lower() in ("year",)), None)
    id_col = code_col or entity_col
    if not id_col or not year_col:
        return [], [], []

    val_cols = [h for h in headers if h not in (entity_col, code_col, year_col) and h]
    keys, dates, vals = [], [], []
    for rec in reader:
        cid = (rec.get(id_col) or "").strip()
        if not cid:
            continue
        try:
            yr = int(float((rec.get(year_col) or "").strip()))
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for col in val_cols:
            raw = rec.get(col, "")
            if not raw or str(raw).strip() in ("", "NA"):
                continue
            try:
                v = float(str(raw).strip())
            except (TypeError, ValueError):
                continue
            if v != v:  # NaN
                continue
            safe = re.sub(r"[^a-zA-Z0-9_]", "_", col)[:30]
            keys.append(f"GPI:{safe}:{cid}")
            dates.append(obs_d)
            vals.append(v)
    return keys, dates, vals


def _slug(header) -> str:
    """Sanitize an indicator column header to [A-Za-z0-9_]."""
    return re.sub(r"[^A-Za-z0-9]+", "_", str(header).strip()).strip("_")


def _parse_gpi_xlsx(data: bytes):
    """IEP GPI public-release workbook -> (keys, dates, vals).

    series_key = "GPI:<indicator_slug>:<ISO3>", obs_date = date(year, 12, 31).

    Layout (verified against GPI_Public_Release_2026.xlsx, 2026-07-06): a
    'Contents' sheet, an 'Overall Scores' sheet, and one sheet PER YEAR
    (2008..latest). Every per-year sheet carries the country name, an ISO3 code
    (column 'geocode'), and every numeric indicator (Overall Score, Rank, and the
    sub-indicators). We read ALL year sheets (the old parser stopped after the
    first). The 'Overall Scores' sheet is intentionally skipped: its <year>_Score/
    <year>_Rank columns are byte-identical to the per-year Overall Score/Rank, so
    re-reading it would duplicate (series_key, obs_date). Gaps (e.g. South Sudan
    pre-2012) are skipped, never fabricated.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    keys, dates, vals = [], [], []

    # Per-year data sheets are those whose name is exactly a 4-digit year.
    year_sheets = [s for s in wb.sheetnames if re.fullmatch(r"\d{4}", s)]

    for sheet_name in year_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Real header row = first row containing both 'country' and 'geocode'
        # (rows above are blank/branding).
        header_idx = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "country" in cells and "geocode" in cells:
                header_idx = i
                break
        if header_idx is None:
            continue

        header = list(rows[header_idx])
        col_country = col_geocode = None
        indicator_cols = []  # (col_index, slug)
        for ci, h in enumerate(header):
            if h is None:
                continue
            hl = str(h).strip().lower()
            if hl == "country":
                col_country = ci
            elif hl == "geocode":
                col_geocode = ci
            elif hl != "year":
                indicator_cols.append((ci, _slug(h)))
        if col_country is None or col_geocode is None:
            continue

        obs_d = dt.date(int(sheet_name), 12, 31)
        for row in rows[header_idx + 1:]:
            if col_country >= len(row) or col_geocode >= len(row):
                continue
            if row[col_country] is None or row[col_geocode] is None:
                continue
            iso3 = str(row[col_geocode]).strip()
            if not re.fullmatch(r"[A-Za-z]{3}", iso3):
                continue  # not a valid ISO3 -> not a country data row
            iso3 = iso3.upper()
            for ci, slug in indicator_cols:
                if ci >= len(row):
                    continue
                v = row[ci]
                if v is None or isinstance(v, bool) or not isinstance(v, (int, float)):
                    continue
                keys.append(f"GPI:{slug}:{iso3}")
                dates.append(obs_d)
                vals.append(float(v))

    wb.close()
    return keys, dates, vals


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "gpi.parquet")
    before = blob.row_count(path)
    tally = Tally()

    keys, dates, vals = [], [], []
    got_data = False
    for url in GPI_URLS:
        try:
            r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError) as e:
            # network/timeout — retry next run, never "no data"
            tally.transient_unit(f"{url[-52:]}: {type(e).__name__}")
            continue

        if r.status_code in _TRANSIENT_HTTP:
            tally.transient_unit(f"{url[-52:]}: HTTP {r.status_code}")
            continue
        if r.status_code != 200 or len(r.content) < 1000:
            # 404 / hard-4xx / empty body for this candidate URL -> structural for this sub-unit.
            tally.structural_unit(
                f"{url[-52:]}: HTTP {r.status_code}, {len(r.content):,} bytes")
            continue

        try:
            if url.endswith(".csv"):
                k, d, v = _parse_owid_csv(r.content)
            else:
                k, d, v = _parse_gpi_xlsx(r.content)
        except Exception as e:  # noqa: BLE001
            # 200 with a body we couldn't parse -> schema break
            tally.structural_unit(f"{url[-52:]}: unparseable body — {type(e).__name__}")
            continue

        if v:
            keys, dates, vals = k, d, v
            got_data = True
            break
        # 200, real body, parsed 0 rows -> structural
        tally.structural_unit(f"{url[-52:]}: real body parsed 0 rows")

    if not got_data:
        # No URL yielded data. finalize() is honest: any structural sub-unit raises
        # DefinitiveError (the source is broken, not quiet); pure transients -> 'partial'.
        return finalize(tally, before, None, source=SOURCE)

    tbl = pa.table({"series_key": pa.array(keys, pa.string()),
                    "obs_date": pa.array(dates, pa.date32()),
                    "value": pa.array(vals, pa.float64())})

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))  # the one successful workbook sub-unit
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
