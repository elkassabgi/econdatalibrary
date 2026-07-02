"""S1 fetcher — KOF Globalisation Index (ETH Zurich), annual, 1970-present.

Free for academic use. Single grouped parquet clean_full/kof_globalization/
kof_globalization.parquet, schema (series_key, obs_date, value). KOF re-estimates
and re-publishes the WHOLE panel each yearly edition (history is revised), so we
re-fetch the full XLSX and MERGE (dedup series_key+obs_date, new wins on revision,
never-shrink).

Vintage signal (registry): "Probe for a new year-stamped edition: HEAD/GET the
ethz.ch dam URL for the next year (e.g. KOFGI_2026_public.xlsx) ... edition = new
URL." So current_vintage() picks the newest year-stamped edition that returns 200
and tokenises it as "<year>:<Last-Modified|Content-Length>" — it moves when a new
edition appears OR the current edition file is re-published. Parse logic is reused
verbatim from jobs/ingest_kof_globalization.py (first sheet yielding obs wins,
KOF:{SAFE_COL}:{ISO} keys, Dec-31 obs) so keys dedup-match the published file.
"""
from __future__ import annotations
import datetime as dt
import io
import os

import pyarrow as pa
import requests

from ... import config, blob, merge
from ._common import Tally, finalize
from ._vintage import UA

SOURCE = "kof_globalization"
DEDUP = ("series_key", "obs_date")

# KOF Zurich provides a direct download of the full dataset per yearly edition.
# Hardcoded URLs reused from jobs/ingest_kof_globalization.py (year-stamped dam paths,
# first >5000 bytes wins). Plus the version-stable time-series API as a final fallback.
KOF_URLS = [
    "https://ethz.ch/content/dam/ethz/special-interest/dual/kof-dam/documents/Globalization/2025/KOFGI_2025_public.xlsx",
    "https://datenservice.kof.ethz.ch/api/v1/public/collections/globidx_v2020?mime=xlsx",
    "https://ethz.ch/content/dam/ethz/special-interest/dual/kof-dam/documents/Globalization/2024/KOFGI_2024_public.xlsx",
]

# Years to probe for a new edition, newest first (e.g. a 2026 release supersedes 2025).
# This MUST stay aligned with KOF_URLS' newest year; extend forward each cycle.
_EDITION_YEARS = (2027, 2026, 2025, 2024)
_DAM_TMPL = ("https://ethz.ch/content/dam/ethz/special-interest/dual/kof-dam/"
             "documents/Globalization/{y}/KOFGI_{y}_public.xlsx")


def _edition_url(year: int) -> str:
    return _DAM_TMPL.format(y=year)


def current_vintage(unit):
    """Cheap probe: HEAD the year-stamped dam URLs newest-first; the first 200 is the
    live edition. Token = "<year>:<Last-Modified or Content-Length>" so it moves on a
    new edition OR a re-publish of the current one. Returns None if none answer (the
    strategy then fetches anyway, which is safe — merge dedups + never-shrinks)."""
    for y in _EDITION_YEARS:
        u = _edition_url(y)
        try:
            r = requests.head(u, headers=UA, timeout=60, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError):
            continue
        if r.status_code == 200:
            tag = r.headers.get("ETag") or r.headers.get("Last-Modified") \
                or r.headers.get("Content-Length") or "200"
            return f"{y}:{tag}"
    return None


def _fetch(url: str):
    """Returns (status_kind, bytes|None). status_kind in {'ok','transient','gone'}."""
    try:
        r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
    except (requests.Timeout, requests.ConnectionError):
        return "transient", None
    if r.status_code in (429, 500, 502, 503, 504):
        return "transient", None
    if r.status_code == 200 and len(r.content) > 5000:
        return "ok", r.content
    if r.status_code in (403, 404):
        return "gone", None
    return "gone", None


def _parse_kof_xlsx(data: bytes):
    """Parse KOF XLSX exactly as jobs/ingest_kof_globalization.py does (first sheet
    yielding obs wins; KOF:{SAFE_COL}:{ISO} keys; year -> Dec-31 obs)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    keys, dates, vals = [], [], []

    for sheet_name in wb.sheetnames:
        if any(s in sheet_name.lower() for s in ["readme", "note", "info", "about", "source"]):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        header_idx = None
        for ri, row in enumerate(rows[:10]):
            row_lower = [str(c).strip().lower() if c else "" for c in row]
            if any(v in row_lower for v in ["code", "country", "iso", "year"]):
                header_idx = ri
                break
        if header_idx is None:
            continue

        header = [str(c).strip().lower() if c else "" for c in rows[header_idx]]
        iso_ci = next((i for i, h in enumerate(header) if h in ("code", "iso3", "iso", "iso3c", "country_code")), None)
        ctry_ci = next((i for i, h in enumerate(header) if h in ("country", "name", "country_name")), None)
        year_ci = next((i for i, h in enumerate(header) if h in ("year",)), None)

        id_ci = iso_ci if iso_ci is not None else ctry_ci
        if id_ci is None or year_ci is None:
            continue

        skip_ci = {iso_ci, ctry_ci, year_ci, None}
        val_col_names = []
        for i, h in enumerate(header):
            if i in skip_ci or not h:
                continue
            val_col_names.append((i, h))
        if not val_col_names:
            continue

        n_before = len(vals)
        for row in rows[header_idx + 1:]:
            if not row or row[id_ci] is None or row[year_ci] is None:
                continue
            try:
                yr = int(float(str(row[year_ci]).strip()))
                obs_d = dt.date(yr, 12, 31)
            except (TypeError, ValueError):
                continue

            id_val = str(row[id_ci]).strip()
            if not id_val or id_val.lower() in ("nan", "none", ""):
                continue

            for col_i, col_name in val_col_names:
                if col_i >= len(row) or row[col_i] is None:
                    continue
                try:
                    v = float(row[col_i])
                    if v != v:
                        continue
                    safe_col = col_name.replace(" ", "_")[:20].upper()
                    keys.append(f"KOF:{safe_col}:{id_val.upper()}")
                    dates.append(obs_d)
                    vals.append(v)
                except (TypeError, ValueError):
                    pass

        if len(vals) - n_before > 0:
            break  # use first successful sheet (matches ingester)

    return keys, dates, vals


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "kof_globalization.parquet")
    before = blob.row_count(path)
    tally = Tally()

    data = None
    any_transient = False
    for url in KOF_URLS:
        kind, body = _fetch(url)
        if kind == "ok":
            data = body
            break
        if kind == "transient":
            any_transient = True
        # 'gone' (403/404) -> try next URL (edition URL-rot is expected between cycles)

    if data is None:
        # No edition fetched. If we saw a transient, retry next tick; otherwise every
        # candidate URL is gone (URL-rot / new edition path) — that's a real blocker,
        # surface it (do NOT fake success) so a human resolves the new-year URL.
        if any_transient:
            tally.transient_unit()
            return finalize(tally, before, None, source=SOURCE)
        tally.structural_unit()  # all URLs 403/404 -> raises DefinitiveError in finalize
        return finalize(tally, before, None, source=SOURCE)

    keys, dates, vals = _parse_kof_xlsx(data)
    tbl = pa.table({
        "series_key": pa.array(keys, pa.string()),
        "obs_date": pa.array(dates, pa.date32()),
        "value": pa.array(vals, pa.float64()),
    })
    if tbl.num_rows == 0:
        tally.structural_unit()  # 200/real body but parsed nothing -> schema break
        return finalize(tally, before, None, source=SOURCE)

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
