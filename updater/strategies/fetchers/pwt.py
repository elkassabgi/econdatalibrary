"""S1 fetcher — Penn World Table 10.0 (annual, ~183 countries, 47 variables).

CC BY 4.0 (Groningen Growth and Development Centre). Single grouped parquet
clean_full/pwt/pwt.parquet, schema (series_key, obs_date, value). series_key is
'<variable>:<ISO3>' (colon separator — distinct from penn_world_table's '|'),
obs_date is annual Dec-31. PWT 10.0 is a FROZEN vintage (workbook last modified
2021-06-18), so the vintage probe will essentially never move; we re-fetch the
whole workbook on a vintage change and MERGE (dedup series_key+obs_date, new wins
on revision, never-shrink). One sub-unit (the workbook); a 200 that parses 0 rows
from a real body is structural.

Reuses jobs/ingest_pwt.py's URL + parse logic. Vintage signal per registry:
HTTP Last-Modified / ETag on the rug.nl workbook (HEAD before the cheap rebuild).
"""
from __future__ import annotations
import datetime as dt
import io
import os

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import http_vintage, UA

# Same URL the ingester uses; HEAD exposes ETag + Last-Modified + Content-Length.
URL = "https://www.rug.nl/ggdc/docs/pwt100.xlsx"
SOURCE = "pwt"
DEDUP = ("series_key", "obs_date")


def current_vintage(unit):
    return http_vintage(URL)


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "pwt.parquet")
    before = blob.row_count(path)
    tally = Tally()

    try:
        r = requests.get(URL, headers=UA, timeout=300)
    except (requests.Timeout, requests.ConnectionError):
        tally.transient_unit()
        return finalize(tally, before, None, source=SOURCE)
    if r.status_code in (429, 500, 502, 503, 504):
        tally.transient_unit()
        return finalize(tally, before, None, source=SOURCE)
    if r.status_code != 200:
        tally.structural_unit()
        return finalize(tally, before, None, source=SOURCE)

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)

    # Find the data sheet (usually "Data") — same selection as the ingester.
    data_sheet = None
    for name in wb.sheetnames:
        if name.lower() in ("data", "pwt100", "pwt10"):
            data_sheet = wb[name]
            break
    if data_sheet is None:
        data_sheet = wb.active

    rows = list(data_sheet.iter_rows(values_only=True))
    if not rows:
        tally.structural_unit()  # 200 but empty sheet -> schema/structural break
        return finalize(tally, before, None, source=SOURCE)

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    country_col = next((i for i, h in enumerate(headers)
                        if h.lower() in ("countrycode", "country_code", "iso", "isocode")), None)
    year_col = next((i for i, h in enumerate(headers)
                     if h.lower() in ("year", "yr")), None)
    country_name_col = next((i for i, h in enumerate(headers)
                             if h.lower() in ("country", "countryname", "country_name")), None)

    if year_col is None:
        tally.structural_unit()  # header present but no year col -> structural break
        return finalize(tally, before, None, source=SOURCE)

    skip_cols = {country_col, year_col, country_name_col}
    val_cols = [(i, headers[i]) for i in range(len(headers))
                if i not in skip_cols and headers[i] and headers[i] not in ("", "None")]

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        yr_raw = row[year_col]
        if yr_raw is None:
            continue
        try:
            yr = int(yr_raw)
            d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue

        iso = (str(row[country_col]).strip() if country_col is not None
               and row[country_col] is not None else "XXX")

        for col_idx, col_name in val_cols:
            v_raw = row[col_idx]
            if v_raw is None:
                continue
            try:
                v = float(v_raw)
            except (ValueError, TypeError):
                continue
            if v != v:  # NaN
                continue
            keys.append(f"{col_name}:{iso}")
            dates.append(d)
            vals.append(v)

    tbl = pa.table({"series_key": pa.array(keys, pa.string()),
                    "obs_date": pa.array(dates, pa.date32()),
                    "value": pa.array(vals, pa.float64())})
    if tbl.num_rows == 0:
        tally.structural_unit()  # 200 + real workbook but parsed nothing -> structural
        return finalize(tally, before, None, source=SOURCE)

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
