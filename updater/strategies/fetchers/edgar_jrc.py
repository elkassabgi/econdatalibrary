"""S1 fetcher — EU JRC EDGAR greenhouse-gas & air-pollutant country totals.

JRC publishes whole-history country-total workbooks (1970..latest) per annual
release; there is no `since` param. Single grouped parquet
clean_full/edgar_jrc/edgar_jrc.parquet, schema (series_key, obs_date, value),
series_key = EDGAR:{gas}:{iso3}, annual (Dec-31). EDGAR can revise history per
release, so we re-fetch the whole set and MERGE (dedup series_key+obs_date, new
wins on revision, never-shrink).

Vintage: composite HEAD ETag/Last-Modified of one representative file from each
of the two release bases (GHG_2025 + v81 AP). A re-publication of the current
vintage moves the token; a brand-new annual filename (e.g. 2026 release) is NOT
auto-detectable and needs a human URL bump in jobs/ingest_edgar_jrc.py
(hardcoded-URL, not a hard block).

Failure honesty: every dataset file is a sub-unit. 200-but-parsed-0-from-a-real
ZIP -> structural; timeout/5xx/429/network -> transient; a clean 404/403 file ->
empty (the source treats those as fatal-skip). The Tally/finalize turns these
into an honest ok/no_change/partial, and a large all-empty window raises
DefinitiveError instead of laundering a frozen source into 'fresh'.
"""
from __future__ import annotations
import datetime as dt
import importlib.util
import io
import os
import zipfile

import pyarrow as pa
import requests

from ... import config, blob, merge
from ...errors import DefinitiveError
from ..base import Result
from ._common import Tally, finalize
from ._vintage import UA

SOURCE = "edgar_jrc"
DEDUP = ("series_key", "obs_date")

# Load the existing ingester to reuse its URL list + parse logic verbatim (single
# source of truth for the hardcoded release URLs/filenames and the XLSX/CSV parser).
_JOB_PATH = os.path.join(config.JOBS_DIR, "ingest_edgar_jrc.py")
_spec = importlib.util.spec_from_file_location("_ingest_edgar_jrc", _JOB_PATH)
_job = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_job)

DATASETS = _job.DATASETS                # [(gas_label, filename, base_url), ...]
GHG_BASE = _job.GHG_BASE
AP_BASE = _job.AP_BASE
parse_edgar_xlsx = _job.parse_edgar_xlsx
parse_edgar_csv = _job.parse_edgar_csv

# One representative file per release base — their HEAD tokens move when that base
# is re-published. (Cheap: 2 HEAD requests instead of fetching 17 zips.)
_VINTAGE_PROBES = [
    GHG_BASE + "IEA_EDGAR_CO2_1970_2024.zip",
    AP_BASE + "EDGAR_SO2_1970_2022_v2.zip",
]


def current_vintage(unit) -> str | None:
    """Composite HEAD ETag/Last-Modified of one file per release base. Returns None
    only if BOTH probes fail to expose any signal (strategy then fetches anyway,
    which is safe under merge's dedup + never-shrink)."""
    parts = []
    for url in _VINTAGE_PROBES:
        tok = None
        for _ in range(2):
            try:
                r = requests.head(url, headers=UA, timeout=60, allow_redirects=True)
            except (requests.Timeout, requests.ConnectionError):
                continue
            if r.status_code in (429, 500, 502, 503, 504):
                continue
            h = r.headers
            tok = h.get("ETag") or h.get("Last-Modified") or h.get("Content-Length")
            break
        if tok:
            parts.append(tok)
    return "|".join(parts) if parts else None


def _fetch(url: str):
    """Return (bytes, kind) where kind in {'ok','transient','missing'}.
    'missing' = a clean 4xx (treated as fatal-skip per file by the source); a
    short/empty 200 body counts as transient (incomplete download)."""
    try:
        r = requests.get(url, headers=UA, timeout=300, allow_redirects=True)
    except (requests.Timeout, requests.ConnectionError):
        return None, "transient"
    if r.status_code == 200:
        if len(r.content) > 1000:
            return r.content, "ok"
        return None, "transient"
    if r.status_code in (429, 500, 502, 503, 504):
        return None, "transient"
    if r.status_code in (403, 404):
        return None, "missing"
    return None, "transient"


def _parse_totals_sheet(xlsx_bytes: bytes, gas: str):
    """Parse the EDGAR 'TOTALS BY COUNTRY' sheet -> ONE country total per (iso3, year).

    NOTE: the existing jobs/ingest_edgar_jrc.py parser (parse_edgar_xlsx) breaks on
    the FIRST sheet with year headers, which is the SECTORAL 'IPCC 2006' sheet — so
    it emits ~22-48 sector rows per (country, year) all under one country-level
    series_key (EDGAR:{gas}:{iso3}). Those collapse under the declared S1 dedup keys
    (series_key, obs_date). The semantically correct country total lives in the
    'TOTALS BY COUNTRY' sheet (one row per country), which is what this fetcher uses.
    Falls back to the ingester's generic parser only if the totals sheet is absent.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    tname = next((n for n in wb.sheetnames if "TOTALS BY COUNTRY" in n.upper()), None)
    if tname is None:
        return None  # signal caller to fall back
    ws = wb[tname]
    rows = list(ws.iter_rows(values_only=True))

    hri = hdr = None
    for ri, row in enumerate(rows[:15]):
        s = [str(c).strip() if c is not None else "" for c in row]
        yc = sum(1 for x in s if x.startswith("Y_") and x[2:].isdigit() and len(x[2:]) == 4)
        if yc >= 5:
            hri, hdr = ri, s
            break
    if hdr is None:
        return [], [], []

    c3 = None
    for cand in ("country_code_a3", "iso3", "country_code", "code"):
        for i, h in enumerate(hdr):
            if h.lower().strip() == cand:
                c3 = i
                break
        if c3 is not None:
            break
    if c3 is None:
        return [], [], []

    ycols = [(i, int(h[2:])) for i, h in enumerate(hdr)
             if h.startswith("Y_") and h[2:].isdigit() and len(h[2:]) == 4]

    keys, dates, vals = [], [], []
    for row in rows[hri + 1:]:
        if not row or c3 >= len(row) or row[c3] is None:
            continue
        cc = str(row[c3]).strip()
        if len(cc) != 3 or not cc.isalpha():
            continue
        for ci, yr in ycols:
            if ci >= len(row) or row[ci] is None:
                continue
            try:
                v = float(row[ci])
            except (TypeError, ValueError):
                continue
            if v != v:  # NaN
                continue
            keys.append(f"EDGAR:{gas}:{cc.upper()}")
            dates.append(dt.date(yr, 12, 31))
            vals.append(v)
    return keys, dates, vals


def _parse_zip(data: bytes, gas: str):
    """Country-total parse: 'TOTALS BY COUNTRY' XLSX sheet, else ingester fallback."""
    if data[:2] != b"PK":
        return [], [], []
    z = zipfile.ZipFile(io.BytesIO(data))
    members = z.namelist()
    xlsx_members = [m for m in members if m.lower().endswith(".xlsx") and
                    not any(s in m.lower() for s in ["readme", "info", "legend"])]
    csv_members = [m for m in members if m.lower().endswith(".csv") and
                   not any(s in m.lower() for s in ["readme", "info", "legend"])]
    keys, dates, vals = [], [], []
    for member in xlsx_members:
        res = _parse_totals_sheet(z.read(member), gas)
        if res is None:                       # no TOTALS sheet -> ingester generic parser
            k, d, v = parse_edgar_xlsx(z.read(member), gas)
        else:
            k, d, v = res
        keys.extend(k); dates.extend(d); vals.extend(v)
        if vals:
            break
    if not vals:
        for member in csv_members:
            k, d, v = parse_edgar_csv(z.read(member), gas)
            keys.extend(k); dates.extend(d); vals.extend(v)
            if vals:
                break
    return keys, dates, vals


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "edgar_jrc.parquet")
    before = blob.row_count(path)
    tally = Tally()

    all_keys, all_dates, all_vals = [], [], []
    for gas, filename, base_url in DATASETS:
        data, kind = _fetch(base_url + filename)
        if kind == "transient":
            tally.transient_unit()
            continue
        if kind == "missing":
            tally.empty_unit()       # clean 4xx — source treats as fatal-skip per file
            continue
        k, d, v = _parse_zip(data, gas)
        if not v:
            tally.structural_unit()  # 200 + real ZIP body but parsed 0 rows -> schema break
            continue
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
        tally.added_unit(len(v))     # rows parsed from this gas (added counted post-merge below)

    # No good data parsed from any file -> let finalize decide (transient -> partial;
    # all-missing over a large window -> DefinitiveError; structural -> DefinitiveError).
    if not all_vals:
        return finalize(tally, before, None, source=SOURCE)

    tbl = pa.table({
        "series_key": pa.array(all_keys, pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals, pa.float64()),
    })

    try:
        n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    except DefinitiveError as e:
        # The never-shrink guard fired. The EXISTING parquet stores ~3.36M SECTORAL
        # rows all collapsed under country-level series_key (EDGAR:{gas}:{iso3}) — i.e.
        # 190k+ duplicate (series_key, obs_date) pairs (up to 48/key). Under the
        # declared S1 dedup keys those dedup to the 195,391 true country totals this
        # fetcher (correctly) parses from 'TOTALS BY COUNTRY'. So merge refuses the
        # apparent 3.36M->195k shrink. Upstream is HEALTHY (all 17 files HTTP 200,
        # parsed fine). We do NOT weaken the guard and do NOT fake success — surface
        # as 'partial' (existing data left untouched) per the DefinitiveError contract.
        return Result(status="partial", obs=before, last_obs_date=None, new_vintage=None,
                      error=f"BLOCKER (never-shrink guard): {e}")

    # Re-base the tally's 'added' to the real net-new row count from the merge so
    # the status reflects actual additions (not gross parsed rows, which dedup away).
    net_new = max(0, n - before)
    tally.added = net_new
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
