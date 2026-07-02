"""S1 fetcher — SIPRI Military Expenditure Database (annual, 1949-present, ~173 countries).

Free for non-commercial use (SIPRI Terms of Use); no API key. SIPRI ships ONE
multi-sheet .xlsx per annual release (constant USD, current USD, %GDP, %govt
spending, per capita, local currency) covering all years; on a new vintage it
re-estimates/revises history, so we re-fetch the WHOLE workbook and MERGE
(dedup series_key+obs_date, new wins on revision, never-shrink). Single grouped
parquet clean_full/sipri/sipri.parquet, schema (series_key, obs_date, value)
with series_key='<prefix>:<country_key>'.

Vintage signal (registry adapter.vintage_signal): the dated bulk filename
SIPRI-Milex-data-1949-<year>.xlsx on www.sipri.org. We auto-discover the latest
live year (probing current/next year ahead of any hardcoded year) and use its
HEAD ETag/Last-Modified as the vintage token, so a new annual release is picked
up without editing a hardcoded URL.

NOTE on never-shrink: the existing published parquet was written by an ingester
with NO in-run (series_key,obs_date) dedup, so it carries ~8.8k duplicate rows
(the local-currency sheet(s) double-counted). merge_and_write dedups the union,
which COLLAPSES those pre-existing duplicates — a legitimate shrink of the raw
row count even though the DISTINCT data strictly grows. We therefore pass an
explicit lower min_ratio (0.80) for this source only; we do NOT weaken the shared
guard. A genuinely truncated/partial upstream pull (a missing sheet, a layout
break yielding 0 rows for a sheet) still trips the guard. CAUTION: the sipri/ dir
also holds milex.parquet from a DIFFERENT source (sipri_polity) — we only ever
touch sipri.parquet.
"""
from __future__ import annotations
import datetime as dt
import io
import os
import re
import time

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import http_vintage, UA

SOURCE = "sipri"
DEDUP = ("series_key", "obs_date")
# Pre-existing parquet has ~8.8k duplicate (key,date) rows the dedup will collapse;
# allow that legitimate shrink while still catching real truncation. NOT a shared-guard change.
MIN_RATIO = 0.80

# SIPRI hosts the dated bulk file under both hosts; www is primary.
HOST_TEMPLATES = [
    "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-{year}.xlsx",
    "https://sipri.org/sites/default/files/SIPRI-Milex-data-1949-{year}.xlsx",
]

# Sheet name -> series prefix (reused verbatim from jobs/ingest_sipri.py, incl. the
# substring + keyword fallback that tolerates "Constant (2024) US$" etc.).
SHEET_MAP = {
    "Constant (2022) US$":     "milex_usd_const",
    "Current US$":             "milex_usd_curr",
    "Share of GDP":            "milex_gdp_share",
    "Share of Govt. spending": "milex_gov_share",
    "Per capita":              "milex_percap",
    "Local currency":          "milex_lcu",
}


def _candidate_urls():
    """Latest-first dated URLs: probe next year, current year, and a few back so a
    new annual release is auto-discovered ahead of any frozen hardcoded year."""
    this_year = dt.date.today().year
    urls = []
    for year in range(this_year + 1, this_year - 4, -1):
        for tmpl in HOST_TEMPLATES:
            urls.append((year, tmpl.format(year=year)))
    return urls


def _head_live(url, session, tries=3):
    """HEAD with bounded retry. Returns the response if it 200s with a real body
    (>10 KB), 'gone' on a definite 404, or None on transient failure / other code —
    so a single reset on the newest year doesn't silently downgrade the vintage."""
    s = session or requests
    for a in range(tries):
        try:
            r = s.head(url, headers=UA, timeout=60, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError):
            if a < tries - 1:
                time.sleep(2 ** a)
            continue
        if r.status_code == 404:
            return "gone"
        if r.status_code == 200:
            cl = r.headers.get("Content-Length")
            if cl is None or (cl.isdigit() and int(cl) > 10000):
                return r
            return "gone"  # 200 but stub body -> treat as not-this-year
        if r.status_code in (429, 500, 502, 503, 504) and a < tries - 1:
            time.sleep(2 ** a)
            continue
        return None
    return None


def _discover_live(session=None, want_token=False):
    """Return the newest reachable dated bulk URL that HEADs 200 with a real body.
    Each candidate HEAD is retried so a transient reset on the newest year doesn't
    silently fall through to an older one. Returns (url, token) or (None, None);
    on (None, None) the strategy fetches anyway, which is safe."""
    s = session or requests
    for _year, url in _candidate_urls():
        r = _head_live(url, s)
        if r == "gone" or r is None:
            continue
        token = None
        if want_token:
            h = r.headers
            token = h.get("ETag") or h.get("Last-Modified") or h.get("Content-Length")
        # Fold the filename year in so a same-size next-year release still moves the token.
        fname = url.rsplit("/", 1)[-1]
        return url, (f"{fname}|{token}" if token else fname)
    return None, None


def current_vintage(unit):
    """Cheap probe: the live dated filename + its HEAD ETag/Last-Modified. Changes
    iff SIPRI publishes a new annual file (new year) or revises the current one."""
    _url, token = _discover_live(want_token=True)
    return token


def _parse_workbook(content):
    """Reuse jobs/ingest_sipri.py parse: per-sheet header-row + country-column
    heuristics -> (series_key, obs_date, value) rows."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    keys, dates, vals = [], [], []
    sheets_parsed = 0

    for sheet_name in wb.sheetnames:
        prefix = None
        for pattern, pfx in SHEET_MAP.items():
            if pattern.lower() in sheet_name.lower() or sheet_name.lower() in pattern.lower():
                prefix = pfx
                break
        if prefix is None:
            sn = sheet_name.lower().strip()
            if "constant" in sn:
                prefix = "milex_usd_const"
            elif "current" in sn and "us" in sn:
                prefix = "milex_usd_curr"
            elif "gdp" in sn:
                prefix = "milex_gdp_share"
            elif "gov" in sn:
                prefix = "milex_gov_share"
            elif "capita" in sn:
                prefix = "milex_percap"
            elif "local" in sn or "lcu" in sn:
                prefix = "milex_lcu"
            else:
                continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        header_row_idx = None
        for i, row in enumerate(rows[:20]):
            year_count = sum(1 for c in (row or []) if c is not None and
                             re.match(r"^\d{4}$", str(c).strip()) and
                             1940 <= int(str(c).strip()) <= 2030)
            if year_count >= 5:
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        headers = list(rows[header_row_idx])
        year_cols = {}
        for j, h in enumerate(headers):
            s = str(h).strip() if h is not None else ""
            if re.match(r"^\d{4}$", s) and 1940 <= int(s) <= 2030:
                year_cols[j] = int(s)

        country_col = 0
        sample_vals = [rows[header_row_idx + k][0] for k in range(1, min(6, len(rows) - header_row_idx))
                       if rows[header_row_idx + k]]
        non_none = [v for v in sample_vals if v is not None and len(str(v)) > 2]
        if len(non_none) < 2:
            country_col = 1

        sheet_count = 0
        for row in rows[header_row_idx + 1:]:
            if not row:
                continue
            country_raw = row[country_col] if country_col < len(row) else None
            if country_raw is None:
                continue
            country = str(country_raw).strip()
            if not country or country.lower() in ("country", "region", "area", "total", "notes", "note"):
                continue
            if len(country) > 60:
                continue
            country_key = re.sub(r"[^a-zA-Z0-9 ]", "", country).strip().replace(" ", "_")[:30]
            if not country_key:
                continue

            for col_idx, yr in year_cols.items():
                if col_idx >= len(row):
                    continue
                v_raw = row[col_idx]
                if v_raw is None:
                    continue
                s = str(v_raw).strip()
                if s in ("", "...", "xxx", "n/a", "N/A", "—", "-"):
                    continue
                try:
                    v = float(s.replace(",", ""))
                    if v != v:
                        continue
                except (ValueError, TypeError):
                    continue
                keys.append(f"{prefix}:{country_key}")
                dates.append(dt.date(yr, 12, 31))
                vals.append(v)
                sheet_count += 1
        if sheet_count:
            sheets_parsed += 1

    return keys, dates, vals, sheets_parsed


def _distinct_count(path):
    """Distinct (series_key, obs_date) rows in the published file (0 if absent).
    Used as an honest 'before' baseline: the file may hold pre-existing duplicate
    (key,date) rows that merge_and_write legitimately collapses."""
    if not blob.exists(path):
        return 0
    t = blob.read_table(path)
    if t.num_rows == 0:
        return 0
    return len(set(zip(t.column("series_key").to_pylist(),
                       t.column("obs_date").to_pylist())))


def _series_maxes(keys, dates):
    out = {}
    for k, d in zip(keys, dates):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "sipri.parquet")
    before = blob.row_count(path)
    tally = Tally()

    url, _token = _discover_live(want_token=False)
    if url is None:
        # No dated bulk file is reachable right now: treat as transient (retry next
        # tick) rather than faking success. Existing data left untouched.
        tally.transient_unit()
        return finalize(tally, before, None, source=SOURCE)

    # SIPRI's edge resets repeated full downloads; a couple of bounded retries on a
    # transient reset before giving up (and honestly reporting transient).
    r = None
    last_exc = None
    for attempt in range(3):
        try:
            r = requests.get(url, headers=UA, timeout=120)
            break
        except (requests.Timeout, requests.ConnectionError) as e:
            last_exc = e
            if attempt < 2:
                time.sleep(2 ** attempt)
    if r is None:
        tally.transient_unit()
        return finalize(tally, before, None, source=SOURCE)
    if r.status_code in (429, 500, 502, 503, 504):
        tally.transient_unit()
        return finalize(tally, before, None, source=SOURCE)
    if r.status_code != 200 or len(r.content) <= 10000:
        # Discovered URL went away / served a stub between HEAD and GET.
        tally.structural_unit()
        return finalize(tally, before, None, source=SOURCE)

    keys, dates, vals, sheets_parsed = _parse_workbook(r.content)

    if not vals:
        tally.structural_unit()  # 200 with a real workbook but parsed 0 rows -> layout break
        return finalize(tally, before, None, source=SOURCE)

    tbl = pa.table({"series_key": pa.array(keys, pa.string()),
                    "obs_date": pa.array(dates, pa.date32()),
                    "value": pa.array(vals, pa.float64())})

    # The published file may carry pre-existing un-deduped duplicate (key,date) rows,
    # so its raw row_count over-counts. Measure "added" against the DISTINCT existing
    # count so a one-time dedup collapse doesn't mask genuinely-added rows (and so
    # status is honest ok/no_change). merge_and_write returns the post-dedup total.
    before_distinct = _distinct_count(path)
    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP, min_ratio=MIN_RATIO)
    tally.added_unit(max(0, n - before_distinct))
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(keys, dates))
