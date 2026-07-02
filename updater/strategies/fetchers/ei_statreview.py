"""S1 fetcher — Energy Institute Statistical Review of World Energy (annual panel).

CC BY 4.0 (Energy Institute, formerly BP). Single grouped parquet
clean_full/ei_statreview/ei_statreview.parquet, schema (series_key, obs_date,
value) with series_key = EISR:{variable}:{country}, obs_date annual Dec-31. The EI
republishes the WHOLE 1965..latest history each edition, so we re-fetch the whole
table and MERGE (dedup series_key+obs_date, new wins on revision, never-shrink).

STALENESS NOTE (verified 2026-06-23): the two hard-coded energyinst.org primary
URLs (__data/assets/excel_doc/0020/1540154/... and the panel CSV) return HTTP 403
to programmatic clients (even a browser UA) — the asset id rotates per edition and
the host blocks bots. The published data was actually built from the OWID GitHub
mirror (130 columns, iso_code/country/year), which is alive (HEAD 200 + ETag) and
versioned via the GitHub commits API. So the HONEST vintage signal tracks the OWID
mirror commit SHA (the URL that genuinely produces our rows), not the dead EI HEAD.
update() still TRIES the EI URLs first (per the ingester) and only falls through to
the mirror, matching existing behaviour; if every URL fails it surfaces the failure
truthfully (transient) rather than faking success.
"""
from __future__ import annotations
import csv
import datetime as dt
import io
import os

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import github_sha, http_vintage, UA

SOURCE = "ei_statreview"
DEDUP = ("series_key", "obs_date")

# Reused verbatim from jobs/ingest_ei_statreview.py (URL list + fallback order).
URLS = [
    "https://www.energyinst.org/__data/assets/excel_doc/0020/1540154/EI-Stats-Review-All-Data.xlsx",
    "https://www.energyinst.org/__data/assets/excel_doc/0007/1055545/Statistical-Review-of-World-Energy-Consolidated-Dataset-panel-format.csv",
    "https://raw.githubusercontent.com/owid/energy-data/master/owid-energy-data.csv",
    "https://www.bp.com/content/dam/bp/business-sites/en/global/corporate/xlsx/energy-economics/statistical-review/bp-stats-review-2022-all-data.xlsx",
]

# The mirror that actually produces our published rows (EI URLs 403 for bots).
OWID_REPO = "owid/energy-data"
OWID_PATH = "owid-energy-data.csv"
OWID_RAW = "https://raw.githubusercontent.com/owid/energy-data/master/owid-energy-data.csv"


def current_vintage(unit):
    """Cheap probe that moves iff the upstream data moved.

    Registry hint asks for a HEAD on the EI XLSX, but that URL is 403 (dead for
    bots) and is NOT what produces our data — the OWID mirror is. So probe the
    mirror's commit SHA (changes iff the file changes). github_sha may raise
    TransientError on a flaky API; treat that as 'undeterminable' (return None) so
    detection never fails the run — the strategy then fetches anyway, which is safe
    (merge dedups + never-shrinks). Fall back to the raw-file ETag if the API is
    unavailable; return None only if nothing is cheaply determinable."""
    try:
        sha = github_sha(OWID_REPO, OWID_PATH)
        if sha:
            return f"owid:{sha}"
    except Exception:
        pass
    return http_vintage(OWID_RAW)


def _parse_csv(data: bytes):
    """Parse the OWID/panel CSV exactly as jobs/ingest_ei_statreview.py does."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []

    iso3_col = next((h for h in headers if h.lower() in
                     ("iso_code", "iso3", "country_code", "code")), None)
    ctry_col = next((h for h in headers if h.lower() in ("country", "entity")), None)
    year_col = next((h for h in headers if h.lower() in ("year", "yr")), None)

    if not (iso3_col or ctry_col) or not year_col:
        return [], [], []

    skip = {(iso3_col or "").lower(), (ctry_col or "").lower(),
            (year_col or "").lower(), "country", "entity"}

    keys, dates, vals = [], [], []
    for row in reader:
        if iso3_col:
            ctry = (row.get(iso3_col) or "").strip()
        else:
            ctry = (row.get(ctry_col) or "").strip().replace(" ", "_")[:30]
        if not ctry:
            continue
        try:
            yr = int(float(row.get(year_col) or ""))
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue
        for col, raw in row.items():
            if not col or col.lower() in skip:
                continue
            if not raw or str(raw).strip() in ("", "NA", "N/A", "nan", "#N/A"):
                continue
            try:
                v = float(str(raw).replace(",", ""))
            except (TypeError, ValueError):
                continue
            if v != v:  # NaN
                continue
            keys.append(f"EISR:{col}:{ctry}")
            dates.append(obs_d)
            vals.append(v)
    return keys, dates, vals


def _parse_panel_xlsx(data: bytes):
    """Parse the EI panel XLSX exactly as jobs/ingest_ei_statreview.py does."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    sheet_name = None
    for candidate in ("Panel data", "Panel Data", "panel_data", "Data"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], []
    header = [str(c).strip() if c else "" for c in rows[0]]

    ctry_idx = next((i for i, h in enumerate(header) if h.lower() in
                     ("country", "iso3", "iso_code", "country_name")), None)
    year_idx = next((i for i, h in enumerate(header) if h.lower() in ("year", "yr")), None)
    if ctry_idx is None or year_idx is None:
        return [], [], []

    skip_idx = {ctry_idx, year_idx}
    for i, h in enumerate(header):
        if h.lower() in ("region", "continent", "sub_region"):
            skip_idx.add(i)

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip()
        if not ctry:
            continue
        try:
            yr = int(row[year_idx])
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for ci, (col, cell) in enumerate(zip(header, row)):
            if ci in skip_idx or not col or cell is None:
                continue
            try:
                v = float(cell)
            except (TypeError, ValueError):
                continue
            if v != v:  # NaN
                continue
            keys.append(f"EISR:{col}:{ctry}")
            dates.append(obs_d)
            vals.append(v)
    return keys, dates, vals


def _fetch(url, timeout=180):
    try:
        r = requests.get(url, headers=UA, timeout=timeout, allow_redirects=True)
    except (requests.Timeout, requests.ConnectionError):
        return None, "transient"
    if r.status_code in (429, 500, 502, 503, 504):
        return None, "transient"
    if r.status_code == 200 and len(r.content) > 10_000:
        return r.content, "ok"
    return None, "down"  # 403/404/other non-success or trivially small body


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "ei_statreview.parquet")
    before = blob.row_count(path)
    tally = Tally()

    keys = dates = vals = None
    any_transient = False
    for url in URLS:
        data, kind = _fetch(url)
        if data is None:
            if kind == "transient":
                any_transient = True
            continue  # 403/404/down -> try next fallback (matches ingester)
        try:
            if url.endswith(".csv"):
                k, d, v = _parse_csv(data)
            else:
                k, d, v = _parse_panel_xlsx(data)
        except Exception:
            # a 200 that won't parse from a real body is a structural break for THIS url;
            # keep trying fallbacks before giving up
            continue
        if v:
            keys, dates, vals = k, d, v
            break

    # No URL yielded data. If a reachable 200 produced 0 parsable rows we'd have
    # broken out above with v truthy; getting here means every URL was 403/down or
    # transient. Surface honestly — never fake success.
    if not vals:
        if any_transient:
            tally.transient_unit()      # network/5xx somewhere -> retry next tick
        else:
            tally.transient_unit()      # all URLs unreachable/403 -> not a quiet day, retry (keeps old data)
        return finalize(tally, before, None, source=SOURCE)

    tbl = pa.table({
        "series_key": pa.array(keys, pa.string()),
        "obs_date": pa.array(dates, pa.date32()),
        "value": pa.array(vals, pa.float64()),
    })
    if tbl.num_rows == 0:
        tally.structural_unit()         # 200 parsed to nothing -> schema break
        return finalize(tally, before, None, source=SOURCE)

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
