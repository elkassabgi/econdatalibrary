"""S1 fetcher — Yale Environmental Performance Index (EPI), biennial country index.

CC BY 4.0 (Yale Center for Environmental Law & Policy). Single grouped parquet
clean_full/yale_epi/yale_epi.parquet, schema (series_key, obs_date, value),
series_key 'EPI:{variable}:{iso}'. The EPI results CSV is re-estimated/re-published
each release (currently epi2024results.csv, a wide format: code/iso/country columns
plus ~146 indicator columns). We re-fetch the WHOLE table and MERGE (dedup
series_key+obs_date, new wins on revision, never-shrink). One sub-unit (the CSV);
a 200 that parses 0 rows from a real body is structural.

RELEASES ARE DISCOVERED, NOT HARDCODED (2026-07-28). This fetcher used to pin
`epi2024results.csv` and probe the vintage of that one URL. Yale published EPI 2026
on 2026-07-07 at a DIFFERENT url — `epi2026results2026-07-07.xlsx` — so the pinned
file never changed, the probe never moved, and the source sat on 2024 data reporting
success while a whole new edition went unnoticed. Watching a fixed URL answers "did
this file change", not "did the publisher release something", and for a biennial
index those diverge exactly when it matters (ledger R73).

So the downloads page is scraped for every `epiYYYYresults*.{csv,xlsx}` link and all
of them are fetched. A page redesign that yields NO links is reported structurally
rather than silently falling back to a stale pin — but the known-good URLs are still
attempted, so a scrape failure costs a red run, never data.
"""
from __future__ import annotations
import csv
import datetime as dt
import io
import os
import re

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import http_vintage, UA

SOURCE = "yale_epi"
DEDUP = ("series_key", "obs_date")

DOWNLOADS = "https://epi.yale.edu/downloads"
# Floor, not the source of truth: these keep historical editions reachable even if
# the downloads page stops listing them (Yale currently lists only the newest).
KNOWN_URLS = [
    ("https://epi.yale.edu/downloads/epi2024results.csv", 2024),
]
_LINK_RE = re.compile(r'href=["\']([^"\']*epi(\d{4})results[^"\']*\.(?:csv|xlsx))',
                      re.I)

# Countries EPI 2026 added that the 2024 edition never listed, so the vocabulary map
# derived from 2024 cannot translate them and they would be dropped (~746 real
# observations). The published `code` column is ISO 3166-1 NUMERIC — verified by
# spot-check: EPI's Afghanistan is 4, and ISO 3166-1 numeric for AFG is 004. These
# two were looked up in the ISO 3166-1 table, NOT typed from memory, and are written
# unpadded to match EPI's own formatting.
EXTRA_COUNTRY_CODES = {
    "PLW": "585",   # Palau
    "KNA": "659",   # Saint Kitts and Nevis
}


def _discover():
    """[(url, year)] for every results file the downloads page currently lists."""
    try:
        r = requests.get(DOWNLOADS, headers=UA, timeout=120, allow_redirects=True)
    except (requests.Timeout, requests.ConnectionError):
        return []
    if r.status_code != 200:
        return []
    out = {}
    for href, yr in _LINK_RE.findall(r.text):
        url = href if href.startswith("http") else (
            "https://epi.yale.edu" + (href if href.startswith("/") else "/" + href))
        out[url] = int(yr)
    return sorted(out.items(), key=lambda kv: kv[1])


def current_vintage(unit):
    """Vintage covers the SET of published releases, not one pinned file.

    Combining the discovered URL list with each file's ETag/Last-Modified means the
    signal moves both when Yale revises an existing edition AND when it publishes a
    new one at a new URL — the case that let EPI 2026 go unnoticed for three weeks.
    """
    found = _discover()
    urls = [u for u, _ in found] or [u for u, _ in KNOWN_URLS]
    parts = ["|".join(sorted(urls))]
    for u in urls:
        v = http_vintage(u)
        if v:
            parts.append(f"{u}={v}")
    return "; ".join(parts) or None


def _country_index(data: bytes):
    """{ISO3 -> the country code our published ids actually use}, read from a file
    that carries both. Returns {} when the file has only one vocabulary."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    num = next((h for h in headers if h.lower() in ("code",)), None)
    alpha = next((h for h in headers if h.lower() in ("iso", "iso3")), None)
    if not (num and alpha):
        return {}
    out = {}
    for row in reader:
        a, n = (row.get(alpha) or "").strip(), (row.get(num) or "").strip()
        if len(a) == 3 and n:
            out[a] = n
    return out


def _parse_epi_csv(data: bytes, default_year: int, iso_index=None):
    """Parse EPI results CSV — wide: country column + many indicator columns.
    Byte-for-byte the same logic as jobs/ingest_yale_epi.parse_epi_csv, plus the
    country-vocabulary translation below."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []

    iso3_col = next((h for h in headers if h.lower() in
                     ("iso", "iso3", "iso_code", "country_iso3", "code")), None)
    year_col = next((h for h in headers if h.lower() in ("year", "yr")), None)
    if not iso3_col:
        return [], [], []

    # WHICH COUNTRY VOCABULARY THE IDS USE. The 2024 CSV carries BOTH `code`
    # (ISO 3166-1 NUMERIC: Afghanistan = 4) and `iso` (alpha-3: AFG), and this picks
    # whichever appears first in the file — which is `code`. So every one of the
    # 21,300 published yale_epi ids is keyed on the NUMERIC code (EPI:AGR.new:4).
    # The 2026 workbook ships only `iso`, so parsing it naively produced
    # EPI:AGR.new:AFG — a completely disjoint id space. Merging that would not have
    # failed: it would have added 63,354 brand-new series while all 21,300 live ones
    # stayed frozen at 2024, and the source's newest observation would read 2026,
    # making the health gate green over a source that had stopped updating. Exactly
    # the failure this whole day has been about.
    #
    # So alpha-3 is translated back to the published numeric vocabulary, using a map
    # read from the edition that carries both rather than a table typed here.
    # (Switching the ids to alpha-3 would be an improvement and a RE-KEY of 21,300
    # live series — not a call to make inside a fetcher.)
    translate = {}
    if iso_index and iso3_col.lower() in ("iso", "iso3", "iso_code", "country_iso3"):
        translate = iso_index

    skip = {(iso3_col or "").lower(), (year_col or "").lower(),
            "country", "region", "continent", "rank", "tier", "country.name"}

    keys, dates, vals = [], [], []
    n_untranslated = 0
    for row in reader:
        iso3 = (row.get(iso3_col) or "").strip()
        if not iso3:
            continue
        if translate:
            mapped = translate.get(iso3)
            if not mapped:
                # A country the reference edition never listed. Emitting it under the
                # wrong vocabulary would silently fork that country's series, so skip
                # and count — a named gap beats a quiet duplicate id space.
                n_untranslated += 1
                continue
            iso3 = mapped
        elif len(iso3) != 3:
            continue

        if year_col and row.get(year_col):
            try:
                yr = int(float(row[year_col]))
            except (ValueError, TypeError):
                yr = default_year
        else:
            yr = default_year
        obs_d = dt.date(yr, 12, 31)

        for col, raw in row.items():
            if col is None or col.lower().strip() in skip or not col:
                continue
            if not raw or str(raw).strip() in ("", "NA", "N/A", "nan", "#N/A", "-"):
                continue
            try:
                v = float(str(raw).replace(",", ""))
                if v != v:
                    continue
                keys.append(f"EPI:{col.strip()}:{iso3}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    if n_untranslated:
        print(f"[yale_epi] {default_year}: {n_untranslated} country row(s) had no "
              f"entry in the reference vocabulary and were SKIPPED", flush=True)
    return keys, dates, vals


def _parse_epi_xlsx(data: bytes, default_year: int, iso_index=None):
    """Parse an EPI results workbook — same wide shape as the CSV, in a 'data' sheet.

    EPI 2026 ships xlsx rather than csv (README + data sheets, 178 rows x 420
    indicator columns). Rows are handed to the CSV parser rather than reimplementing
    the value/skip/ISO rules, so the two formats cannot drift apart in how they
    decide what is an observation.
    """
    import openpyxl                                          # lazy: see requirements

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["data"] if "data" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = ["" if c is None else str(c).strip() for c in next(rows)]
    except StopIteration:
        return [], [], []
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(header)
    for r in rows:
        w.writerow(["" if c is None else c for c in r])
    return _parse_epi_csv(buf.getvalue().encode("utf-8"), default_year, iso_index)


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "yale_epi.parquet")
    before = blob.row_count(path)
    tally = Tally()

    found = _discover()
    if not found:
        # Loud, not silent. A downloads page that lists no results file means the
        # scrape broke or Yale restructured — either way the next release would be
        # missed, which is the whole failure being fixed here. Report it and still
        # fetch the known URLs so a scrape problem never costs data.
        tally.structural_unit("downloads page listed no epiYYYYresults file")
    urls, seen = [], set()
    for url, yr in list(found) + KNOWN_URLS:
        if url not in seen:
            seen.add(url)
            urls.append((url, yr))

    # TWO PASSES, because the country vocabulary is defined by one edition and
    # needed by another: fetch everything first, learn ISO3 -> published code from
    # whichever file carries both columns, then parse. Parsing as we download would
    # have meant the 2026 workbook was decoded before the 2024 CSV taught us what
    # its countries are called in our ids.
    bodies = []
    for url, yr in urls:
        try:
            r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError):
            tally.transient_unit()
            continue
        if r.status_code in (429, 500, 502, 503, 504):
            tally.transient_unit()
            continue
        if r.status_code != 200 or len(r.content) <= 500:
            # A wholesale 404 / tiny body is a structural break (Yale moved or
            # renamed the file), not a quiet day.
            tally.structural_unit(f"{yr}: HTTP {r.status_code}, {len(r.content)} B")
            continue
        bodies.append((url, yr, r.content))

    iso_index = {}
    for url, yr, body in bodies:
        if not url.lower().endswith(".xlsx"):
            iso_index = _country_index(body) or iso_index
    if iso_index:
        # Supplement, never override: a code the reference edition supplies wins.
        for k, v in EXTRA_COUNTRY_CODES.items():
            iso_index.setdefault(k, v)
    if iso_index:
        print(f"[yale_epi] country vocabulary: {len(iso_index)} ISO3 -> published "
              f"code mappings", flush=True)

    all_keys, all_dates, all_vals = [], [], []
    for url, yr, body in bodies:
        try:
            if url.lower().endswith(".xlsx"):
                k, d, v = _parse_epi_xlsx(body, default_year=yr, iso_index=iso_index)
            else:
                k, d, v = _parse_epi_csv(body, default_year=yr)
        except Exception as e:                               # noqa: BLE001
            # A workbook we cannot open is a structural break on THAT release, named
            # so the log says which one rather than "yale_epi failed".
            print(f"[yale_epi] {url}: {type(e).__name__}: {e}", flush=True)
            tally.structural_unit(f"{yr}: unreadable ({type(e).__name__})")
            continue
        if not v:
            tally.structural_unit(f"{yr}: 200 but parsed 0 rows")  # schema break
            continue
        print(f"[yale_epi] {yr}: {len(v):,} obs from {url.rsplit('/', 1)[-1]}",
              flush=True)
        all_keys.extend(k)
        all_dates.extend(d)
        all_vals.extend(v)

    tbl = pa.table({
        "series_key": pa.array(all_keys, pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals, pa.float64()),
    })

    if tbl.num_rows == 0:
        # Nothing parsed across all URLs; Tally already recorded transient/structural.
        return finalize(tally, before, None, source=SOURCE)

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    tally.added_unit(max(0, n - before))
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
