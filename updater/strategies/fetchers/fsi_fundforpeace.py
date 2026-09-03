"""S1 fetcher — Fragile States Index (FSI), Fund for Peace (annual, 2006-present).

CC BY-NC 3.0 (Fund for Peace). Single grouped parquet
clean_full/fsi_fundforpeace/fsi_fundforpeace.parquet, schema
(series_key, obs_date, value). series_key = 'FSI_FP:{indicator}:{country}',
obs_date = Dec-31 of the index year. Each annual release is a discrete static
XLSX; FSI occasionally revises prior years, so we re-fetch ALL years and MERGE
(dedup series_key+obs_date, new wins on revision, never-shrink).

VINTAGE / URL note (verified 2026-06): the registry's named primary signal — the
GitHub CSV mirror raw.githubusercontent.com/ksreyes/tidy-fragile-states-index —
is DEAD (repo deleted, both the API and the raw CSV return 404). The hardcoded
per-year fragilestatesindex.org URLs in jobs/ingest_fsi_fundforpeace.py were also
stale for 2006-2020 (wrong WordPress upload paths -> 404). The resilient source
of truth is the FSI "excel" page, which lists the current canonical XLSX URL for
every year; we scrape it (so a new release is picked up automatically) and fall
back to a hardcoded current-URL table if the page is unreachable. The cheap
vintage probe content-hashes the discovered URL list (changes when a year is
added or a URL moves) and falls back to a HEAD on the newest year's XLSX.
"""
from __future__ import annotations
import datetime as dt
import io
import os
import re

import pyarrow as pa
import requests

from ... import config, blob, merge
from ..base import Result
from ._common import Tally, finalize
from ._vintage import http_vintage, content_hash, UA

SOURCE = "fsi_fundforpeace"
DEDUP = ("series_key", "obs_date")

# Canonical FSI download index — lists the current XLSX URL for every year.
EXCEL_PAGE = "https://fragilestatesindex.org/excel/"

# Current (verified-live 2026-06) per-year XLSX URLs from the excel page, used as a
# fallback if the page scrape fails. Years 2006-2017 live under /uploads/data/,
# later years under year/month upload folders. Keep newest first so the vintage
# probe's HEAD hits the most-recently-revised file.
FALLBACK_URLS = [
    "https://fragilestatesindex.org/wp-content/uploads/2023/06/FSI-2023-DOWNLOAD.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/2022/07/fsi-2022-download.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/2021/05/fsi-2021.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/2020/05/fsi-2020.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/2019/04/fsi-2019.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/2018/04/fsi-2018.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2017.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2016.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2015.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2014.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2013.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2012.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2011.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2010.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2009.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2008.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2007.xlsx",
    "https://fragilestatesindex.org/wp-content/uploads/data/fsi-2006.xlsx",
]

# A browser-ish UA for the HTML index page (WordPress 403s some bot agents).
PAGE_UA = {"User-Agent": "Mozilla/5.0 (compatible; Econ-Fin Data Library admin@hfdatalibrary.com)"}

# Header columns that are not indicator values.
_SKIP_COLS = {"country", "country name", "countryname", "name", "rank", "change", "trend", "year"}


def _discover_urls() -> list[str]:
    """Scrape the FSI excel index page for all year XLSX links (newest first).

    Returns [] on any failure so the caller falls back to FALLBACK_URLS. Never
    raises — vintage/discovery failing must not fail the run."""
    try:
        r = requests.get(EXCEL_PAGE, headers=PAGE_UA, timeout=60)
    except (requests.Timeout, requests.ConnectionError):
        return []
    if r.status_code != 200:
        return []
    links = re.findall(r"""href=["']?([^"' >]+\.xlsx)""", r.text, re.I)
    seen, out = set(), []
    for u in links:
        if not u.lower().startswith("http"):
            continue
        if _year_of(u) is None:
            continue
        if u not in seen:
            seen.add(u)
            out.append(u)
    out.sort(key=lambda u: (_year_of(u) or 0), reverse=True)
    return out


def _year_of(url: str) -> int | None:
    m = re.search(r"(20\d\d)", url)
    return int(m.group(1)) if m else None


def current_vintage(unit):
    """Cheap probe: content-hash of the discovered XLSX URL list (changes when a
    year is added or a URL moves). Falls back to a HEAD vintage on the newest
    year's XLSX, then None (strategy fetches anyway, which is safe)."""
    urls = _discover_urls()
    if urls:
        return "urls:" + content_hash("\n".join(sorted(urls)).encode("utf-8"))
    head = http_vintage(FALLBACK_URLS[0])
    return ("head:" + head) if head else None


def _parse_xlsx(data: bytes, year: int):
    """Parse one FSI annual XLSX -> (keys, dates, vals). obs_date = Dec-31 of year."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], []
    header = [str(c).strip() if c else "" for c in rows[0]]
    ctry_idx = next((i for i, h in enumerate(header)
                     if h.lower() in ("country", "country name", "countryname", "name")), None)
    if ctry_idx is None:
        return [], [], []
    skip_idx = {i for i, h in enumerate(header) if h.lower() in _SKIP_COLS}
    skip_idx.add(ctry_idx)

    obs_d = dt.date(year, 12, 31)
    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip()
        if not ctry or ctry.lower() in ("country", "nan"):
            continue
        for ci, (col, cell) in enumerate(zip(header, row)):
            if ci in skip_idx or not col or cell is None:
                continue
            try:
                v = float(cell)
            except (TypeError, ValueError):
                continue
            if v != v or v < 0:  # NaN or negative -> not a valid index value
                continue
            keys.append(f"FSI_FP:{col}:{ctry}")
            dates.append(obs_d)
            vals.append(v)
    return keys, dates, vals


def _series_maxes(tbl):
    out = {}
    if tbl.num_rows == 0:
        return out
    for k, d in zip(tbl.column("series_key").to_pylist(), tbl.column("obs_date").to_pylist()):
        if d is None:
            continue
        if k not in out or d > out[k]:
            out[k] = d
    return {k: v.isoformat() for k, v in out.items()}


def update(unit, since) -> Result:
    out_dir = config.source_dir(SOURCE)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "fsi_fundforpeace.parquet")
    before = blob.row_count(path)
    tally = Tally()

    urls = _discover_urls() or list(FALLBACK_URLS)

    all_keys, all_dates, all_vals = [], [], []
    for url in urls:
        year = _year_of(url)
        if year is None:
            continue
        try:
            r = requests.get(url, headers=UA, timeout=60, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError) as e:
            tally.transient_unit(f"{year}: {type(e).__name__} fetching {url[-46:]}")
            continue
        if r.status_code in (429, 500, 502, 503, 504):
            tally.transient_unit(f"{year}: HTTP {r.status_code}")
            continue
        if r.status_code == 404:
            # one stale per-year URL; not fatal — the year may simply have moved.
            tally.empty_unit(f"{year}: 404, the release URL may have moved")
            continue
        if r.status_code != 200 or len(r.content) < 5000:
            tally.structural_unit(
                f"{year}: HTTP {r.status_code}, {len(r.content):,} bytes "
                f"(under the 5,000-byte floor)")
            continue
        try:
            k, d, v = _parse_xlsx(r.content, year)
        except Exception as e:  # noqa: BLE001
            tally.structural_unit(f"{year}: XLSX will not parse — {type(e).__name__}")
            continue
        if not v:
            # 200 with a real XLSX body but parsed 0 rows -> schema break
            tally.structural_unit(f"{year}: real XLSX parsed 0 rows")
            continue
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
        tally.added_unit(len(v))  # provisional; net new vs merge resolved below

    tbl = pa.table({
        "series_key": pa.array(all_keys, pa.string()),
        "obs_date": pa.array(all_dates, pa.date32()),
        "value": pa.array(all_vals, pa.float64()),
    })

    if tbl.num_rows == 0:
        # Nothing parsed from any year. If every attempt transient-failed, finalize
        # surfaces 'partial'; otherwise it raises (structural/empty-window) — honest.
        return finalize(tally, before, None, source=SOURCE)

    n, md = merge.merge_and_write(path, tbl, mode="merge", dedup_keys=DEDUP)
    # Reset the added counter to the TRUE net-new (post-dedup) so status is honest.
    net_new = max(0, n - before)
    tally.added = net_new
    return finalize(tally, n, md, source=SOURCE, series_cursors=_series_maxes(tbl))
