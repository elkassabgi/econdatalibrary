"""S1/S5 fetcher — IEP Global Terrorism Index (GTI).

CC BY-NC-SA 4.0 (Institute for Economics & Peace; granted to the Elkassabgi Data
Library for non-commercial re-hosting 2026-07-06). Single grouped parquet
clean_full/gti/gti.parquet, schema (series_key, obs_date, value),
series_key = "GTI:<indicator>:<ISO3>" (annual, Dec-31).

Workbook layout (verified against GTI_PublicReleaseData_2026.xlsx): a 'Contents'
sheet (licence text only) plus a 'Data' sheet in LONG format — one row per
country-year. The real header is a few rows down (branding above); columns are
Country | iso3c | year | rank | Score | Incidents | Fatalities | Injuries |
Hostages. ISO3 (iso3c) is native and complete, so it is the series id. Every
non-identity numeric column becomes an indicator. Higher Score = greater impact
of terrorism (0-10); rank 1 = worst-affected.
"""
from __future__ import annotations
import io
import re
from datetime import date

from ._iep import probe_vintage, run_update

SOURCE = "gti"
URLS = [
    "https://www.economicsandpeace.org/wp-content/uploads/2026/03/GTI_PublicReleaseData_2026.xlsx",
]


def _slug(s):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(s).strip()).strip("_")


def parse(xlsx_bytes):
    """IEP GTI 'Data' sheet (long format) -> (keys, dates, vals)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    ws = wb["Data"] if "Data" in wb.sheetnames else None
    candidates = [ws] if ws is not None else list(wb.worksheets)

    rows = header_idx = header = None
    for sheet in candidates:
        srows = list(sheet.iter_rows(values_only=True))
        for i, r in enumerate(srows):
            lc = [str(c).strip().lower() if c is not None else None for c in r]
            if "country" in lc:
                rows, header_idx, header = srows, i, r
                break
        if rows is not None:
            break
    if rows is None:
        raise ValueError("GTI: no data sheet with a 'Country' header row found")

    col = {str(name).strip(): ci for ci, name in enumerate(header) if name is not None}
    country_ci = col["Country"]

    iso_ci = None
    for cand in ("iso3c", "iso3", "ISO3", "Country code", "geocode", "iso_3", "ISO3C"):
        if cand in col:
            iso_ci = col[cand]
            break
    if iso_ci is None:
        for name, ci in col.items():
            if name.lower() in ("iso3c", "iso3", "country code", "geocode"):
                iso_ci = ci
                break

    year_ci = None
    for cand in ("year", "Year", "YEAR"):
        if cand in col:
            year_ci = col[cand]
            break
    if year_ci is None:
        raise ValueError("GTI: no 'year' column found")

    identity = {country_ci, iso_ci, year_ci}
    indicator_cols = [ci for ci, name in enumerate(header) if name is not None and ci not in identity]

    keys, dates, vals = [], [], []
    for r in rows[header_idx + 1:]:
        if all(c is None for c in r):
            continue
        country, year_raw = r[country_ci], r[year_ci]
        if country is None or year_raw is None:
            continue
        try:
            year = int(year_raw)
        except (TypeError, ValueError):
            continue

        ident = None
        if iso_ci is not None:
            i3 = r[iso_ci]
            if isinstance(i3, str) and len(i3.strip()) == 3 and i3.strip().isalpha():
                ident = i3.strip().upper()
        if ident is None:
            ident = str(country).strip()

        obs = date(year, 12, 31)
        for ci in indicator_cols:
            v = r[ci]
            if v is None or isinstance(v, bool) or not isinstance(v, (int, float)):
                continue
            keys.append(f"GTI:{_slug(header[ci])}:{ident}")
            dates.append(obs)
            vals.append(float(v))

    wb.close()
    return keys, dates, vals


def current_vintage(unit):
    return probe_vintage(URLS)


def update(unit, since):
    return run_update(SOURCE, URLS, parse)
