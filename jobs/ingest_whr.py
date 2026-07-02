#!/usr/bin/env python3
"""World Happiness Report (WHR) ingest.

Source: https://worldhappiness.report/data/
License: CC BY 4.0 (Gallup / Sustainable Development Solutions Network)
Coverage: ~155 countries, 2005-present, life satisfaction + 6 explanatory variables.

Downloads WHR data from the official GitHub/Kaggle/OWID mirrors.
series_key: WHR:{variable}:{iso3}  e.g. WHR:Life.Ladder:FIN

Output: data/clean_full/whr/whr.parquet
Run: python jobs/ingest_whr.py
"""
from __future__ import annotations
import csv, datetime as dt, io, os, time
import requests, pyarrow as pa, pyarrow.parquet as pq

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "whr")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# WHR data: Gallup World Poll panel (2005-2023)
# OWID grapher CSV endpoints (confirmed working)
URLS = [
    # OWID Cantril ladder (life satisfaction scores) - confirmed accessible
    "https://ourworldindata.org/grapher/happiness-cantril-ladder.csv",
    # OWID contributing factors
    "https://ourworldindata.org/grapher/log-gdp-per-capita-whr.csv",
    "https://ourworldindata.org/grapher/social-support-whr.csv",
    "https://ourworldindata.org/grapher/healthy-life-expectancy-whr.csv",
    "https://ourworldindata.org/grapher/freedom-to-make-life-choices.csv",
    "https://ourworldindata.org/grapher/generosity-whr.csv",
    "https://ourworldindata.org/grapher/perceptions-of-corruption-whr.csv",
    # GitHub mirrors (may work)
    "https://raw.githubusercontent.com/erikgahner/PolData/master/Data/WHR/WorldHappinessReport2023.csv",
    "https://raw.githubusercontent.com/WardF/world-happiness-report/main/data/whr_main.csv",
]

# The full dataset from the official WHR appendix
# S3 bucket requires Referer header from worldhappiness.report
PANEL_URLS = [
    # 2025 report (2011-2024 data) — en-dash URL-encoded as %E2%80%93
    "https://happiness-report.s3.us-east-1.amazonaws.com/2025/Data+for+Figure+2.1+(2011%E2%80%932024).xlsx",
    # Older naming conventions
    "https://happiness-report.s3.us-east-1.amazonaws.com/2024/DataForTable2.1WHR2024.xlsx",
    "https://happiness-report.s3.amazonaws.com/2024/DataForTable2.1WHR2024.xls",
    "https://happiness-report.s3.amazonaws.com/2023/DataForTable2.1WHR2023.xls",
    # Figshare mirror (open access)
    "https://figshare.com/ndownloader/files/49981044",  # Figshare WHR 2020-2024
]
# Keep old names for backward compatibility in main()
PANEL_URL  = PANEL_URLS[0]
PANEL_URL2 = PANEL_URLS[1]
PANEL_URL3 = PANEL_URLS[2]
PANEL_URL4 = PANEL_URLS[3]


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url):
    try:
        r = requests.get(url, headers=UA, timeout=60, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 500:
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_whr_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    log(f"  Headers ({len(headers)}): {headers[:10]}")

    # Key columns: country name (use iso3 from pycountry or fallback), year
    ctry_col = next((h for h in headers if h.lower().strip() in
                     ("country.name", "country_name", "country name", "country",
                      "entity", "name")), None)
    iso3_col = next((h for h in headers if h.lower().strip() in
                     ("iso_code", "iso3", "code", "country_code")), None)
    year_col = next((h for h in headers if h.lower().strip() in ("year", "yr")), None)

    if not (ctry_col or iso3_col) or not year_col:
        log(f"  Missing country/year columns"); return [], [], []

    skip = {(ctry_col or "").lower(), (iso3_col or "").lower(),
            (year_col or "").lower(), "country", "entity", "region"}

    keys, dates, vals = [], [], []
    for row in reader:
        iso3 = ""
        if iso3_col:
            iso3 = (row.get(iso3_col) or "").strip()
        if not iso3 and ctry_col:
            # Use country name as key (no ISO3 available)
            iso3 = (row.get(ctry_col) or "").strip().replace(" ", "_")[:30]
        if not iso3:
            continue

        yr_raw = (row.get(year_col) or "").strip()
        try:
            yr = int(float(yr_raw))
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue

        for col, raw in row.items():
            if col.lower().strip() in skip or not col:
                continue
            if not raw or str(raw).strip() in ("", "NA", "N/A", "nan", "#N/A"):
                continue
            try:
                v = float(str(raw).replace(",", ""))
                if v != v:
                    continue
                keys.append(f"WHR:{col.strip()}:{iso3}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    return keys, dates, vals


def parse_xls_panel(data: bytes):
    """Parse WHR Excel appendix file (DataForTable2.1)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    log(f"  Columns: {header[:10]}")

    ctry_idx = next((i for i, h in enumerate(header) if "country" in h.lower()), None)
    year_idx = next((i for i, h in enumerate(header) if h.lower() in ("year", "yr")), None)
    if ctry_idx is None or year_idx is None:
        return [], [], []

    skip_idx = {ctry_idx, year_idx}
    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip().replace(" ", "_")[:30]
        try:
            yr = int(row[year_idx])
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for ci, (col, cell) in enumerate(zip(header, row)):
            if ci in skip_idx or not col or cell is None:
                continue
            try:
                v = float(cell)
                if v != v:
                    continue
                keys.append(f"WHR:{col}:{ctry}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass
    return keys, dates, vals


def main():
    os.makedirs(OUT, exist_ok=True)
    out = os.path.join(OUT, "whr.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows
        log(f"WHR: already {n:,} rows"); return

    log("=== World Happiness Report Ingest ===")
    all_keys, all_dates, all_vals = [], [], []

    # Try XLS panel first (most complete historical data)
    for url in [PANEL_URL, PANEL_URL2, PANEL_URL3, PANEL_URL4]:
        log(f"Trying XLS panel: {url[-70:]}...")
        data = fetch(url)
        if data:
            try:
                k, d, v = parse_xls_panel(data)
                if v:
                    log(f"  Got {len(v):,} obs from XLS")
                    all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
                    break
            except Exception as e:
                log(f"  XLS parse error: {e}")
        time.sleep(1)

    # Try CSV mirrors — accumulate ALL (each OWID endpoint = one variable)
    for url in URLS:
        log(f"Trying CSV: {url[-70:]}...")
        data = fetch(url)
        if data:
            k, d, v = parse_whr_csv(data)
            if v:
                log(f"  Got {len(v):,} obs")
                all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
        time.sleep(0.5)

    if not all_vals:
        log("0 observations parsed"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out, compression="zstd")
    n = pq.read_metadata(out).num_rows
    log(f"=== WHR DONE: {n:,} obs ===")


if __name__ == "__main__":
    main()
