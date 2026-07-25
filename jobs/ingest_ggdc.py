#!/usr/bin/env python3
"""GGDC (Groningen Growth and Development Centre) data ingest.

Sources:
  Penn World Tables 10.0:
    URL: https://www.rug.nl/ggdc/docs/pwt100.xlsx
    183 countries, 1950-2019, 48 macro/productivity variables
    Ref: Feenstra, Inklaar & Timmer (2015), AER

  Maddison Project Database 2020:
    URL: https://www.rug.nl/ggdc/historicaldevelopment/maddison/data/mpd2020.xlsx
    169 countries, 1 AD–2018, GDP pc and population (2011 int'l $)
    Ref: Maddison Project Database 2020, Bolt & van Zanden (2020)

Output:
  data/clean_full/ggdc/pwt10.parquet
  data/clean_full/ggdc/maddison2020.parquet

Series key format:
  PWT10:{variable}:{countrycode}   e.g. PWT10:rgdpe:USA
  MADDISON:{variable}:{countrycode} e.g. MADDISON:gdppc:CHN

Run: python jobs/ingest_ggdc.py
"""
from __future__ import annotations
import datetime as dt
import io
import os
import time

import requests
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
OUT  = os.path.join(ROOT, "data", "clean_full", "ggdc")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

PWT_URL      = "https://www.rug.nl/ggdc/docs/pwt100.xlsx"
MADDISON_URL = "https://www.rug.nl/ggdc/historicaldevelopment/maddison/data/mpd2020.xlsx"

# Newer versions on Dataverse (GET works, HEAD returns 403)
PWT11_URL    = "https://dataverse.nl/api/access/datafile/554105"   # PWT 11.0, 185 ctry, 1950-2023
MAD23_URL    = "https://dataverse.nl/api/access/datafile/421302"   # Maddison 2023, 169 ctry, to 2022

# PWT id columns to skip (not numeric variables)
PWT_ID_COLS = {"countrycode", "country", "currency_unit", "year"}


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url: str, retries: int = 3) -> bytes | None:
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
            if r.status_code == 200 and len(r.content) > 10_000:
                return r.content
            log(f"  HTTP {r.status_code} or tiny ({len(r.content)} bytes): {url[-80:]}")
        except Exception as e:
            log(f"  ERR attempt {attempt+1}: {e}")
        time.sleep(5 * (attempt + 1))
    return None


def save(keys, dates, vals, path):
    tbl = pa.table({
        "series_key": pa.array(keys,  pa.string()),
        "obs_date":   pa.array(dates, pa.date32()),
        "value":      pa.array(vals,  pa.float64()),
    })
    pq.write_table(tbl, path, compression="zstd")
    n = pq.read_metadata(path).num_rows
    log(f"  -> {n:,} obs saved to {os.path.basename(path)}")
    return n


# ---------------------------------------------------------------------------
# Penn World Tables 10.0
# ---------------------------------------------------------------------------

def ingest_pwt():
    out = os.path.join(OUT, "pwt10.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"PWT10: already {n:,} rows"); return n

    log("Downloading PWT 10.0...")
    data = fetch(PWT_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024:,} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Data"]
    log(f"  Sheet 'Data': {ws.max_row:,} rows × {ws.max_column} cols")

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Build column map: col_idx -> variable name (skip id cols)
    col_map = {}
    ctry_idx = year_idx = None
    for ci, col in enumerate(header):
        if col is None:
            continue
        col = str(col).strip()
        if col == "countrycode":
            ctry_idx = ci
        elif col == "year":
            year_idx = ci
        elif col not in PWT_ID_COLS:
            col_map[ci] = col

    if ctry_idx is None or year_idx is None:
        log("  Missing countrycode/year column"); return 0

    log(f"  {len(col_map)} value columns: {list(col_map.values())[:10]}...")

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None or row[year_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip()
        try:
            yr = int(row[year_idx])
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue

        for ci, varname in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            try:
                v = float(cell)
                if v != v or v in (float("inf"), float("-inf")):
                    continue
                keys.append(f"PWT10:{varname}:{ctry}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    log(f"  Parsed {len(vals):,} obs from {ws.max_row-1:,} country-year rows")
    return save(keys, dates, vals, out)


# ---------------------------------------------------------------------------
# Maddison Project Database 2020
# ---------------------------------------------------------------------------

def ingest_maddison():
    out = os.path.join(OUT, "maddison2020.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"Maddison2020: already {n:,} rows"); return n

    log("Downloading Maddison Project Database 2020...")
    data = fetch(MADDISON_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024:,} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Full data"]
    log(f"  Sheet 'Full data': {ws.max_row:,} rows × {ws.max_column} cols")

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Columns: countrycode, country, year, gdppc, pop
    col_idx = {str(col).strip(): ci for ci, col in enumerate(header) if col is not None}
    log(f"  Columns: {list(col_idx.keys())}")

    ctry_ci = col_idx.get("countrycode")
    year_ci = col_idx.get("year")
    val_cols = [(col, ci) for col, ci in col_idx.items()
                if col not in ("countrycode", "country", "year")]

    if ctry_ci is None or year_ci is None:
        log("  Missing countrycode/year"); return 0

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_ci] is None or row[year_ci] is None:
            continue
        ctry = str(row[ctry_ci]).strip()
        try:
            yr = int(row[year_ci])
            # Maddison covers 1 AD onwards; use Dec 31 as obs date
            if yr < 1:
                continue
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue

        for varname, ci in val_cols:
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            try:
                v = float(cell)
                if v != v or v in (float("inf"), float("-inf")):
                    continue
                keys.append(f"MADDISON:{varname}:{ctry}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    log(f"  Parsed {len(vals):,} obs")
    return save(keys, dates, vals, out)


# ---------------------------------------------------------------------------
# Penn World Tables 11.0
# ---------------------------------------------------------------------------

def ingest_pwt11():
    out = os.path.join(OUT, "pwt11.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"PWT11: already {n:,} rows"); return n

    log("Downloading PWT 11.0 from Dataverse...")
    data = fetch(PWT11_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024:,} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")
    ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
    log(f"  Sheet '{ws.title}': {ws.max_row:,} rows × {ws.max_column} cols")

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_map = {}
    ctry_idx = year_idx = None
    for ci, col in enumerate(header):
        if col is None:
            continue
        col = str(col).strip()
        if col == "countrycode":
            ctry_idx = ci
        elif col == "year":
            year_idx = ci
        elif col not in PWT_ID_COLS:
            col_map[ci] = col

    if ctry_idx is None or year_idx is None:
        log("  Missing countrycode/year column"); return 0

    log(f"  {len(col_map)} value columns")

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_idx] is None or row[year_idx] is None:
            continue
        ctry = str(row[ctry_idx]).strip()
        try:
            yr = int(row[year_idx])
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for ci, varname in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            try:
                v = float(cell)
                if v != v or v in (float("inf"), float("-inf")):
                    continue
                keys.append(f"PWT11:{varname}:{ctry}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    log(f"  Parsed {len(vals):,} obs")
    return save(keys, dates, vals, out)


# ---------------------------------------------------------------------------
# Maddison Project Database 2023
# ---------------------------------------------------------------------------

def ingest_maddison2023():
    out = os.path.join(OUT, "maddison2023.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"Maddison2023: already {n:,} rows"); return n

    log("Downloading Maddison 2023 from Dataverse...")
    data = fetch(MAD23_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024:,} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    # Try "Full data" or similar sheet
    ws = None
    for sn in ["Full data", "Data", "data", "full_data", wb.sheetnames[0]]:
        if sn in wb.sheetnames:
            ws = wb[sn]; break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    log(f"  Using sheet '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_idx = {str(col).strip(): ci for ci, col in enumerate(header) if col is not None}
    log(f"  Columns: {list(col_idx.keys())}")

    # Find country code and year — use explicit None check (index 0 is falsy!)
    def find_col(*names):
        for n in names:
            if n in col_idx:
                return col_idx[n]
        return None

    ctry_ci = find_col("countrycode", "iso3", "code")
    year_ci = find_col("year")
    val_cols = [(col, ci) for col, ci in col_idx.items()
                if col not in ("countrycode", "country", "year", "iso3", "code", "region")]

    if ctry_ci is None or year_ci is None:
        log(f"  Missing countrycode/year. Cols: {list(col_idx.keys())}"); return 0

    keys, dates, vals = [], [], []
    for row in rows[1:]:
        if not row or row[ctry_ci] is None or row[year_ci] is None:
            continue
        ctry = str(row[ctry_ci]).strip()
        try:
            yr = int(row[year_ci])
            if yr < 1:
                continue
            obs_d = dt.date(yr, 12, 31)
        except (TypeError, ValueError):
            continue
        for varname, ci in val_cols:
            if ci >= len(row) or row[ci] is None:
                continue
            try:
                v = float(row[ci])
                if v != v or v in (float("inf"), float("-inf")):
                    continue
                keys.append(f"MADDISON23:{varname}:{ctry}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass

    log(f"  Parsed {len(vals):,} obs")
    return save(keys, dates, vals, out)


def main():
    os.makedirs(OUT, exist_ok=True)
    log("=== GGDC Data Ingest (PWT 10+11 + Maddison 2020+2023) ===")
    total = 0

    log("--- Penn World Tables 10.0 ---")
    total += ingest_pwt()
    time.sleep(1)

    log("--- Penn World Tables 11.0 ---")
    total += ingest_pwt11()
    time.sleep(1)

    log("--- Maddison Project Database 2020 ---")
    total += ingest_maddison()
    time.sleep(1)

    log("--- Maddison Project Database 2023 ---")
    total += ingest_maddison2023()

    log(f"=== GGDC TOTAL: {total:,} observations ===")


if __name__ == "__main__":
    main()
