#!/usr/bin/env python3
"""Freedom House — Freedom in the World (FIW), 1973–2024.

License: Creative Commons Attribution 4.0
Source: https://freedomhouse.org/report/freedom-world
No API key required (direct XLSX download).

Coverage:
  * Political Rights score (1–7, lower = more free), ~210 countries
  * Civil Liberties score (1–7, lower = more free), ~210 countries
  * Freedom Status (Free/Partly Free/Not Free)
  * Electoral Democracy designation (0/1)
  * Annual, 1973–2024

Run: python jobs/ingest_freedomhouse.py
"""
from __future__ import annotations
import datetime as dt, io, os, re, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "freedomhouse")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

URLS = [
    "https://freedomhouse.org/sites/default/files/2025-03/Country_and_Territory_Ratings_and_Statuses_FIW_1973-2025.xlsx",
    "https://freedomhouse.org/sites/default/files/2024-02/Country_and_Territory_Ratings_and_Statuses_FIW_1973-2024.xlsx",
    "https://freedomhouse.org/sites/default/files/2023-02/Country_and_Territory_Ratings_and_Statuses_FIW_1973-2023.xlsx",
]

# Also: subscores file (A-G indicators)
SUBSCORE_URLS = [
    "https://freedomhouse.org/sites/default/files/2024-02/FITW_Data_FIW_2013-2024.xlsx",
    "https://freedomhouse.org/sites/default/files/2023-02/FITW_Data_FIW_2013-2023.xlsx",
]

STATUS_MAP = {"F": 1.0, "PF": 2.0, "NF": 3.0}  # Freedom status → numeric


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii','replace').decode()}", flush=True)


def try_download(url: str) -> bytes | None:
    try:
        r = requests.get(url, headers=UA, timeout=120)
        if r.status_code == 200 and len(r.content) > 5000:
            log(f"  OK: {len(r.content):,} bytes from {url[-70:]}")
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_fiw_xlsx(content: bytes) -> list[tuple[str, dt.date, float]]:
    """Parse Freedom in the World XLSX.

    The main format is wide: rows = countries, columns = years (PR, CL, Status for each year).
    Header row has years like 1973, 1974, ... with PR/CL/Status sub-columns.
    """
    import openpyxl
    results = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        log(f"  Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            if "rating" not in sheet_name.lower() and "fiw" not in sheet_name.lower():
                log(f"  Skipping sheet: {sheet_name}")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 5:
                continue

            log(f"  Sheet '{sheet_name}': {len(rows)} rows")

            # Find header rows — look for rows containing year numbers
            year_row_idx = None
            col_row_idx = None  # row with PR/CL/Status sub-headers

            for i, row in enumerate(rows[:10]):
                year_count = sum(1 for c in (row or []) if c is not None and
                                 re.match(r"^(19|20)\d{2}$", str(c).strip()))
                if year_count >= 3:
                    year_row_idx = i
                    break

            # Look for a row with PR/CL/Status pattern
            for i, row in enumerate(rows[:15]):
                pr_count = sum(1 for c in (row or []) if c is not None and
                               str(c).strip().upper() in ("PR", "CL", "STATUS", "C.L.", "P.R."))
                if pr_count >= 3:
                    col_row_idx = i
                    break

            if year_row_idx is None:
                log(f"  No year row in '{sheet_name}'"); continue

            # Use either the year row or col row as the primary header
            header_row = rows[year_row_idx]
            subheader_row = rows[col_row_idx] if col_row_idx is not None else None

            # Build column mapping: col_idx → (year, metric)
            col_map: dict[int, tuple[int, str]] = {}
            current_year = None

            for j, cell in enumerate(header_row):
                if cell is None:
                    continue
                s = str(cell).strip()
                m = re.match(r"^(19|20)\d{2}$", s)
                if m:
                    current_year = int(s)

            # If sub-headers exist, map more precisely
            if subheader_row and current_year:
                # Scan both rows together
                current_yr = None
                for j, (yr_cell, sub_cell) in enumerate(zip(header_row, subheader_row)):
                    yr_s = str(yr_cell).strip() if yr_cell is not None else ""
                    sub_s = str(sub_cell).strip().upper() if sub_cell is not None else ""
                    if re.match(r"^(19|20)\d{2}$", yr_s):
                        current_yr = int(yr_s)
                    if current_yr and sub_s in ("PR", "P.R."):
                        col_map[j] = (current_yr, "political_rights")
                    elif current_yr and sub_s in ("CL", "C.L."):
                        col_map[j] = (current_yr, "civil_liberties")
                    elif current_yr and sub_s in ("STATUS", "STATUS (F/PF/NF)"):
                        col_map[j] = (current_yr, "freedom_status")
            else:
                # Single header row, no sub-headers
                # Columns might be: Country, PR2024, CL2024, Status2024, ...
                for j, cell in enumerate(header_row):
                    s = str(cell).strip() if cell is not None else ""
                    m = re.match(r"^(PR|CL|Status)[.\s]*(\d{4})$", s, re.IGNORECASE)
                    if m:
                        metric_map = {"pr": "political_rights", "cl": "civil_liberties",
                                      "status": "freedom_status"}
                        metric = metric_map.get(m.group(1).lower(), m.group(1))
                        col_map[j] = (int(m.group(2)), metric)
                    else:
                        # Try year only in header, PR/CL/Status as every-3-columns
                        yr_m = re.match(r"^(19|20)\d{2}$", s)
                        if yr_m:
                            yr = int(s)
                            col_map[j]   = (yr, "political_rights")
                            col_map[j+1] = (yr, "civil_liberties")
                            col_map[j+2] = (yr, "freedom_status")

            # Find country column (usually col 0 or 1)
            data_start = max(year_row_idx, col_row_idx if col_row_idx else 0) + 1
            country_col = 0  # default

            # Check if first column is numeric (rank) or text (country)
            sample = [rows[data_start+k][0] for k in range(3) if data_start+k < len(rows)]
            if all(isinstance(s, (int, float)) for s in sample if s is not None):
                country_col = 1  # rank in col 0

            log(f"  {len(col_map)} year-metric columns, data from row {data_start}")

            if not col_map:
                continue

            for row in rows[data_start:]:
                if not row:
                    continue
                country_raw = row[country_col] if country_col < len(row) else None
                if country_raw is None or str(country_raw).strip() in ("", "Country"):
                    continue
                country = str(country_raw).strip()[:60]
                if not country or country.lower() in ("nan", "none"):
                    continue

                for col_idx, (yr, metric) in col_map.items():
                    if col_idx >= len(row):
                        continue
                    v_raw = row[col_idx]
                    if v_raw is None or str(v_raw).strip() in ("", "N/A", "-"):
                        continue

                    v_str = str(v_raw).strip()
                    if metric == "freedom_status":
                        v = STATUS_MAP.get(v_str.upper().replace(" ", ""))
                        if v is None:
                            continue
                    else:
                        try:
                            v = float(v_str)
                            if v != v:
                                continue
                        except (ValueError, TypeError):
                            continue

                    results.append((f"{metric}:{country}", dt.date(yr, 12, 31), v))

    except Exception as e:
        log(f"  XLSX parse error: {e}")
        import traceback; traceback.print_exc()
    return results


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "freedomhouse.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    all_keys, all_dates, all_vals = [], [], []
    seen: set[tuple] = set()

    for url in URLS:
        log(f"Trying {url[-70:]}")
        content = try_download(url)
        if not content:
            continue
        results = parse_fiw_xlsx(content)
        for key, d, v in results:
            tok = (key, d)
            if tok not in seen:
                seen.add(tok)
                all_keys.append(key)
                all_dates.append(d)
                all_vals.append(v)
        log(f"  After this file: {len(all_vals):,} total obs")
        if len(all_vals) > 10000:
            break

    if not all_vals:
        log("0 observations — check URLs"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} Freedom House FIW observations ({len(set(all_keys))} series)")


if __name__ == "__main__":
    main()
