#!/usr/bin/env python3
"""Penn World Tables (PWT) 10.0 — real GDP, TFP, capital, labour across 183 countries.

License: CC BY 4.0
Source: https://www.rug.nl/ggdc/productivity/pwt/
No API key required (direct XLSX download).

Variables (47): rgdpe, rgdpo, pop, emp, avh, hc, ccon, cda, cgdpe, cgdpo,
  cn, ck, ctfp, cwtfp, rconna, rdana, irr, delta, xr, pl_con, pl_da, pl_gdpo,
  i_cig, i_xm, i_xr, i_outlier, i_irr, cor_exp, statcap, csh_c, csh_i,
  csh_g, csh_x, csh_m, csh_r, pl_c, pl_i, pl_g, pl_x, pl_m, pl_n, pl_k,
  labsh, ulc, tfpna, rnna, rcon

Run: python jobs/ingest_pwt.py
"""
from __future__ import annotations
import datetime as dt, io, os, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "pwt")
URL  = "https://www.rug.nl/ggdc/docs/pwt100.xlsx"
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}


def log(m): print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "pwt.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    log(f"Downloading Penn World Tables 10.0 from {URL}")
    try:
        r = requests.get(URL, headers=UA, timeout=300, stream=True)
        r.raise_for_status()
        content = r.content
    except Exception as e:
        log(f"ERROR: {e}"); return

    log(f"Downloaded {len(content):,} bytes; parsing XLSX...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        log(f"Sheets: {wb.sheetnames}")

        # Find the data sheet (usually "Data")
        data_sheet = None
        for name in wb.sheetnames:
            if name.lower() in ("data", "pwt100", "pwt10"):
                data_sheet = wb[name]; break
        if data_sheet is None:
            data_sheet = wb.active

        # Read all rows
        rows = list(data_sheet.iter_rows(values_only=True))
        if not rows:
            log("ERROR: no rows in data sheet"); return

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        log(f"Columns ({len(headers)}): {headers[:10]}...")

        # Identify key columns
        country_col = next((i for i, h in enumerate(headers)
                            if h.lower() in ("countrycode", "country_code", "iso", "isocode")), None)
        year_col    = next((i for i, h in enumerate(headers)
                            if h.lower() in ("year", "yr")), None)
        country_name_col = next((i for i, h in enumerate(headers)
                                 if h.lower() in ("country", "countryname", "country_name")), None)

        if year_col is None:
            log("ERROR: cannot find year column"); return

        # All numeric columns that aren't metadata
        skip_cols = {country_col, year_col, country_name_col}
        val_cols = [(i, headers[i]) for i in range(len(headers))
                    if i not in skip_cols and headers[i] and headers[i] not in ("", "None")]

        all_keys, all_dates, all_vals = [], [], []
        for row in rows[1:]:
            yr_raw = row[year_col] if year_col is not None else None
            if yr_raw is None:
                continue
            try:
                yr = int(yr_raw)
                d  = dt.date(yr, 12, 31)
            except (ValueError, TypeError):
                continue

            iso = (str(row[country_col]).strip() if country_col is not None
                   and row[country_col] is not None else "XXX")

            for col_idx, col_name in val_cols:
                v_raw = row[col_idx]
                if v_raw is None:
                    continue
                try:
                    v = float(v_raw)
                except (ValueError, TypeError):
                    continue
                all_keys.append(f"{col_name}:{iso}")
                all_dates.append(d)
                all_vals.append(v)

    except Exception as e:
        log(f"XLSX parse error: {e}"); return

    if not all_vals:
        log("0 observations"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} PWT observations")


if __name__ == "__main__":
    main()
