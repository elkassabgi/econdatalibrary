#!/usr/bin/env python3
"""UNDP Human Development Report (HDR) Statistical Annexes — 190+ countries.

License: CC BY 3.0 IGO
Source: https://hdr.undp.org/data-center/documentation-and-downloads
No API key required (direct XLSX download from hdr.undp.org).

Coverage:
  * HDI (Human Development Index), 1990–2023
  * GDI (Gender Development Index)
  * GII (Gender Inequality Index)
  * IHDI (Inequality-adjusted HDI)
  * MPI (Multidimensional Poverty Index)
  * Education Index, Life Expectancy Index, GNI per capita
  * Mean years of schooling, Expected years of schooling
  * Life expectancy at birth
  * GNI per capita (2017 PPP$)

Run: python jobs/ingest_undp_hdr.py
"""
from __future__ import annotations
import datetime as dt, io, os, re, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "undp_hdr")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# UNDP HDR 2023/24 Statistical Annexes
# Each file has country rows with multiple year columns
BASE_URL = "https://hdr.undp.org/sites/default/files/2023-24_HDR"

FILES = [
    # (filename, series_prefix, description)
    ("HDR23-24_Statistical_Annex_HDI_Table.xlsx",   "hdi",   "Human Development Index"),
    ("HDR23-24_Statistical_Annex_GDI_Table.xlsx",   "gdi",   "Gender Development Index"),
    ("HDR23-24_Statistical_Annex_GII_Table.xlsx",   "gii",   "Gender Inequality Index"),
    ("HDR23-24_Statistical_Annex_IHDI_Table.xlsx",  "ihdi",  "Inequality-adjusted HDI"),
    ("HDR23-24_Statistical_Annex_MPI_Table.xlsx",   "mpi",   "Multidimensional Poverty Index"),
]

# Also try the composite HDR data CSV (full time series, all indicators)
COMPOSITE_URLS = [
    "https://hdr.undp.org/sites/default/files/2023-24_HDR/HDR23-24_Composite_indices_complete_time_series.csv",
    "https://hdr.undp.org/sites/default/files/2024_statistical_annex/HDR2024_Statistical_Annex_Tables.zip",
    "https://hdr.undp.org/sites/default/files/2023-24_HDR/HDR23-24_Statistical_Annex_HDI_Trends_Table.xlsx",
]


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii','replace').decode()}", flush=True)


def try_download(url: str, timeout: int = 120) -> bytes | None:
    try:
        r = requests.get(url, headers=UA, timeout=timeout)
        if r.status_code == 200 and len(r.content) > 1000:
            log(f"  OK: {len(r.content):,} bytes from {url[-70:]}")
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_year_cols(headers: list) -> dict[int, int]:
    """Return {col_index: year} for columns whose header looks like a 4-digit year."""
    result = {}
    for i, h in enumerate(headers):
        s = str(h).strip() if h is not None else ""
        m = re.match(r"^(\d{4})$", s)
        if m:
            yr = int(m.group(1))
            if 1980 <= yr <= 2030:
                result[i] = yr
    return result


def ingest_composite_csv(content: bytes, all_keys, all_dates, all_vals) -> int:
    """Parse the composite time-series CSV.

    Supports two layouts:
    1. Long format:  iso3, year, indicator_columns...
    2. Wide format:  iso3, indicator_1990, indicator_1991, ... (col names like 'hdi_2022')
    """
    import csv
    count = 0
    try:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        log(f"  CSV headers ({len(headers)}): {headers[:15]}")

        iso_col = next((h for h in headers if h.lower() in ("iso3", "iso_code", "country_code", "code")), None)
        yr_col  = next((h for h in headers if h.lower() == "year"), None)
        skip_meta = {"country", "country_name", "hdicode", "region", "hdi_rank_2022",
                     "hdi_rank_2023", "hdi_rank", iso_col, yr_col}

        if yr_col:
            # Long format: one row per country/year
            val_cols = [h for h in headers if h not in skip_meta and h]
            log(f"  Long format: iso_col={iso_col}, yr_col={yr_col}, {len(val_cols)} value cols")
            for row in reader:
                iso = str(row.get(iso_col, "")).strip()[:20] if iso_col else "WLD"
                if not iso: continue
                try:
                    yr = int(float(str(row.get(yr_col, "")).strip()))
                    obs_date = dt.date(yr, 12, 31)
                except (ValueError, TypeError):
                    continue
                for col in val_cols:
                    v_raw = row.get(col, "")
                    if not v_raw or str(v_raw).strip() in ("", "..", "NA", "n.a.", "N/A"):
                        continue
                    try:
                        v = float(str(v_raw).replace(",", ""))
                        if v != v: continue
                    except (ValueError, TypeError):
                        continue
                    all_keys.append(f"{col.lower().strip()}:{iso}")
                    all_dates.append(obs_date)
                    all_vals.append(v)
                    count += 1
        else:
            # Wide format: columns named like 'hdi_1990', 'gni_per_capita_2022'
            # Split on last underscore to extract year
            wide_cols = []  # (col_name, series_name, year)
            for h in headers:
                if h in skip_meta or not h:
                    continue
                m = re.search(r"_(\d{4})$", h)
                if m:
                    yr = int(m.group(1))
                    if 1980 <= yr <= 2030:
                        series = h[:m.start()].lower().strip()
                        wide_cols.append((h, series, yr))
            log(f"  Wide format: {len(wide_cols)} year-indicator columns")
            for row in reader:
                iso = str(row.get(iso_col, "")).strip()[:20] if iso_col else "WLD"
                if not iso: continue
                for col, series, yr in wide_cols:
                    v_raw = row.get(col, "")
                    if not v_raw or str(v_raw).strip() in ("", "..", "NA", "n.a.", "N/A"):
                        continue
                    try:
                        v = float(str(v_raw).replace(",", ""))
                        if v != v: continue
                    except (ValueError, TypeError):
                        continue
                    all_keys.append(f"{series}:{iso}")
                    all_dates.append(dt.date(yr, 12, 31))
                    all_vals.append(v)
                    count += 1

    except Exception as e:
        log(f"  CSV parse error: {e}")
        import traceback; traceback.print_exc()
    return count


def ingest_annex_xlsx(content: bytes, series_prefix: str, all_keys, all_dates, all_vals) -> int:
    """Parse an UNDP HDR Statistical Annex XLSX.

    Format: rows = countries, columns = years (wide) plus indicator sub-columns.
    The exact layout varies by table; we scan for numeric year headers and ISO codes.
    """
    import openpyxl
    count = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 5:
                continue

            # Find header row(s) with year numbers
            header_row_idx = None
            for i, row in enumerate(rows[:20]):
                year_count = sum(1 for c in (row or []) if c is not None and
                                 re.match(r"^\d{4}$", str(c).strip()) and
                                 1980 <= int(str(c).strip()) <= 2030)
                if year_count >= 3:
                    header_row_idx = i
                    break

            if header_row_idx is None:
                log(f"  Sheet '{sheet_name}': no year columns found, skipping")
                continue

            headers = list(rows[header_row_idx])
            year_cols = parse_year_cols(headers)
            log(f"  Sheet '{sheet_name}': {len(year_cols)} year columns, header row {header_row_idx}")

            if not year_cols:
                continue

            # Find ISO/country column: look for 3-letter uppercase in data rows
            iso_col_idx = None
            for j, h in enumerate(headers):
                col_vals = [rows[k][j] for k in range(header_row_idx+1, min(header_row_idx+10, len(rows)))
                            if rows[k] and j < len(rows[k])]
                iso_like = sum(1 for v in col_vals if v and re.match(r"^[A-Z]{3}$", str(v).strip()))
                if iso_like >= 3:
                    iso_col_idx = j
                    break

            if iso_col_idx is None:
                # Try col 0 for HDI rank, col 1 for country name, see if there's an HDI value
                iso_col_idx = 1  # fallback to country name

            for row in rows[header_row_idx+1:]:
                if not row:
                    continue
                iso_raw = row[iso_col_idx] if iso_col_idx < len(row) else None
                if iso_raw is None:
                    continue
                iso = str(iso_raw).strip()[:20]
                if not iso or iso.lower() in ("", "nan", "none"):
                    continue

                for col_idx, yr in year_cols.items():
                    if col_idx >= len(row):
                        continue
                    v_raw = row[col_idx]
                    if v_raw is None or str(v_raw).strip() in ("", "..", "NA"):
                        continue
                    try:
                        v = float(str(v_raw).replace(",", ""))
                        if v != v:
                            continue
                    except (ValueError, TypeError):
                        continue
                    all_keys.append(f"{series_prefix}:{iso}")
                    all_dates.append(dt.date(yr, 12, 31))
                    all_vals.append(v)
                    count += 1

    except Exception as e:
        log(f"  XLSX parse error ({series_prefix}): {e}")
        import traceback; traceback.print_exc()
    return count


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "undp_hdr.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    all_keys, all_dates, all_vals = [], [], []
    total = 0

    # 1. Try composite time-series CSV (best source — one file, all indicators, all years)
    log("Trying composite time-series CSV...")
    for url in COMPOSITE_URLS:
        content = try_download(url)
        if content:
            if url.endswith(".csv"):
                n = ingest_composite_csv(content, all_keys, all_dates, all_vals)
                log(f"  Composite CSV: {n:,} obs")
                total += n
                if total > 0:
                    break
            elif url.endswith(".xlsx"):
                # HDI trends table — treat as annex
                n = ingest_annex_xlsx(content, "hdi_trend", all_keys, all_dates, all_vals)
                log(f"  Trends XLSX: {n:,} obs")
                total += n

    # 2. Download individual Statistical Annex XLSXes
    log("Downloading individual Statistical Annex files...")
    for fname, prefix, desc in FILES:
        url = f"{BASE_URL}/{fname}"
        content = try_download(url)
        if not content:
            # try alternate naming
            url2 = f"{BASE_URL}/{fname.replace('23-24', '2024')}"
            content = try_download(url2)
        if not content:
            log(f"  Skipping {fname}: not accessible")
            continue
        n = ingest_annex_xlsx(content, prefix, all_keys, all_dates, all_vals)
        log(f"  {prefix}: {n:,} obs from {fname}")
        total += n

    if not all_vals:
        log("0 observations — UNDP HDR endpoints may have changed"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} UNDP HDR observations ({len(set(all_keys))} series)")


if __name__ == "__main__":
    main()
