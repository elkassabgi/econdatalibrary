#!/usr/bin/env python3
"""Social Progress Index (SPI) ingest.

Source: https://www.socialprogress.org/
Publisher: Social Progress Imperative
License: Creative Commons Attribution-NonCommercial 4.0
Coverage: 170+ countries, 2011-2023, 3 dimensions, 12 components, 54 indicators

Dimensions:
  1. Basic Human Needs (nutrition, water, shelter, safety)
  2. Foundations of Wellbeing (education, health, environment, info access)
  3. Opportunity (rights, freedom, inclusiveness, advanced education)

series_key: SPI:{indicator}:{iso3}   e.g. SPI:SPI_score:USA

Output: data/clean_full/spi/spi.parquet
Run: python jobs/ingest_social_progress.py
"""
from __future__ import annotations
import csv, datetime as dt, io, os, re, time
import requests
import pyarrow as pa, pyarrow.parquet as pq

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "spi")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# Multiple URL attempts
SPI_URLS = [
    # GitHub official mirror
    "https://raw.githubusercontent.com/social-progress-imperative/open-data/main/data/spi_complete.csv",
    "https://raw.githubusercontent.com/social-progress-imperative/open-data/master/data/spi_complete.csv",
    # Direct from socialprogress.org (may redirect)
    "https://www.socialprogress.org/static/ee39ec5d9b4c5b0f6e7a04b9aece43f2/2024-spi-results.xlsx",
    # Tableau / dashboards
    "https://www.socialprogress.org/static/f7d4e5c8b2a9d1e3f6c7a8b9c0d1e2f3/spi-2023-full.csv",
    # OWID
    "https://raw.githubusercontent.com/owid/owid-datasets/master/datasets/Social%20Progress%20Index%20(Social%20Progress%20Imperative)/Social%20Progress%20Index%20(Social%20Progress%20Imperative).csv",
    # Harvard Dataverse mirror
    "https://dataverse.harvard.edu/api/access/datafile/:persistentId/?persistentId=doi:10.7910/DVN/IXA4NB",
    # Kaggle public
    "https://raw.githubusercontent.com/public-data-archives/social-progress-index/main/data/spi.csv",
    # World Bank microdata
    "https://raw.githubusercontent.com/datasets/social-progress-index/master/data/spi.csv",
    # Alternative: the 2023 Excel from Social Progress Imperative CDN
    "https://cdn.socialprogress.org/files/2023-social-progress-index-results.xlsx",
    "https://cdn.socialprogress.org/files/2024-social-progress-index-results.xlsx",
    "https://www.socialprogress.org/assets/downloads/2023-Social-Progress-Index-Results-and-Trends.xlsx",
    "https://www.socialprogress.org/assets/downloads/2024-Social-Progress-Index-Results-and-Trends.xlsx",
    # IDB / InterAmerican Development Bank mirror
    "https://publications.iadb.org/en/node/33847/download",
    # OPHI / Oxford
    "https://ophi.org.uk/wp-content/uploads/2023-spi.csv",
]


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url: str) -> bytes | None:
    for attempt in range(2):
        try:
            r = requests.get(url, headers=UA, timeout=60, allow_redirects=True)
            if r.status_code == 200 and len(r.content) > 1000:
                # Check it's not HTML
                if b"<html" in r.content[:200].lower() or b"<!DOCTYPE" in r.content[:200]:
                    log(f"  Got HTML page")
                    return None
                log(f"  {len(r.content)//1024:,} KB from {url[-60:]}")
                return r.content
            log(f"  HTTP {r.status_code}")
            if r.status_code in (403, 404):
                return None
        except Exception as e:
            log(f"  ERR: {e}")
        time.sleep(2)
    return None


def parse_spi_csv(data: bytes) -> tuple[list, list, list]:
    """Parse SPI CSV with columns: country, year, SPI score, dimensions, indicators."""
    try:
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or [])]
        log(f"  CSV columns ({len(headers)}): {headers[:10]}")

        def find_col(*candidates):
            for c in candidates:
                for h in headers:
                    if h.lower() == c.lower() or h.lower().strip("_") == c.lower().strip("_"):
                        return h
            return None

        code_col = find_col("Code", "ISO3", "iso3", "Country Code", "countrycode",
                            "iso3c", "iso_code")
        name_col = find_col("Country", "country", "Name", "country_name")
        year_col = find_col("Year", "year")
        id_col   = code_col or name_col

        if not id_col:
            log(f"  No country identifier found in: {headers[:8]}")
            return [], [], []

        skip_cols = {code_col, name_col, year_col,
                     find_col("region"), find_col("continent")}
        skip_cols = {c for c in skip_cols if c}

        val_cols = []
        for h in headers:
            if h in skip_cols or not h:
                continue
            safe = re.sub(r"[^a-zA-Z0-9_]", "_", h)[:30].strip("_")
            if safe:
                val_cols.append((h, safe))

        log(f"  Value columns: {len(val_cols)}")
        if not val_cols:
            return [], [], []

        keys, dates, vals = [], [], []
        snapshot_date = dt.date(2023, 12, 31)  # default year if no year column

        for rec in reader:
            cid = (rec.get(id_col) or "").strip()
            if not cid:
                continue
            entity = re.sub(r"[^a-zA-Z0-9_]", "_", cid)[:20].strip("_")

            if year_col:
                try:
                    yr = int(float((rec.get(year_col) or "").strip()))
                    obs_d = dt.date(yr, 12, 31)
                except (ValueError, TypeError):
                    obs_d = snapshot_date
            else:
                obs_d = snapshot_date

            for col_name, col_label in val_cols:
                raw = rec.get(col_name, "")
                if not raw or str(raw).strip() in ("", "NA", "N/A", ".."):
                    continue
                try:
                    v = float(str(raw).strip())
                    if v != v:
                        continue
                    keys.append(f"SPI:{col_label}:{entity}")
                    dates.append(obs_d)
                    vals.append(v)
                except (ValueError, TypeError):
                    pass

        log(f"  Parsed {len(vals):,} obs")
        return keys, dates, vals

    except Exception as e:
        log(f"  Parse error: {e}")
        return [], [], []


def parse_spi_xlsx(data: bytes) -> tuple[list, list, list]:
    """Parse SPI Excel file."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        log(f"  Sheets: {wb.sheetnames}")
        keys, dates, vals = [], [], []

        for sn in wb.sheetnames:
            if any(s in sn.lower() for s in ["cover", "note", "method", "about"]):
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 5:
                continue

            # Find header row
            header_idx = None
            for ri, row in enumerate(rows[:10]):
                row_lower = [str(c).lower() if c else "" for c in row]
                if any(v in row_lower for v in ["country", "iso3", "code", "year"]):
                    header_idx = ri; break

            if header_idx is None:
                continue

            header = [str(c).strip() if c else "" for c in rows[header_idx]]
            log(f"  Sheet '{sn}': header at row {header_idx}: {header[:6]}")

            code_ci = next((i for i, h in enumerate(header)
                           if h.lower() in ("code", "iso3", "iso_code", "country code")), None)
            name_ci = next((i for i, h in enumerate(header)
                           if h.lower() in ("country", "country name", "name")), None)
            year_ci = next((i for i, h in enumerate(header)
                           if h.lower() in ("year",)), None)
            id_ci   = code_ci if code_ci is not None else name_ci

            if id_ci is None:
                continue

            skip_ci = {code_ci, name_ci, year_ci}
            val_cols = [(i, re.sub(r"[^a-zA-Z0-9_]", "_", h)[:30].strip("_"))
                        for i, h in enumerate(header)
                        if i not in skip_ci and h]

            snapshot_date = dt.date(2023, 12, 31)

            for row in rows[header_idx + 1:]:
                if not row or row[id_ci] is None:
                    continue
                cid = str(row[id_ci]).strip()
                if not cid or cid.lower() in ("nan", "none"):
                    continue
                entity = re.sub(r"[^a-zA-Z0-9_]", "_", cid)[:20]

                if year_ci is not None and year_ci < len(row) and row[year_ci] is not None:
                    try:
                        yr = int(float(str(row[year_ci]).strip()))
                        obs_d = dt.date(yr, 12, 31)
                    except (TypeError, ValueError):
                        obs_d = snapshot_date
                else:
                    obs_d = snapshot_date

                for col_i, col_label in val_cols:
                    if col_i >= len(row) or row[col_i] is None:
                        continue
                    try:
                        v = float(row[col_i])
                        if v != v:
                            continue
                        keys.append(f"SPI:{col_label}:{entity}")
                        dates.append(obs_d)
                        vals.append(v)
                    except (TypeError, ValueError):
                        pass

            if keys:
                break

        return keys, dates, vals

    except Exception as e:
        log(f"  XLSX parse error: {e}")
        return [], [], []


def main():
    os.makedirs(OUT, exist_ok=True)
    out = os.path.join(OUT, "spi.parquet")

    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows
        log(f"SPI: already {n:,} rows"); return

    log("=== Social Progress Index Ingest ===")

    for url in SPI_URLS:
        log(f"Trying {url[-70:]}...")
        data = fetch(url)
        if not data:
            continue

        if url.endswith(".xlsx"):
            k, d, v = parse_spi_xlsx(data)
        else:
            k, d, v = parse_spi_csv(data)

        if v:
            tbl = pa.table({
                "series_key": pa.array(k,  pa.string()),
                "obs_date":   pa.array(d, pa.date32()),
                "value":      pa.array(v,  pa.float64()),
            })
            pq.write_table(tbl, out, compression="zstd")
            n = pq.read_metadata(out).num_rows
            log(f"=== SPI DONE: {n:,} obs ===")
            return

        log("  0 obs from this URL")

    log("All SPI URLs failed")


if __name__ == "__main__":
    main()
