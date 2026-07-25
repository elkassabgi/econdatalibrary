#!/usr/bin/env python3
"""SIPRI Military Expenditure + Polity/CSP Political Data ingest.

SIPRI:
  URL: https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2025_v1.2.xlsx
  Sheets: Current USD (millions), Constant (2022) USD, Share of GDP, Share of govt spending, % Change
  Output: data/clean_full/sipri/milex.parquet

Polity 5 (Center for Systemic Peace):
  p5v2018.xls — Polity 5 country-year dataset (democracy/autocracy scores)
  MEPVv2018.xls — Major Episodes of Political Violence (country-year)
  Output: data/clean_full/polity/*.parquet

Series key format:
  SIPRI:milex:{sheet}:{country}
  POLITY:{variable}:{country}

Run: python jobs/ingest_sipri_polity.py
"""
from __future__ import annotations
import datetime as dt
import io
import os
import time

import requests
import pyarrow as pa
import pyarrow.parquet as pq
import openpyxl
import xlrd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
SIPRI_OUT  = os.path.join(ROOT, "data", "clean_full", "sipri")
POLITY_OUT = os.path.join(ROOT, "data", "clean_full", "polity")
UA = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}
RATE = 0.5


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url: str, retries: int = 4) -> bytes | None:
    for attempt in range(retries):
        try:
            log(f"  GET {url[-70:]}")
            r = requests.get(url, headers=UA, timeout=120)
            if r.status_code == 200:
                return r.content
            if r.status_code == 404:
                log(f"  404"); return None
            log(f"  HTTP {r.status_code}")
        except Exception as e:
            log(f"  ERR attempt {attempt+1}: {e}")
        time.sleep(5 * (attempt + 1))
    return None


def save(keys, dates, vals, out_path):
    tbl = pa.table({
        "series_key": pa.array(keys,  pa.string()),
        "obs_date":   pa.array(dates, pa.date32()),
        "value":      pa.array(vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"  -> {n:,} obs saved to {os.path.basename(out_path)}")
    return n


# ---------------------------------------------------------------------------
# SIPRI
# ---------------------------------------------------------------------------

# SIPRI xlsx sheet name → short label for series key
SIPRI_SHEETS = {
    "Current US$": "usd_curr",
    "Constant (2024) US$": "usd_const",
    "Share of GDP": "gdp_share",
    "Share of Govt. spending": "govt_share",
    "Per capita": "per_capita",
}


def parse_sipri_sheet(ws, sheet_label: str) -> tuple[list, list, list]:
    """Parse one SIPRI sheet. Row 1+ = countries, Col 1 = country name, Col 2+ = years."""
    keys, dates, vals = [], [], []

    # Find the header row (contains years as integers)
    header_row = None
    year_cols = {}  # col_idx -> year
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Look for a row where most cells are year integers (e.g. 1949, 1950, ...)
        year_count = sum(1 for c in row if isinstance(c, (int, float)) and 1900 < c < 2100)
        if year_count > 10:
            header_row = r_idx
            for c_idx, cell in enumerate(row):
                if isinstance(cell, (int, float)) and 1900 < cell < 2100:
                    year_cols[c_idx] = int(cell)
            break

    if header_row is None:
        log(f"  SIPRI {sheet_label}: no header row found"); return [], [], []
    log(f"  SIPRI {sheet_label}: {len(year_cols)} years, first={min(year_cols.values())}, last={max(year_cols.values())}")

    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row:
            continue
        if not row or row[0] is None:
            continue
        country = str(row[0]).strip()
        if not country or country.lower() in ("country", "notes", "note", "source", "sources"):
            continue
        for c_idx, yr in year_cols.items():
            if c_idx >= len(row):
                continue
            cell_val = row[c_idx]
            if cell_val is None or cell_val == "" or str(cell_val).strip() in ("", "...", "xxx", "NA", "N/A"):
                continue
            try:
                v = float(cell_val)
                if v != v or v in (float("inf"), float("-inf")):  # nan/inf check
                    continue
                key = f"SIPRI:milex:{sheet_label}:{country}"
                keys.append(key)
                dates.append(dt.date(yr, 12, 31))
                vals.append(v)
            except (TypeError, ValueError):
                pass

    return keys, dates, vals


def ingest_sipri():
    out = os.path.join(SIPRI_OUT, "milex.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"SIPRI: already {n:,} rows"); return n

    log("SIPRI: downloading SIPRI-Milex-data-1949-2025_v1.2.xlsx...")
    data = fetch("https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2025_v1.2.xlsx")
    if not data:
        return 0

    log(f"  SIPRI: {len(data)/1024:.0f} KB, opening workbook...")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []
    for sheet_name, sheet_label in SIPRI_SHEETS.items():
        # Try exact match then fuzzy
        ws = None
        for s in wb.sheetnames:
            if sheet_name.lower() in s.lower() or s.lower() in sheet_name.lower():
                ws = wb[s]; break
        if ws is None:
            log(f"  Sheet '{sheet_name}' not found, available: {wb.sheetnames}"); continue
        k, d, v = parse_sipri_sheet(ws, sheet_label)
        log(f"  {sheet_label}: {len(v):,} obs")
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)

    if not all_vals:
        log("  SIPRI: 0 obs"); return 0
    return save(all_keys, all_dates, all_vals, out)


# ---------------------------------------------------------------------------
# Polity 5
# ---------------------------------------------------------------------------

POLITY_SKIP = {
    "cyear", "ccode", "scode", "country", "year", "byear", "bmonth", "bday",
    "eyear", "emonth", "eday", "flag", "fragment", "prior", "emonth2", "eday2",
    "eyear2", "eseq", "post", "lead", "change", "d5", "sf", "regtrans", "p5"
}

def ingest_polity():
    """Polity 5 country-year political regime dataset (XLS format via xlrd)."""
    out = os.path.join(POLITY_OUT, "polity5.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"Polity5: already {n:,} rows"); return n

    log("Polity5: downloading p5v2018.xls...")
    data = fetch("https://www.systemicpeace.org/inscr/p5v2018.xls")
    if not data:
        return 0

    log(f"  {len(data)/1024:.0f} KB, reading XLS...")
    wb = xlrd.open_workbook(file_contents=data)
    # Use the first sheet
    ws = wb.sheets()[0]
    log(f"  Sheets: {[s.name for s in wb.sheets()]}")
    ws = wb.sheets()[0]

    # Get headers from first row
    headers = [str(ws.cell(0, c).value).strip().lower() for c in range(ws.ncols)]
    log(f"  Headers: {headers[:15]}")

    year_i   = headers.index("year") if "year" in headers else None
    ctry_i   = headers.index("scode") if "scode" in headers else (
               headers.index("ccode") if "ccode" in headers else None)
    if year_i is None:
        log("  Polity5: no year column"); return 0

    # All non-skip columns are value columns
    num_idx = [(h, i) for i, h in enumerate(headers) if h not in POLITY_SKIP and i not in [year_i, ctry_i]]
    log(f"  {ws.nrows-1} rows, value cols: {[h for h,_ in num_idx[:10]]}...")

    keys, dates, vals = [], [], []
    for r in range(1, ws.nrows):
        try:
            yr = int(ws.cell(r, year_i).value)
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue
        ctry = str(ws.cell(r, ctry_i).value).strip() if ctry_i is not None else ""
        for col, ci in num_idx:
            cell = ws.cell(r, ci)
            if cell.value is None or str(cell.value).strip() in ("", "-99", "-88", "-77", "-66"):
                continue
            try:
                v = float(cell.value)
                if v not in (-99.0, -88.0, -77.0, -66.0):
                    key = f"POLITY:{col}:{ctry}" if ctry else f"POLITY:{col}"
                    keys.append(key); dates.append(obs_d); vals.append(v)
            except (TypeError, ValueError):
                pass
    return save(keys, dates, vals, out)


def ingest_mepv():
    """Major Episodes of Political Violence (MEPV) — country-year."""
    out = os.path.join(POLITY_OUT, "mepv.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"MEPV: already {n:,} rows"); return n

    log("MEPV: downloading MEPVv2018.xls...")
    data = fetch("https://www.systemicpeace.org/inscr/MEPVv2018.xls")
    if not data:
        return 0

    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheets()[0]
    headers = [str(ws.cell(0, c).value).strip().lower() for c in range(ws.ncols)]
    log(f"  Headers: {headers[:15]}")

    year_i = headers.index("year") if "year" in headers else None
    ctry_i = headers.index("scode") if "scode" in headers else (
             headers.index("ccode") if "ccode" in headers else
             headers.index("country") if "country" in headers else None)
    if year_i is None:
        log("  MEPV: no year column"); return 0

    skip = {"cyear", "ccode", "scode", "country", "year", "version"}
    num_idx = [(h, i) for i, h in enumerate(headers) if h not in skip and i not in [year_i, ctry_i]]
    log(f"  {ws.nrows-1} rows, value cols: {[h for h,_ in num_idx[:10]]}...")

    keys, dates, vals = [], [], []
    for r in range(1, ws.nrows):
        try:
            yr = int(ws.cell(r, year_i).value)
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue
        ctry = str(ws.cell(r, ctry_i).value).strip() if ctry_i is not None else ""
        for col, ci in num_idx:
            cell = ws.cell(r, ci)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            try:
                v = float(cell.value)
                key = f"MEPV:{col}:{ctry}" if ctry else f"MEPV:{col}"
                keys.append(key); dates.append(obs_d); vals.append(v)
            except (TypeError, ValueError):
                pass
    return save(keys, dates, vals, out)


def main():
    os.makedirs(SIPRI_OUT, exist_ok=True)
    os.makedirs(POLITY_OUT, exist_ok=True)
    total = 0

    log("=== SIPRI Military Expenditure ===")
    total += ingest_sipri()
    time.sleep(RATE)

    log("=== Polity 5 / CSP ===")
    total += ingest_polity()
    time.sleep(RATE)
    total += ingest_mepv()

    log(f"=== GRAND TOTAL: {total:,} SIPRI+Polity observations ===")


if __name__ == "__main__":
    main()
