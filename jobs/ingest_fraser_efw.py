#!/usr/bin/env python3
"""Fraser Institute — Economic Freedom of the World (EFW), 1970–2022.

License: Creative Commons Attribution 4.0 (Fraser Institute data policy)
Source: https://www.fraserinstitute.org/economic-freedom/dataset
No API key required (direct download from EFW data portal).

Coverage:
  * 165 jurisdictions, 1970–2022 (not every year for early periods)
  * Overall Economic Freedom score (0–10, higher = more free)
  * Area 1: Size of Government
  * Area 2: Legal System and Property Rights
  * Area 3: Sound Money
  * Area 4: Freedom to Trade Internationally
  * Area 5: Regulation
  * ~40 sub-components and sub-sub-components
  * Summary ratings + component ratings

Run: python jobs/ingest_fraser_efw.py
"""
from __future__ import annotations
import datetime as dt, io, os, re, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
OUT  = os.path.join(ROOT, "data", "clean_full", "fraser_efw")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# Fraser Institute hosts the data on their own CDN + EFW data portal
# Try multiple URLs in order
URLS = [
    # Direct 2024 edition link (covers 1970–2022)
    "https://www.fraserinstitute.org/sites/default/files/economic-freedom-of-the-world-2024-dataset.xlsx",
    "https://www.fraserinstitute.org/sites/default/files/economic-freedom-of-the-world-2023-dataset.xlsx",
    "https://efwdata.com/sites/default/files/Economic%20Freedom%20of%20the%20World%202024%20Dataset.xlsx",
    # Fallback: EFW data portal direct download
    "https://efwdata.com/grid/downloadEFWdata?type=excel",
    # Mirror on GitHub (if above fail)
    "https://raw.githubusercontent.com/FraseInstitute/efw-data/main/efw_data.xlsx",
]

# Which columns to try to identify year and country
COUNTRY_COLS = ("countries", "country", "economy", "jurisdiction", "name", "entity")
YEAR_COLS    = ("year", "yr", "year_of_rating", "period")
ISO_COLS     = ("iso_code", "iso", "country_code", "code", "iso3")


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii','replace').decode()}", flush=True)


def try_download(url: str) -> bytes | None:
    try:
        r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 10_000:
            log(f"  OK: {len(r.content):,} bytes from {url[-70:]}")
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_efw_xlsx(content: bytes) -> list[tuple[str, dt.date, float]]:
    """Parse Fraser EFW XLSX. Long format: country, year, area1, area2, ..."""
    import openpyxl
    results = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        log(f"  Sheets: {wb.sheetnames}")

        # Try each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 5:
                continue

            # Find header row (look for 'year' or 'country')
            header_idx = None
            for i, row in enumerate(rows[:10]):
                cells = [str(c).strip().lower() if c is not None else "" for c in row]
                if any(c in cells for c in ("year", "countries", "country", "economy")):
                    header_idx = i
                    break

            if header_idx is None:
                log(f"  No header in sheet '{sheet_name}'"); continue

            headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
            log(f"  Sheet '{sheet_name}': {len(rows)} rows, headers: {headers[:10]}")

            # Find key columns
            year_col = next((j for j, h in enumerate(headers) if h in YEAR_COLS), None)
            country_col = next((j for j, h in enumerate(headers) if h in COUNTRY_COLS), None)
            iso_col = next((j for j, h in enumerate(headers) if h in ISO_COLS), None)

            if year_col is None or country_col is None:
                log(f"  Missing year/country columns, headers: {headers[:15]}"); continue

            # Identify numeric data columns (all except year/country/iso/rank/quartile/etc.)
            skip_cols = {year_col, country_col}
            if iso_col is not None:
                skip_cols.add(iso_col)
            # Also skip rank, quartile, decile, etc.
            for j, h in enumerate(headers):
                if h in ("rank", "quartile", "decile", "quintile", "tercile",
                          "percentile", "status", "iso_num", "continent"):
                    skip_cols.add(j)

            data_cols = [(j, headers[j]) for j in range(len(headers))
                         if j not in skip_cols and headers[j]]

            log(f"  {len(data_cols)} data columns from row {header_idx+1}")

            for row in rows[header_idx + 1:]:
                if not row:
                    continue
                # Extract year
                yr_raw = row[year_col] if year_col < len(row) else None
                if yr_raw is None:
                    continue
                try:
                    yr = int(float(str(yr_raw).strip()))
                    if not (1960 <= yr <= 2030):
                        continue
                except (ValueError, TypeError):
                    continue

                # Extract country/jurisdiction
                country_raw = row[country_col] if country_col < len(row) else None
                if country_raw is None or str(country_raw).strip() in ("", "nan", "None"):
                    continue
                country = str(country_raw).strip()[:60]
                if not country or country.lower() in ("country", "economy", "countries"):
                    continue

                # Optional ISO code for series key
                iso = ""
                if iso_col is not None and iso_col < len(row):
                    iso_raw = row[iso_col]
                    if iso_raw is not None:
                        iso = str(iso_raw).strip().upper()[:10]

                obs_date = dt.date(yr, 12, 31)
                entity = iso if iso else country

                for col_idx, col_name in data_cols:
                    if col_idx >= len(row):
                        continue
                    v_raw = row[col_idx]
                    if v_raw is None or str(v_raw).strip() in ("", "N/A", "NA", "-", ".."):
                        continue
                    try:
                        v = float(str(v_raw).strip())
                        if v != v:
                            continue
                    except (ValueError, TypeError):
                        continue

                    # Clean column name for series key
                    col_clean = re.sub(r"[^a-z0-9_]", "_", col_name)[:40].strip("_")
                    series_key = f"EFW:{col_clean}:{entity}"
                    results.append((series_key, obs_date, v))

    except Exception as e:
        log(f"  XLSX parse error: {e}")
        import traceback; traceback.print_exc()
    return results


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "fraser_efw.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    content = None
    for url in URLS:
        log(f"Trying {url[-80:]}")
        content = try_download(url)
        if content:
            break

    if not content:
        log("All URLs failed — check Fraser Institute download page"); return

    log(f"Parsing {len(content):,} bytes...")
    results = parse_efw_xlsx(content)
    log(f"Parsed {len(results):,} raw observations")

    if not results:
        log("0 observations — check parsing logic"); return

    # Deduplicate
    all_keys, all_dates, all_vals = [], [], []
    seen: set[tuple] = set()
    for key, d, v in results:
        tok = (key, d)
        if tok not in seen:
            seen.add(tok)
            all_keys.append(key)
            all_dates.append(d)
            all_vals.append(v)

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} Fraser EFW observations ({len(set(all_keys))} series)")


if __name__ == "__main__":
    main()
