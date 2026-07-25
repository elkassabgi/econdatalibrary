#!/usr/bin/env python3
"""Global Carbon Budget (GCB) 2024 ingest.

Sources:
  Global Carbon Budget spreadsheet:  https://globalcarbonbudget.org/download/1725/
  National fossil CO2 emissions:     https://globalcarbonbudget.org/download/1728/
  National land-use change:          https://globalcarbonbudget.org/download/1731/
  GCP fossil flat CSV (Zenodo):      https://zenodo.org/records/14106218 -> GCB2024v18_MtCO2_flat.csv

Output: data/clean_full/gcb/gcb_budget.parquet
         data/clean_full/gcb/gcb_country.parquet

Series key format:
  GCB:global:{var}              – global budget aggregates (GtC/yr), 1959-2023
  GCB:historical:{var}          – historical budget (GtC/yr), 1750-2023
  GCB:fossil_cat:{var}          – global fossil by fuel category (MtC/yr), 1850-2023
  GCB:ocean:{model}             – ocean sink by model (GtC/yr)
  GCB:land:{model}              – terrestrial sink by DGVM (GtC/yr)
  GCB:cement_carb:{model}       – cement carbonation by model (GtC/yr)
  GCB:territorial:{country}     – national territorial emissions (MtC/yr), 1850-2023
  GCB:consumption:{country}     – national consumption emissions (MtC/yr), 1990-2022
  GCB:transfers:{country}       – national emissions transfers (MtC/yr)
  GCB:luc_BLUE:{country}        – LUC BLUE model by country (TgC/yr), 1850-2023
  GCB:luc_HNC:{country}         – LUC H&C2023 model by country
  GCB:luc_OSCAR:{country}       – LUC OSCAR model by country
  GCB:luc_LUCE:{country}        – LUC LUCE model by country
  GCB:fossil_flat:{country}     – GCP flat CSV country fossil CO2 (MtCO2), redundant check

Run: python jobs/ingest_gcb.py
"""
from __future__ import annotations
import csv
import datetime as dt
import io
import os
import time

import requests
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
OUT  = os.path.join(ROOT, "data", "clean_full", "gcb")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

GCB_BUDGET_URL   = "https://globalcarbonbudget.org/download/1725/"
GCB_FOSSIL_URL   = "https://globalcarbonbudget.org/download/1728/"
GCB_LUC_URL      = "https://globalcarbonbudget.org/download/1731/"
GCB_FLAT_URL     = "https://zenodo.org/records/14106218/files/GCB2024v18_MtCO2_flat.csv"


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url: str, retries: int = 3) -> bytes | None:
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=UA, timeout=120, allow_redirects=True)
            if r.status_code == 200:
                return r.content
            log(f"  HTTP {r.status_code}: {url[-80:]}")
        except Exception as e:
            log(f"  ERR attempt {attempt+1}: {e}")
        time.sleep(5 * (attempt + 1))
    return None


def save(keys, dates, vals, path):
    tbl = pa.table({
        "series_key": pa.array(keys,  pa.string()),
        "obs_date":   pa.array(dates, pa.date32()),
        "value":      pa.array(vals,  pa.float64()),
    })
    pq.write_table(tbl, path, compression="zstd")
    n = pq.read_metadata(path).num_rows
    log(f"  -> {n:,} obs saved to {os.path.basename(path)}")
    return n


def is_year(v):
    """Return True if v looks like a calendar year."""
    try:
        return 1700 <= int(v) <= 2100
    except (TypeError, ValueError):
        return False


def parse_global_sheet(ws, prefix: str, skip_cols=None) -> tuple[list, list, list]:
    """Parse a global/aggregate sheet where year is in col 0 (index 0).
    Finds header row (first row where col 0 is a string not a year), then data rows."""
    skip_cols = skip_cols or set()
    rows = list(ws.iter_rows(values_only=True))

    # Find the data start (first row where col 0 is a valid year)
    data_start = None
    for i, row in enumerate(rows):
        if is_year(row[0]):
            data_start = i
            break
    if data_start is None:
        log(f"  {prefix}: no data rows found"); return [], [], []

    # Header is the row just before data_start
    header_row = rows[data_start - 1] if data_start > 0 else None
    if header_row is None:
        log(f"  {prefix}: no header row"); return [], [], []

    # Build column map
    col_map = {}  # col_idx -> label
    for ci, cell in enumerate(header_row):
        if ci == 0:
            continue
        if cell is None:
            continue
        label = str(cell).strip()
        if not label or label in skip_cols:
            continue
        col_map[ci] = label.replace(" ", "_").replace("/", "_").replace(".", "_")

    log(f"  {prefix}: {len(col_map)} columns, first={rows[data_start][0]}, last={rows[-1][0]}")

    keys, dates, vals = [], [], []
    for row in rows[data_start:]:
        if not is_year(row[0]):
            continue
        yr = int(row[0])
        obs_d = dt.date(yr, 12, 31)
        for ci, label in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None or str(cell).strip() in ("", "N/A", "n/a", "NA"):
                continue
            try:
                v = float(cell)
                if v != v:  # NaN check
                    continue
                keys.append(f"{prefix}:{label}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass
    return keys, dates, vals


def parse_wide_country_sheet(ws, prefix: str, country_row_idx: int = None) -> tuple[list, list, list]:
    """Parse wide-format country sheet: year in col 0, countries in header row.

    Auto-detects the country header row as the last non-year row immediately
    before the first data row (where col 0 is a valid year).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find data start
    data_start = None
    for i, row in enumerate(rows):
        if is_year(row[0]):
            data_start = i
            break
    if data_start is None:
        log(f"  {prefix}: no data rows"); return [], [], []

    # Find the country name row: scan backwards from data_start for the
    # last row that has strings in cols 1+
    ctry_row = None
    if country_row_idx is not None:
        ctry_row = rows[country_row_idx]
    else:
        for i in range(data_start - 1, -1, -1):
            r = rows[i]
            names = [str(c).strip() for c in r[1:] if c is not None and str(c).strip()]
            if len(names) >= 3:
                ctry_row = r
                break

    if ctry_row is None:
        log(f"  {prefix}: no country header found"); return [], [], []

    # Build col → country map, using proper-case names
    col_map = {}
    for ci, cell in enumerate(ctry_row):
        if ci == 0:
            continue
        if cell is None:
            continue
        name = str(cell).strip()
        if not name or name.upper() == name and len(name) > 2 and name not in ("USA", "UK"):
            # Skip ALL-CAPS rows (those are the duplicate uppercase header rows)
            # Heuristic: if > 50% of non-empty cells are all-caps, skip this row
            pass
        col_map[ci] = name

    # Validate: if mostly uppercase, this might be the uppercase header; skip
    upper_count = sum(1 for n in col_map.values() if n == n.upper() and n.isalpha())
    if col_map and upper_count / len(col_map) > 0.6:
        # All-caps row — try the proper-case row (one row later if it exists)
        next_idx = rows.index(ctry_row) + 1 if ctry_row in rows else -1
        if 0 <= next_idx < data_start:
            ctry_row2 = rows[next_idx]
            names2 = [str(c).strip() for c in ctry_row2[1:] if c is not None]
            if names2:
                col_map = {}
                for ci, cell in enumerate(ctry_row2):
                    if ci == 0 or cell is None:
                        continue
                    name = str(cell).strip()
                    if name:
                        col_map[ci] = name

    log(f"  {prefix}: {len(col_map)} countries, rows {data_start} to {len(rows)-1}")

    keys, dates, vals = [], [], []
    for row in rows[data_start:]:
        if not is_year(row[0]):
            continue
        yr = int(row[0])
        obs_d = dt.date(yr, 12, 31)
        for ci, country in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None or str(cell).strip() in ("", "N/A"):
                continue
            try:
                v = float(cell)
                if v != v:
                    continue
                keys.append(f"{prefix}:{country}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass
    return keys, dates, vals


# ---------------------------------------------------------------------------
# Global Carbon Budget XLSX
# ---------------------------------------------------------------------------

# sheet name → (prefix, description)
GLOBAL_SHEETS = [
    ("Global Carbon Budget",       "GCB:global",       "global budget aggregates 1959-2023"),
    ("Historical Budget",          "GCB:historical",   "historical budget 1750-2023"),
    ("Fossil Emissions by Category","GCB:fossil_cat",  "global fossil by category 1850-2023"),
    ("Ocean Sink",                  "GCB:ocean",        "ocean sink model ensemble"),
    ("Terrestrial Sink",            "GCB:land",         "terrestrial sink DGVM ensemble"),
    ("Cement Carbonation Sink",     "GCB:cement_carb",  "cement carbonation models"),
    ("Land-Use Change Emissions",   "GCB:luc_global",   "global LUC emissions model ensemble"),
]


def ingest_gcb_budget():
    out = os.path.join(OUT, "gcb_budget.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"gcb_budget: already {n:,} rows"); return n

    log("Downloading Global Carbon Budget XLSX...")
    data = fetch(GCB_BUDGET_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []
    for sname, prefix, desc in GLOBAL_SHEETS:
        if sname not in wb.sheetnames:
            log(f"  Sheet '{sname}' not found"); continue
        ws = wb[sname]
        k, d, v = parse_global_sheet(ws, prefix)
        log(f"  {sname}: {len(v):,} obs  ({desc})")
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)

    if not all_vals:
        log("  0 obs total"); return 0
    return save(all_keys, all_dates, all_vals, out)


# ---------------------------------------------------------------------------
# National fossil CO2 XLSX (Territorial + Consumption + Transfers)
# ---------------------------------------------------------------------------

FOSSIL_SHEETS = [
    ("Territorial Emissions", "GCB:territorial"),
    ("Consumption Emissions", "GCB:consumption"),
    ("Emissions Transfers",   "GCB:transfers"),
]


def ingest_gcb_national_fossil():
    out = os.path.join(OUT, "gcb_national_fossil.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"gcb_national_fossil: already {n:,} rows"); return n

    log("Downloading National Fossil Emissions XLSX...")
    data = fetch(GCB_FOSSIL_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []
    for sname, prefix in FOSSIL_SHEETS:
        if sname not in wb.sheetnames:
            log(f"  Sheet '{sname}' not found"); continue
        ws = wb[sname]
        k, d, v = parse_wide_country_sheet(ws, prefix)
        log(f"  {sname}: {len(v):,} obs")
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)

    if not all_vals:
        log("  0 obs total"); return 0
    return save(all_keys, all_dates, all_vals, out)


# ---------------------------------------------------------------------------
# National LUC XLSX (BLUE, H&C2023, OSCAR, LUCE — all country-level)
# ---------------------------------------------------------------------------

LUC_SHEETS = [
    ("BLUE",   "GCB:luc_BLUE"),
    ("H&C2023","GCB:luc_HNC"),
    ("OSCAR",  "GCB:luc_OSCAR"),
    ("LUCE",   "GCB:luc_LUCE"),
]


def ingest_gcb_luc():
    out = os.path.join(OUT, "gcb_luc.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"gcb_luc: already {n:,} rows"); return n

    log("Downloading National LUC Emissions XLSX...")
    data = fetch(GCB_LUC_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []
    for sname, prefix in LUC_SHEETS:
        if sname not in wb.sheetnames:
            log(f"  Sheet '{sname}' not found"); continue
        ws = wb[sname]
        k, d, v = parse_wide_country_sheet(ws, prefix)
        log(f"  {sname}: {len(v):,} obs")
        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)

    if not all_vals:
        log("  0 obs total"); return 0
    return save(all_keys, all_dates, all_vals, out)


# ---------------------------------------------------------------------------
# GCP Fossil flat CSV from Zenodo (country-level MtCO2, long format)
# ---------------------------------------------------------------------------

def ingest_gcb_flat():
    """GCB2024v18_MtCO2_flat.csv from Zenodo 14106218 — country × year × fuel flat CSV."""
    out = os.path.join(OUT, "gcb_fossil_flat.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows; log(f"gcb_fossil_flat: already {n:,} rows"); return n

    log("Downloading GCB2024 fossil flat CSV from Zenodo...")
    data = fetch(GCB_FLAT_URL)
    if not data:
        log("  FAILED"); return 0
    log(f"  {len(data)//1024} KB")

    text = data.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames
    log(f"  Columns: {headers[:15]}")

    # Identify year column and country/name column
    year_col   = next((h for h in (headers or []) if h.lower() in ("year", "yr")), None)
    name_col   = next((h for h in (headers or []) if h.lower() in ("country", "name", "nation")), None)
    # Value columns: all numeric-named ones (Total, Coal, Oil, Gas, Cement, Flaring, Other, Per.Capita, etc.)
    skip_cols  = {year_col, name_col, "isocode", "iso3", "Country.Code",
                  "country_code", "country_id", "ISO3166_1_Alpha_3"}
    val_cols   = [h for h in (headers or []) if h not in skip_cols and h is not None and h.strip()]

    if year_col is None:
        log("  No year column found"); return 0

    keys, dates, vals = [], [], []
    for row in reader:
        try:
            yr = int(float(row[year_col]))
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError, KeyError):
            continue
        country = str(row.get(name_col, "") or "").strip() if name_col else ""
        for vc in val_cols:
            raw = str(row.get(vc, "") or "").strip()
            if not raw or raw in ("", "NA", "N/A", "NaN", "nan", "None"):
                continue
            try:
                v = float(raw)
                if v != v:
                    continue
                label = vc.replace(".", "_").replace(" ", "_")
                k = f"GCB:fossil_flat:{label}:{country}" if country else f"GCB:fossil_flat:{label}"
                keys.append(k); dates.append(obs_d); vals.append(v)
            except (TypeError, ValueError):
                pass

    if not keys:
        log("  0 obs total"); return 0
    return save(keys, dates, vals, out)


def main():
    os.makedirs(OUT, exist_ok=True)
    log("=== Global Carbon Budget 2024 Ingest ===")
    total = 0

    log("--- Global budget aggregates ---")
    total += ingest_gcb_budget()
    time.sleep(1)

    log("--- National fossil emissions ---")
    total += ingest_gcb_national_fossil()
    time.sleep(1)

    log("--- National LUC emissions ---")
    total += ingest_gcb_luc()
    time.sleep(1)

    log("--- GCP fossil flat CSV (Zenodo) ---")
    total += ingest_gcb_flat()

    log(f"=== GCB TOTAL: {total:,} observations ===")


if __name__ == "__main__":
    main()
