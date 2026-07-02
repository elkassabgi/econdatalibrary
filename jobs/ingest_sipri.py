#!/usr/bin/env python3
"""SIPRI Military Expenditure Database — 173 countries, 1949–present.

License: Free for non-commercial use (SIPRI Terms of Use)
Source: https://www.sipri.org/databases/milex
No API key required (direct XLSX download).

Coverage:
  * Military expenditure in constant 2022 USD (billions)
  * Military expenditure in current USD (billions)
  * Military expenditure as % of GDP
  * Military expenditure as % of government spending
  * Military expenditure per capita (USD)
  * Military expenditure in local currency (current)
  * Annual, 1949–2023, ~173 countries

Run: python jobs/ingest_sipri.py
"""
from __future__ import annotations
import datetime as dt, io, os, re, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "sipri")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# SIPRI MILEX main file — multiple sheets for different indicators
URLS = [
    "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2024.xlsx",
    "https://sipri.org/sites/default/files/SIPRI-Milex-data-1949-2024.xlsx",
    "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2023.xlsx",
]

# Sheet name → series prefix mapping
SHEET_MAP = {
    "Constant (2022) US$":     "milex_usd_const",    # constant 2022 USD billions
    "Current US$":             "milex_usd_curr",      # current USD billions
    "Share of GDP":            "milex_gdp_share",     # % of GDP
    "Share of Govt. spending": "milex_gov_share",     # % of govt spending
    "Per capita":              "milex_percap",        # USD per capita
    "Local currency":          "milex_lcu",           # local currency millions
}


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii','replace').decode()}", flush=True)


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "sipri.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    content = None
    for url in URLS:
        log(f"Trying {url}")
        try:
            r = requests.get(url, headers=UA, timeout=120)
            if r.status_code == 200 and len(r.content) > 10000:
                content = r.content
                log(f"Downloaded {len(content):,} bytes"); break
            log(f"HTTP {r.status_code}")
        except Exception as e:
            log(f"ERR: {e}")

    if not content:
        log("ERROR: Could not download SIPRI data"); return

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    log(f"Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []

    for sheet_name in wb.sheetnames:
        # Match sheet to a known series prefix
        prefix = None
        for pattern, pfx in SHEET_MAP.items():
            if pattern.lower() in sheet_name.lower() or sheet_name.lower() in pattern.lower():
                prefix = pfx
                break
        if prefix is None:
            # Try to infer from sheet name
            sn = sheet_name.lower().strip()
            if "constant" in sn:
                prefix = "milex_usd_const"
            elif "current" in sn and "us" in sn:
                prefix = "milex_usd_curr"
            elif "gdp" in sn:
                prefix = "milex_gdp_share"
            elif "gov" in sn:
                prefix = "milex_gov_share"
            elif "capita" in sn:
                prefix = "milex_percap"
            elif "local" in sn or "lcu" in sn:
                prefix = "milex_lcu"
            else:
                log(f"  Skipping sheet: {sheet_name}")
                continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Find header row with years (4-digit numbers)
        header_row_idx = None
        for i, row in enumerate(rows[:20]):
            year_count = sum(1 for c in (row or []) if c is not None and
                             re.match(r"^\d{4}$", str(c).strip()) and
                             1940 <= int(str(c).strip()) <= 2030)
            if year_count >= 5:
                header_row_idx = i
                break

        if header_row_idx is None:
            log(f"  Sheet '{sheet_name}': no year header found"); continue

        headers = list(rows[header_row_idx])
        # Build year index
        year_cols = {}
        for j, h in enumerate(headers):
            s = str(h).strip() if h is not None else ""
            if re.match(r"^\d{4}$", s) and 1940 <= int(s) <= 2030:
                year_cols[j] = int(s)

        # Country name is usually column 0 or 1 (SIPRI uses country names, not ISO codes)
        # First column might be region, second might be country
        country_col = 0
        # Check if col 0 looks like country names
        sample_vals = [rows[header_row_idx+k][0] for k in range(1, min(6, len(rows)-header_row_idx))
                       if rows[header_row_idx+k]]
        # If first few values are all None or short, try col 1
        non_none = [v for v in sample_vals if v is not None and len(str(v)) > 2]
        if len(non_none) < 2:
            country_col = 1

        count = 0
        for row in rows[header_row_idx+1:]:
            if not row:
                continue
            country_raw = row[country_col] if country_col < len(row) else None
            if country_raw is None:
                continue
            country = str(country_raw).strip()
            # Skip header-like rows, totals, notes
            if not country or country.lower() in ("country", "region", "area", "total", "notes", "note"):
                continue
            if len(country) > 60:  # skip very long note-like strings
                continue
            # Normalize to ~20 char key
            country_key = re.sub(r"[^a-zA-Z0-9 ]", "", country).strip().replace(" ", "_")[:30]
            if not country_key:
                continue

            for col_idx, yr in year_cols.items():
                if col_idx >= len(row):
                    continue
                v_raw = row[col_idx]
                if v_raw is None:
                    continue
                s = str(v_raw).strip()
                if s in ("", "...", "xxx", "n/a", "N/A", "—", "-"):
                    continue
                try:
                    v = float(s.replace(",", ""))
                    if v != v:
                        continue
                except (ValueError, TypeError):
                    continue
                all_keys.append(f"{prefix}:{country_key}")
                all_dates.append(dt.date(yr, 12, 31))
                all_vals.append(v)
                count += 1

        log(f"  Sheet '{sheet_name}' ({prefix}): {count:,} obs")

    if not all_vals:
        log("0 observations"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} SIPRI MILEX observations ({len(set(all_keys))} series)")


if __name__ == "__main__":
    main()
