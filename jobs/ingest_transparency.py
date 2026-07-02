#!/usr/bin/env python3
"""Transparency International — Corruption Perceptions Index (CPI), 180 countries.

License: CC BY-ND 4.0 (TI) / CC BY 4.0 (OWID)
Source: https://www.transparency.org/en/cpi
        via OWID: https://ourworldindata.org/grapher/ti-corruption-perception-index.csv

Coverage:
  * CPI Score (0=highly corrupt, 100=very clean), 2012–2024
  * ~180 countries

series_key: TI_CPI:cpi_score:{iso3}  e.g. TI_CPI:cpi_score:AFG

Output: data/clean_full/transparency_ti/transparency_ti.parquet
Run: python jobs/ingest_transparency.py
"""
from __future__ import annotations
import csv, datetime as dt, io, os, time
import requests
import pyarrow as pa, pyarrow.parquet as pq

ROOT = r"D:/research/econfindatalibrary"
OUT  = os.path.join(ROOT, "data", "clean_full", "transparency_ti")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# OWID mirrors TI CPI — confirmed working 2025-06
OWID_URLS = [
    "https://ourworldindata.org/grapher/ti-corruption-perception-index.csv",
]

# TI CDN (often 403, kept as fallback)
TI_URLS = [
    "https://files.transparencycdn.org/images/CPI2024-Results-and-trends.xlsx",
    "https://files.transparencycdn.org/images/CPI2023-Results-and-trends.xlsx",
    "https://files.transparencycdn.org/images/CPI2022-Results-and-trends.xlsx",
]


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii','replace').decode()}", flush=True)


def fetch(url: str, timeout: int = 60) -> bytes | None:
    try:
        r = requests.get(url, headers=UA, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 200:
            log(f"  OK {len(r.content):,} bytes: {url[-70:]}")
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_owid_csv(data: bytes) -> tuple[list, list, list]:
    """Parse OWID grapher CSV: Entity, Code, Year, Corruption Perceptions Index, ..."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    log(f"  Columns: {headers}")

    # Find columns
    code_col  = next((h for h in headers if h.strip() in ("Code", "iso3", "ISO3", "code")), None)
    entity_col = next((h for h in headers if h.strip().lower() in ("entity", "country", "name")), None)
    year_col  = next((h for h in headers if h.strip().lower() in ("year", "yr")), None)
    score_col = next((h for h in headers if "corruption" in h.lower() or "cpi" in h.lower()
                      or "perception" in h.lower()), None)

    if not year_col or not score_col:
        log(f"  Missing year or score column")
        return [], [], []

    keys, dates, vals = [], [], []
    for row in reader:
        iso3 = (row.get(code_col) or "").strip() if code_col else ""
        if not iso3 and entity_col:
            iso3 = (row.get(entity_col) or "").strip().replace(" ", "_")[:30]
        if not iso3:
            continue

        yr_raw = (row.get(year_col) or "").strip()
        try:
            yr = int(float(yr_raw))
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue

        v_raw = (row.get(score_col) or "").strip()
        if not v_raw or v_raw in ("", "NA", "N/A", "nan"):
            continue
        try:
            v = float(v_raw)
            if v != v:
                continue
            keys.append(f"TI_CPI:cpi_score:{iso3}")
            dates.append(obs_d)
            vals.append(v)
        except (ValueError, TypeError):
            pass

    log(f"  Parsed {len(vals):,} obs")
    return keys, dates, vals


def parse_cpi_xlsx(content: bytes) -> tuple[list, list, list]:
    """Parse TI CPI Excel file — fallback."""
    import openpyxl, re
    keys, dates, vals = [], [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue
            # Find header row
            hdr_idx = None
            for i, row in enumerate(rows[:10]):
                vals_r = [str(v).strip().lower() for v in (row or []) if v is not None]
                if any(v in ("country", "country/territory", "iso3") for v in vals_r):
                    hdr_idx = i; break
            if hdr_idx is None:
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[hdr_idx]]
            iso_col = next((i for i, h in enumerate(headers)
                            if h.lower() in ("iso3", "iso", "country code", "code")), None)
            ctry_col = next((i for i, h in enumerate(headers)
                             if h.lower() in ("country", "country/territory")), None)
            id_col = iso_col if iso_col is not None else ctry_col
            # Score columns by year
            score_cols = []
            for i, h in enumerate(headers):
                m = re.search(r"\b(20\d{2})\b", h)
                if m and "score" in h.lower():
                    score_cols.append((i, int(m.group(0))))
                elif h.lower().strip() in ("score", "cpi score", "cpi_score"):
                    ym = re.search(r"\b(20\d{2})\b", sheet_name)
                    if ym:
                        score_cols.append((i, int(ym.group(0))))
            if not score_cols or id_col is None:
                continue
            for row in rows[hdr_idx+1:]:
                if not row:
                    continue
                cid_raw = row[id_col] if id_col < len(row) else None
                if not cid_raw:
                    continue
                cid = str(cid_raw).strip()[:30]
                if not cid or cid.lower() in ("nan", "none"):
                    continue
                for ci, yr in score_cols:
                    if ci >= len(row) or row[ci] is None:
                        continue
                    try:
                        v = float(str(row[ci]).replace(",", ""))
                        if v != v:
                            continue
                        keys.append(f"TI_CPI:cpi_score:{cid}")
                        dates.append(dt.date(yr, 12, 31))
                        vals.append(v)
                    except (ValueError, TypeError):
                        pass
    except Exception as e:
        log(f"  XLSX parse error: {e}")
    return keys, dates, vals


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "transparency_ti.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    log("=== Transparency International CPI Ingest ===")
    all_keys, all_dates, all_vals = [], [], []
    seen: set[tuple] = set()

    # Primary: OWID CSV (confirmed working, 2012-2024)
    for url in OWID_URLS:
        log(f"Trying OWID CSV: {url[-70:]}")
        data = fetch(url)
        if data:
            k, d, v = parse_owid_csv(data)
            for ki, di, vi in zip(k, d, v):
                tok = (ki, di)
                if tok not in seen:
                    seen.add(tok)
                    all_keys.append(ki); all_dates.append(di); all_vals.append(vi)
        time.sleep(0.5)

    # Fallback: TI CDN Excel
    if not all_vals:
        log("OWID failed, trying TI CDN...")
        for url in TI_URLS:
            log(f"Trying XLSX: {url[-70:]}")
            data = fetch(url)
            if data:
                k, d, v = parse_cpi_xlsx(data)
                for ki, di, vi in zip(k, d, v):
                    tok = (ki, di)
                    if tok not in seen:
                        seen.add(tok)
                        all_keys.append(ki); all_dates.append(di); all_vals.append(vi)
            time.sleep(0.5)

    if not all_vals:
        log("0 observations — all sources failed"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"=== TI CPI DONE: {n:,} obs ({len(set(all_keys))} series) ===")


if __name__ == "__main__":
    main()
