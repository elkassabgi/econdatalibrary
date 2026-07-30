#!/usr/bin/env python3
"""Maddison Project Database 2023 — historical GDP per capita, ~180 countries, 1 AD–2022.

License: CC BY 4.0
Source: https://www.rug.nl/ggdc/historicaldevelopment/maddison/
No API key required (direct XLSX download).

Variables: gdppc (real GDP per capita, 2011 USD PPP), pop (population, thousands),
           cgdppc (current GDP per capita), rgdpnapc (national-accounts per capita)

Run: python jobs/ingest_maddison.py
"""
from __future__ import annotations
import datetime as dt, io, os, time
import pyarrow as pa, pyarrow.parquet as pq
import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
OUT  = os.path.join(ROOT, "data", "clean_full", "maddison")
# mpd2023 not yet on main site; use 2020 (most complete) + Dataverse 2023 as fallback
URLS = [
    "https://www.rug.nl/ggdc/historicaldevelopment/maddison/data/mpd2020.xlsx",
    "https://dataverse.nl/api/access/datafile/421302",   # Maddison 2023 on Dataverse.nl
]
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

_enc = "utf-8"


def log(m):
    try:
        print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
    except UnicodeEncodeError:
        print(f"[{time.strftime('%H:%M:%S')}] {str(m).encode('ascii', 'replace').decode()}", flush=True)


def parse_xlsx(content, log=log):
    """Parse a Maddison Project workbook into (keys, dates, values).

    PURE: no download, no skip rule, no writer, and no error policy — the caller decides
    what a parse failure means. EXTRACTED 2026-07-30 so the fetcher and this ingest share
    ONE parser; copying it would break the duplication invariant the moment either side
    was edited (same reason ingest_bis_cbs_lbs.iter_rows exists).
    """
    all_keys, all_dates, all_vals = [], [], []
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    log(f"Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        log(f"  Sheet '{sheet_name}': {len(headers)} cols, {len(rows)-1} rows")
        log(f"    Headers: {headers[:10]}")

        # Look for standard Maddison columns
        # Expected: countrycode, country, year, gdppc, pop, cgdppc, rgdpnapc
        country_col = next((i for i, h in enumerate(headers)
                            if h in ("countrycode", "iso3", "isocode", "iso")), None)
        year_col = next((i for i, h in enumerate(headers)
                         if h in ("year", "yr")), None)

        if year_col is None:
            log(f"  Skipping sheet '{sheet_name}': no year column")
            continue

        skip_cols = {country_col, year_col}
        skip_names = {"country", "countryname", "country_name", "name", "region", "note"}
        val_cols = [(i, headers[i]) for i in range(len(headers))
                    if i not in skip_cols and headers[i] and headers[i] not in skip_names
                    and headers[i] not in ("", "none")]

        sheet_count = 0
        for row in rows[1:]:
            yr_raw = row[year_col] if year_col is not None else None
            if yr_raw is None:
                continue
            try:
                yr = int(yr_raw)
                if yr < 1 or yr > 2100:
                    continue
                d = dt.date(yr, 12, 31)
            except (ValueError, TypeError, OverflowError):
                continue

            iso = (str(row[country_col]).strip()[:20] if country_col is not None
                   and row[country_col] is not None else "WLD")

            for col_idx, col_name in val_cols:
                v_raw = row[col_idx]
                if v_raw is None:
                    continue
                try:
                    v = float(v_raw)
                except (ValueError, TypeError):
                    continue
                all_keys.append(f"{col_name}:{iso}")
                all_dates.append(d)
                all_vals.append(v)
                sheet_count += 1

        log(f"  Sheet '{sheet_name}': {sheet_count:,} obs collected")
    return all_keys, all_dates, all_vals


def main():
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "maddison.parquet")
    if os.path.exists(out_path):
        n = pq.read_metadata(out_path).num_rows
        log(f"Already done: {n:,} rows"); return

    content = None
    for url in URLS:
        log(f"Trying {url}")
        try:
            r = requests.get(url, headers=UA, timeout=300, stream=True)
            if r.status_code == 200:
                content = r.content
                log(f"Downloaded {len(content):,} bytes from {url}")
                break
            else:
                log(f"HTTP {r.status_code}: {url}")
        except Exception as e:
            log(f"ERR: {e}")
    if not content:
        log("ERROR: could not download from any URL"); return

    log(f"Parsing XLSX ({len(content):,} bytes)...")

    try:
        all_keys, all_dates, all_vals = parse_xlsx(content)
    except Exception as e:
        log(f"XLSX parse error: {e}"); return

    if not all_vals:
        log("0 observations"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out_path, compression="zstd")
    n = pq.read_metadata(out_path).num_rows
    log(f"DONE: {n:,} Maddison Project observations")


if __name__ == "__main__":
    main()
