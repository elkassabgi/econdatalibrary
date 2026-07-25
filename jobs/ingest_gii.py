#!/usr/bin/env python3
"""Global Innovation Index (GII) ingest.

Source: https://www.globalinnovationindex.org/analysis-indicator
License: CC BY 3.0 IGO (WIPO / Cornell University / INSEAD)
Coverage: ~130 countries, 2011-present, 80+ innovation indicators.

Downloads GII historical data from WIPO's open data portal.
series_key: GII:{indicator}:{iso3}  e.g. GII:Score:USA

Output: data/clean_full/gii/gii.parquet
Run: python jobs/ingest_gii.py
"""
from __future__ import annotations
import csv, datetime as dt, io, os, time, zipfile
import requests, pyarrow as pa, pyarrow.parquet as pq

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # derived, never hardcoded
OUT  = os.path.join(ROOT, "data", "clean_full", "gii")
UA   = {"User-Agent": "Econ-Fin Data Library admin@hfdatalibrary.com"}

# WIPO GII data URLs
URLS = [
    # WIPO open data portal
    "https://www.wipo.int/edocs/pubdocs/en/wipo_pub_2000_2024.xlsx",
    "https://www.wipo.int/edocs/pubdocs/en/wipo_pub_2000_2023.xlsx",
    # GitHub mirror maintained by WIPO
    "https://raw.githubusercontent.com/wipo-analytics/gii-data/main/gii_results.csv",
    # Kaggle-hosted open dataset
    "https://raw.githubusercontent.com/datasets/global-innovation-index/main/data/gii.csv",
    # Our World in Data version
    "https://raw.githubusercontent.com/owid/owid-datasets/master/datasets/Global%20Innovation%20Index%20(GII)/Global%20Innovation%20Index%20(GII).csv",
]


def log(m):
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def fetch(url, timeout=120):
    try:
        r = requests.get(url, headers=UA, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 500:
            log(f"  Downloaded {len(r.content)//1024:,} KB")
            return r.content
        log(f"  HTTP {r.status_code}: {url[-70:]}")
    except Exception as e:
        log(f"  ERR: {e}")
    return None


def parse_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    log(f"  Headers ({len(headers)}): {headers[:12]}")

    iso3_col = next((h for h in headers if h.lower().strip() in
                     ("iso3", "iso", "country_code", "code", "iso_code")), None)
    year_col  = next((h for h in headers if h.lower().strip() in ("year", "edition", "yr")), None)
    ctry_col  = next((h for h in headers if h.lower().strip() in ("entity", "country", "country_name")), None)

    if not (iso3_col or ctry_col) or not year_col:
        log(f"  Missing key columns"); return [], [], []

    skip = {(iso3_col or "").lower(), (ctry_col or "").lower(),
            (year_col or "").lower(), "country", "entity", "region", "rank"}

    keys, dates, vals = [], [], []
    for row in reader:
        if iso3_col:
            iso3 = (row.get(iso3_col) or "").strip()
        else:
            iso3 = (row.get(ctry_col) or "").strip().replace(" ", "_")[:30]
        if not iso3:
            continue

        try:
            yr = int(float(row.get(year_col, "") or ""))
            obs_d = dt.date(yr, 12, 31)
        except (ValueError, TypeError):
            continue

        for col, raw in row.items():
            if col.lower().strip() in skip or not col:
                continue
            if not raw or str(raw).strip() in ("", "NA", "N/A", "nan", "#N/A", "-", ".."):
                continue
            try:
                v = float(str(raw).replace(",", ""))
                if v != v:
                    continue
                keys.append(f"GII:{col.strip()}:{iso3}")
                dates.append(obs_d)
                vals.append(v)
            except (TypeError, ValueError):
                pass
    return keys, dates, vals


def parse_xlsx(data: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")

    all_keys, all_dates, all_vals = [], [], []
    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in ("readme", "notes", "cover")):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]

        iso3_idx = next((i for i, h in enumerate(header) if h.lower() in
                         ("iso3", "iso", "country_code", "code")), None)
        year_idx = next((i for i, h in enumerate(header) if h.lower() in ("year", "edition")), None)
        if iso3_idx is None or year_idx is None:
            continue

        skip_idx = {iso3_idx, year_idx}
        for row in rows[1:]:
            if not row or row[iso3_idx] is None:
                continue
            iso3 = str(row[iso3_idx]).strip()
            if not iso3 or len(iso3) > 5:
                continue
            try:
                yr = int(row[year_idx])
                obs_d = dt.date(yr, 12, 31)
            except (TypeError, ValueError):
                continue
            for ci, (col, cell) in enumerate(zip(header, row)):
                if ci in skip_idx or not col or cell is None:
                    continue
                try:
                    v = float(cell)
                    if v != v:
                        continue
                    all_keys.append(f"GII:{col}:{iso3}")
                    all_dates.append(obs_d)
                    all_vals.append(v)
                except (TypeError, ValueError):
                    pass
    return all_keys, all_dates, all_vals


def main():
    os.makedirs(OUT, exist_ok=True)
    out = os.path.join(OUT, "gii.parquet")
    if os.path.exists(out):
        n = pq.read_metadata(out).num_rows
        log(f"GII: already {n:,} rows"); return

    log("=== Global Innovation Index Ingest ===")
    all_keys, all_dates, all_vals = [], [], []

    for url in URLS:
        log(f"Trying {url[-70:]}...")
        data = fetch(url)
        if not data:
            time.sleep(1); continue

        if data[:2] == b"PK":  # ZIP or XLSX
            try:
                k, d, v = parse_xlsx(data)
                if v:
                    all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
                    break
            except Exception as e:
                log(f"  XLSX/ZIP error: {e}")
        else:
            if data[:1] in (b"\xd0", b"\x50"):  # OLE or XLSX magic
                try:
                    k, d, v = parse_xlsx(data)
                    if v:
                        all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
                        break
                except Exception as e:
                    log(f"  XLSX error: {e}")
            else:
                k, d, v = parse_csv(data)
                if v:
                    all_keys.extend(k); all_dates.extend(d); all_vals.extend(v)
                    break
        time.sleep(1)

    if not all_vals:
        log("0 observations parsed"); return

    tbl = pa.table({
        "series_key": pa.array(all_keys,  pa.string()),
        "obs_date":   pa.array(all_dates, pa.date32()),
        "value":      pa.array(all_vals,  pa.float64()),
    })
    pq.write_table(tbl, out, compression="zstd")
    n = pq.read_metadata(out).num_rows
    log(f"=== GII DONE: {n:,} obs ===")


if __name__ == "__main__":
    main()
