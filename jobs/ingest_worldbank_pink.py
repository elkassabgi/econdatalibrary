#!/usr/bin/env python3
"""FULL-COVERAGE ingest of the World Bank "Pink Sheet" (Commodity Markets).

Two workbooks downloaded from https://www.worldbank.org/en/research/commodity-markets :
  * CMO-Historical-Data-Monthly.xlsx  -> sheets "Monthly Prices", "Monthly Indices"
  * CMO-Historical-Data-Annual.xlsx   -> sheets "Annual Prices (Nominal)",
        "Annual Indices (Nominal)", "Annual Prices (Real)", "Annual Indices (Real)"

We parse ALL commodity columns (energy, agriculture, fertilizers, metals, precious
metals) AND all the price-index columns, for every frequency / nominal-real variant.

GROUPED storage (mirrors jobs/ingest_eurostat.py + jobs/ingest_worldbank_esg.py):
ONE Parquet per SHEET (the natural "dataset"); every commodity / index column is a
series inside it -> long columns (series_key, obs_date, value). 6 sheets => 6 grouped
Parquet files for the whole source. Never one-file-per-series.

  series_key = "<freq>:<variant>:<slug>"     e.g. "m:price:crude_oil_brent",
  "m:index:energy", "a:price_real:gold", "a:index_nominal:precious_metals".
The slug is a stable URL-safe id derived from the commodity / index label.

SHEET LAYOUTS (observed June-2026 monthly / March-2026 annual vintage):
  Monthly Prices    : names row 5, units row 6, data row 7+ ; period col A = "YYYYMmm".
  Monthly Indices   : staggered header rows 6-9 (one non-blank cell per column),
                      data row 10+ ; period col A = "YYYYMmm".
  Annual Prices (*) : names row 7, units row 8, data row 9+ ; period col A = year.
  Annual Indices(*) : staggered header rows 6-9, data row 10+ ; period col A = year.
Missing values are the replacement char U+FFFD (shows as a box) or non-numeric -> skipped.

License: cc-by-4.0 (reservable id from configs/sources.yaml -> worldbank_pink).

Does NOT touch data/catalog.db or data/clean/ (per task constraints).

Run:
  python jobs/ingest_worldbank_pink.py --dry    # parse, print per-series summary, no writes
  python jobs/ingest_worldbank_pink.py          # full run -> data/clean_full/worldbank_pink/
  python jobs/ingest_worldbank_pink.py --no-download   # reuse already-downloaded xlsx
"""
from __future__ import annotations

import datetime as dt
import io
import json
import os
import re
import sys
import time
import unicodedata

import pyarrow as pa
import pyarrow.parquet as pq

ROOT = r"D:/research/econfindatalibrary"
sys.path.insert(0, ROOT)

SOURCE_ID = "worldbank_pink"
LICENSE_ID = "cc-by-4.0"
UA = "Econ-Fin Data Library admin@hfdatalibrary.com"

RAW = os.path.join(ROOT, "data", "raw", SOURCE_ID)
OUT = os.path.join(ROOT, "data", "clean_full", SOURCE_ID)
os.makedirs(RAW, exist_ok=True)
os.makedirs(OUT, exist_ok=True)

LANDING = "https://www.worldbank.org/en/research/commodity-markets"
# Current (June 2026) vintage doc id, scraped 2026-06-04. _resolve_urls() tries to
# rediscover the live links from the landing page first and only falls back here.
FALLBACK_DOC = "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026"
FALLBACK = {
    "monthly": f"https://thedocs.worldbank.org/en/doc/{FALLBACK_DOC}/related/CMO-Historical-Data-Monthly.xlsx",
    "annual": f"https://thedocs.worldbank.org/en/doc/{FALLBACK_DOC}/related/CMO-Historical-Data-Annual.xlsx",
}
_LINK_RE = {
    "monthly": re.compile(r"https://thedocs\.worldbank\.org/[^\"'\s]*CMO-Historical-Data-Monthly\.xlsx"),
    "annual": re.compile(r"https://thedocs\.worldbank\.org/[^\"'\s]*CMO-Historical-Data-Annual\.xlsx"),
}
FILES = {
    "monthly": "CMO-Historical-Data-Monthly.xlsx",
    "annual": "CMO-Historical-Data-Annual.xlsx",
}

REPL = "�"  # World Bank uses this glyph for "no observation"

# Per-sheet parse spec. header_rows are 0-INDEXED.
#   kind 'price' : flat header -> name_row, unit_row.
#   kind 'index' : staggered header -> coalesce over coalesce_rows (one label per col).
# variant feeds the series_key namespace; data_row is the 0-indexed first data row.
SHEETS = [
    # workbook, sheet name, freq, variant, kind, name_row, unit_row/coalesce_rows, data_row
    ("monthly", "Monthly Prices",           "M", "price",        "price", 4, 5, 6),
    ("monthly", "Monthly Indices",          "M", "index",        "index", None, [5, 6, 7, 8], 9),
    ("annual",  "Annual Prices (Nominal)",  "A", "price",        "price", 6, 7, 8),
    ("annual",  "Annual Indices (Nominal)", "A", "index_nominal","index", None, [5, 6, 7, 8], 9),
    ("annual",  "Annual Prices (Real)",     "A", "price_real",   "price", 6, 7, 8),
    ("annual",  "Annual Indices (Real)",    "A", "index_real",   "index", None, [5, 6, 7, 8], 9),
]


# --------------------------------------------------------------------------- net

def http_get(url: str, tries: int = 6) -> bytes:
    """GET with polite UA + exponential backoff."""
    import requests
    last = None
    for attempt in range(tries):
        try:
            r = requests.get(url, headers={"User-Agent": UA}, timeout=180)
            r.raise_for_status()
            return r.content
        except Exception as e:  # noqa: BLE001
            last = e
            wait = min(2 ** attempt, 30)
            print(f"    retry {attempt + 1}/{tries} after {wait}s ({e})", flush=True)
            time.sleep(wait)
    raise RuntimeError(f"GET failed after {tries} tries: {url} ({last})")


def resolve_urls() -> dict:
    """Rediscover the live xlsx links from the landing page; fall back to known."""
    urls = dict(FALLBACK)
    try:
        html = http_get(LANDING, tries=3).decode("utf-8", "ignore")
        for key, rx in _LINK_RE.items():
            m = rx.search(html)
            if m:
                urls[key] = m.group(0)
        print(f"[{SOURCE_ID}] resolved URLs from landing page", flush=True)
    except Exception as e:  # noqa: BLE001
        print(f"[{SOURCE_ID}] landing scrape failed ({e}); using fallback URLs", flush=True)
    return urls


def download(no_download: bool) -> dict:
    """Ensure both workbooks are present locally; return {key: path}."""
    paths = {k: os.path.join(RAW, fn) for k, fn in FILES.items()}
    if no_download and all(os.path.exists(p) for p in paths.values()):
        print(f"[{SOURCE_ID}] --no-download: reusing cached workbooks", flush=True)
        return paths
    urls = resolve_urls()
    for key, fn in FILES.items():
        p = paths[key]
        data = http_get(urls[key])
        with open(p, "wb") as f:
            f.write(data)
        print(f"[{SOURCE_ID}] downloaded {fn}: {len(data):,} bytes", flush=True)
    return paths


# ------------------------------------------------------------------------- parse

def clean_label(s) -> str | None:
    """Collapse newlines / runs of whitespace; drop the asterisk footnote markers."""
    if s is None:
        return None
    txt = str(s).replace("\n", " ").replace("\r", " ")
    txt = " ".join(txt.split())
    if not txt or txt == REPL:
        return None
    return txt


def display_name(s) -> str | None:
    """Human label with footnote asterisks stripped for the title."""
    txt = clean_label(s)
    if txt is None:
        return None
    return re.sub(r"\s*\*+\s*$", "", txt).strip() or txt


def slug(name: str) -> str:
    """Stable URL-safe id. 'Crude oil, Brent' -> 'crude_oil_brent';
    'Rice, Thai 5% ' -> 'rice_thai_5pct'; 'Base Metals (ex. iron ore)' ->
    'base_metals_ex_iron_ore'. Footnote '*'/'**' are dropped first."""
    s = re.sub(r"\*+", "", name)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower().replace("%", "pct")
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def parse_month(cell) -> dt.date | None:
    if not isinstance(cell, str):
        return None
    m = re.fullmatch(r"\s*(\d{4})M(\d{1,2})\s*", cell)
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    if 1 <= mo <= 12:
        return dt.date(y, mo, 1)
    return None


def parse_year(cell) -> dt.date | None:
    y = None
    if isinstance(cell, (int, float)) and float(cell).is_integer():
        y = int(cell)
    elif isinstance(cell, str):
        m = re.fullmatch(r"\s*(\d{4})\s*", cell)
        if m:
            y = int(m.group(1))
    if y is not None and 1900 <= y <= 2100:
        return dt.date(y, 12, 31)  # annual stamped at year-end (ESG convention)
    return None


def header_for_sheet(rows, spec):
    """Return {col_index: (slug, display_name, unit)} for one sheet."""
    _, _, _, _, kind, name_row, unit_spec, _ = spec
    out = {}
    if kind == "price":
        nr = rows[name_row]
        ur = rows[unit_spec] if unit_spec is not None and unit_spec < len(rows) else ()
        for c in range(1, len(nr)):
            nm = clean_label(nr[c])
            if not nm:
                continue
            unit = clean_label(ur[c]) if c < len(ur) else None
            out[c] = (slug(nm), display_name(nr[c]), unit)
    else:  # index: coalesce the single non-blank label down the staggered rows
        coalesce_rows = unit_spec
        maxc = max((len(rows[r]) for r in coalesce_rows if r < len(rows)), default=0)
        for c in range(1, maxc):
            label = None
            for r in coalesce_rows:
                if r < len(rows) and c < len(rows[r]):
                    lab = clean_label(rows[r][c])
                    if lab:
                        label = lab
                        break
            if label:
                out[c] = (slug(label), display_name(label), None)
    return out


def parse_sheet(ws, spec):
    """Yield (slug, display_name, unit, obs_date, value) for one sheet.

    Returns a generator; also exposes the resolved header via the .header attr set
    by the caller. Here we just stream rows.
    """
    freq = spec[2]
    data_row = spec[7]
    period_fn = parse_month if freq == "M" else parse_year

    rows = list(ws.iter_rows(values_only=True))
    header = header_for_sheet(rows, spec)

    series_keys, dates, vals, names_seen = [], [], [], {}
    units = {}
    for ridx in range(data_row, len(rows)):
        row = rows[ridx]
        if not row:
            continue
        od = period_fn(row[0])
        if od is None:
            continue
        for c, (sl, disp, unit) in header.items():
            v = row[c] if c < len(row) else None
            if not isinstance(v, (int, float)):
                continue  # missing (U+FFFD), blank, or label string -> skip
            fv = float(v)
            series_keys.append(sl)
            dates.append(od)
            vals.append(fv)
            names_seen.setdefault(sl, disp)
            if unit is not None:
                units.setdefault(sl, unit)
    return header, series_keys, dates, vals, names_seen, units


# -------------------------------------------------------------------------- main

def main() -> int:
    dry = "--dry" in sys.argv
    no_download = "--no-download" in sys.argv

    import openpyxl

    paths = download(no_download)

    grand_series = set()
    grand_obs = 0
    manifest = {
        "source_id": SOURCE_ID,
        "license": LICENSE_ID,
        "landing": LANDING,
        "downloaded": {k: os.path.basename(v) for k, v in paths.items()},
        "doc_id_fallback": FALLBACK_DOC,
        "datasets": [],
    }
    vintage = {}

    # cache loaded workbooks (each opened once)
    wbs = {}
    for key, p in paths.items():
        wbs[key] = openpyxl.load_workbook(io.BytesIO(open(p, "rb").read()),
                                          read_only=True, data_only=True)

    for spec in SHEETS:
        wb_key, sheet, freq, variant, kind, _, _, _ = spec
        wb = wbs[wb_key]
        if sheet not in wb.sheetnames:
            print(f"[{SOURCE_ID}] WARNING sheet {sheet!r} missing in {wb_key}; skipping", flush=True)
            continue
        ws = wb[sheet]

        # capture the "Updated on ..." banner (row 4) for provenance
        try:
            banner = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 3 and row and isinstance(row[0], str):
                    banner = row[0].strip()
                if i >= 3:
                    break
            vintage[sheet] = banner
        except Exception:  # noqa: BLE001
            pass
        # re-open ws (iter_rows in read_only is one-shot)
        ws = wb[sheet]

        header, series_keys, dates, vals, names_seen, units = parse_sheet(ws, spec)
        n_series = len(set(series_keys))
        n_obs = len(series_keys)

        # grouped dataset id: one file per sheet
        ds_id = f"{freq.lower()}_{variant}"  # e.g. m_price, a_index_real
        out_name = ds_id + ".parquet"

        if dry:
            sample = (series_keys[0], dates[0], vals[0]) if series_keys else ("-", "-", "-")
            print(f"  {sheet:26} cols={len(header):>3} series={n_series:>3} "
                  f"obs={n_obs:>7,} range=[{min(dates) if dates else '-'}..{max(dates) if dates else '-'}] "
                  f"sample={sample}", flush=True)
        else:
            if series_keys:
                # prefix each series_key with freq:variant namespace so keys are globally unique
                ns = f"{freq.lower()}:{variant}:"
                full_keys = [ns + k for k in series_keys]
                tbl = pa.table({
                    "series_key": full_keys,
                    "obs_date": pa.array(dates, type=pa.date32()),
                    "value": vals,
                })
                pq.write_table(tbl, os.path.join(OUT, out_name))
            # per-series catalog inside this dataset
            series_rows = []
            # recompute per-series start/end + obs count
            from collections import defaultdict
            cnt = defaultdict(int)
            smin = {}
            smax = {}
            for k, d in zip(series_keys, dates):
                cnt[k] += 1
                if k not in smin or d < smin[k]:
                    smin[k] = d
                if k not in smax or d > smax[k]:
                    smax[k] = d
            for k in sorted(cnt):
                series_rows.append({
                    "series_key": f"{freq.lower()}:{variant}:{k}",
                    "slug": k,
                    "name": names_seen.get(k, k),
                    "unit": units.get(k),
                    "n_obs": cnt[k],
                    "start": str(smin[k]),
                    "end": str(smax[k]),
                })
            manifest["datasets"].append({
                "dataset_id": ds_id,
                "parquet": out_name,
                "sheet": sheet,
                "workbook": FILES[wb_key],
                "frequency": freq,
                "variant": variant,
                "vintage_banner": vintage.get(sheet),
                "n_series": n_series,
                "n_obs": n_obs,
                "date_min": str(min(dates)) if dates else None,
                "date_max": str(max(dates)) if dates else None,
                "series": series_rows,
            })
            print(f"  wrote {out_name:22} series={n_series:>3} obs={n_obs:>7,} "
                  f"[{min(dates) if dates else '-'}..{max(dates) if dates else '-'}]", flush=True)

        grand_series.update(f"{freq.lower()}:{variant}:{k}" for k in set(series_keys))
        grand_obs += n_obs

    for wb in wbs.values():
        wb.close()

    if not dry:
        manifest["n_datasets"] = len(manifest["datasets"])
        manifest["n_series_total"] = len(grand_series)
        manifest["n_obs_total"] = grand_obs
        manifest["vintage"] = vintage
        with open(os.path.join(OUT, "_manifest.json"), "w", encoding="utf-8") as fh:
            json.dump(manifest, fh, indent=2)

    print(f"[{SOURCE_ID}] {'DRY' if dry else 'DONE'}: "
          f"{len(manifest['datasets']) if not dry else len(SHEETS)} datasets / "
          f"{len(grand_series)} series / {grand_obs:,} observations", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
