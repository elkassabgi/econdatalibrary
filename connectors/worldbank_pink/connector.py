"""World Bank Pink Sheet -- monthly commodity prices (CC BY 4.0).

The "Pink Sheet" is the World Bank's monthly Commodity Markets price release.
There is no JSON API: the canonical machine-readable form is a single Excel
workbook (CMO-Historical-Data-Monthly.xlsx) linked from the commodity-markets
page. We download that workbook, parse the "Monthly Prices" sheet, and emit one
monthly series per commodity.

Sheet layout (observed June 2026 vintage):
  row 1-4 : title / "Updated on <date>" banner
  row 5   : commodity names      (col A blank; "Crude oil, Brent", ...)
  row 6   : units                ("($/bbl)", "($/mt)", "($/troy oz)", ...)
  row 7+  : data; col A is the period as "YYYYMmm" (e.g. "1960M01");
            missing values are the single char U+2026 ("...") -- skipped.

The download URL embeds a per-release doc id that rotates every month, so we
scrape the current link from the commodity-markets page and fall back to a
last-known URL if scraping fails.

Series id format: worldbank_pink:<slug>   (slug derived from the commodity name).
"""
from __future__ import annotations

import datetime as dt
import io
import os
import re
import sys
import time
import unicodedata
from typing import Optional

import requests

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)
from connectors.base import Connector, SeriesMeta, Observation  # noqa: E402

UA = "Econ-Fin Data Library admin@hfdatalibrary.com"
LANDING = "https://www.worldbank.org/en/research/commodity-markets"
# Last-known direct URL (June 2026 vintage). Used only if scraping the landing
# page fails -- the live URL is discovered dynamically in _resolve_url().
FALLBACK_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)
_LINK_RE = re.compile(
    r"https://thedocs\.worldbank\.org/[^\"'\s]*CMO-Historical-Data-Monthly\.xlsx"
)

# Curated starter set: the most-watched Pink Sheet commodities. Keyed by the
# EXACT commodity name in row 5 of the "Monthly Prices" sheet (trailing spaces
# in a couple of source labels are tolerated by matching on the stripped name).
# Quality over completeness -- the parser is generic, so extending this list is
# just a matter of adding names.
STARTER = [
    "Crude oil, average",
    "Crude oil, Brent",
    "Crude oil, Dubai",
    "Crude oil, WTI",
    "Coal, Australian",
    "Natural gas, US",
    "Natural gas, Europe",
    "Liquefied natural gas, Japan",
    "Cocoa",
    "Coffee, Arabica",
    "Coffee, Robusta",
    "Palm oil",
    "Soybeans",
    "Maize",
    "Rice, Thai 5%",
    "Wheat, US HRW",
    "Sugar, world",
    "Cotton, A Index",
    "Aluminum",
    "Iron ore, cfr spot",
    "Copper",
    "Nickel",
    "Zinc",
    "Gold",
    "Platinum",
    "Silver",
]
_STARTER_SET = {s.strip().lower() for s in STARTER}


def _slug(name: str) -> str:
    """Stable, URL-safe identifier from a commodity label.

    "Crude oil, Brent" -> "crude_oil_brent"; "Rice, Thai 5% " -> "rice_thai_5".
    """
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    s = s.lower().replace("%", "pct")
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _parse_period(cell) -> Optional[dt.date]:
    """'1960M01' -> date(1960, 1, 1). Use the first day of the month."""
    if not isinstance(cell, str):
        return None
    m = re.fullmatch(r"\s*(\d{4})M(\d{1,2})\s*", cell)
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    if not (1 <= month <= 12):
        return None
    return dt.date(year, month, 1)


class WorldBankPinkConnector(Connector):
    source_id = "worldbank_pink"
    name = "World Bank Pink Sheet (commodities)"
    license_id = "cc-by-4.0"
    schedule = "0 7 5 * *"  # monthly; the Pink Sheet posts in the first week
    attribution = (
        "Source: World Bank Commodity Markets ('Pink Sheet') (CC BY 4.0)"
    )
    homepage = "https://www.worldbank.org/en/research/commodity-markets"
    SHEET = "Monthly Prices"

    # ---- networking ---------------------------------------------------------

    def _get(self, url: str, timeout: int = 90) -> requests.Response:
        last = None
        for attempt in range(4):
            try:
                r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout)
                r.raise_for_status()
                return r
            except requests.RequestException as e:
                last = e
                if attempt < 3:
                    time.sleep(2 * (attempt + 1))  # 2s, 4s, 6s backoff
        raise RuntimeError(f"GET failed after retries: {url}: {last}")

    def _resolve_url(self) -> str:
        """Find the current monthly-xlsx link; fall back to the known URL."""
        try:
            html = self._get(LANDING, timeout=60).text
            m = _LINK_RE.search(html)
            if m:
                return m.group(0)
        except Exception:
            pass
        return FALLBACK_URL

    # ---- parsing ------------------------------------------------------------

    def _load_columns(self):
        """Return (periods, [(name, unit, [values...])]) for the Monthly sheet.

        values are aligned 1:1 with periods; non-numeric / missing cells -> None.
        """
        import openpyxl

        raw = self._get(self._resolve_url()).content
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        if self.SHEET not in wb.sheetnames:
            raise RuntimeError(
                f"sheet {self.SHEET!r} not found; have {wb.sheetnames}"
            )
        ws = wb[self.SHEET]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 8:
            raise RuntimeError("Monthly Prices sheet has too few rows")
        names = rows[4]   # row 5 (0-indexed 4): commodity names
        units = rows[5]   # row 6: units
        ncol = len(names)

        # Identify the commodity columns (col index >= 1, has a name).
        cols = []  # (col_index, name, unit)
        for c in range(1, ncol):
            nm = names[c]
            if isinstance(nm, str) and nm.strip():
                unit = units[c] if c < len(units) else None
                unit = unit.strip() if isinstance(unit, str) else None
                cols.append((c, nm.strip(), unit))

        periods: list[dt.date] = []
        series_vals: dict[int, list] = {c: [] for c, _, _ in cols}
        for row in rows[6:]:  # row 7+ data
            if not row:
                continue
            d = _parse_period(row[0])
            if d is None:
                continue
            periods.append(d)
            for c, _, _ in cols:
                v = row[c] if c < len(row) else None
                series_vals[c].append(v if isinstance(v, (int, float)) else None)

        out = [(nm, unit, series_vals[c]) for c, nm, unit in cols]
        return periods, out

    # ---- contract -----------------------------------------------------------

    def _meta(self, name: str, unit: Optional[str]) -> SeriesMeta:
        sid = f"{self.source_id}:{_slug(name)}"
        return SeriesMeta(
            sid,
            title=f"{name} ({unit})" if unit else name,
            frequency="M",
            unit=unit,
            geography="World",
            category="commodities",
            license_id=self.license_id,
            metadata={"commodity": name, "sheet": self.SHEET},
        )

    def discover(self) -> list[SeriesMeta]:
        """List the curated starter commodities (offline; no download)."""
        return [self._meta(name, None) for name in STARTER]

    def fetch(self, since: Optional[dt.date] = None):
        periods, columns = self._load_columns()
        for name, unit, values in columns:
            if name.strip().lower() not in _STARTER_SET:
                continue  # restrict to curated starter set
            sid = f"{self.source_id}:{_slug(name)}"
            obs = []
            for d, v in zip(periods, values):
                if v is None:
                    continue
                if since is not None and d < since:
                    continue
                obs.append(Observation(sid, d, float(v), version="clean"))
            if obs:
                yield self._meta(name, unit), obs
